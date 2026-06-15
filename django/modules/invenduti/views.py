"""
Modulo views per l'app Invenduti.

Gestisce la visualizzazione e l'esportazione degli articoli invenduti
letti dalla tabella t_invendutiTot del database GoldReport.
Sono disponibili quattro opzioni di filtro predefinite (Deposito, PDV,
Stato K, Vista Completa) selezionabili dalla pagina principale.
"""

from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Dizionario delle opzioni disponibili: chiave = codice opzione, value = titolo + clausola WHERE
OPZIONI = {
    '1': {
        'titolo': 'Opzione 1 – Invenduti Deposito (E/I/S/P)',
        'where': "St IN ('E','I','S','P') AND Gestione = 'DEP'",
    },
    '2': {
        'titolo': 'Opzione 2 – Invenduti PDV (E/I/S/P)',
        'where': "St IN ('E','I','S','P') AND Gestione = 'PDV'",
    },
    '3': {
        'titolo': 'Opzione 3 – Articoli Stato K',
        'where': "St = 'K'",
    },
    '4': {
        'titolo': 'Opzione 4 – Vista Completa',
        'where': '1=1',   # Nessun filtro: restituisce tutti i record
    },
}

# Mappa dei campi DB alle etichette colonna per la visualizzazione e l'export
COLONNE = [
    ('Dtaaggio',             'Data Agg.'),
    ('Cor',                  'Corridoio'),
    ('Camp',                 'Campo'),
    ('rep',                  'Reparto'),
    ('CCOM',                 'CCOM'),
    ('DESCRCCOM',            'Descrizione CCOM'),
    ('CodArticolo',          'Cod. Articolo'),
    ('Descrizione_Articolo', 'Descrizione'),
    ('St',                   'Stato'),
    ('Ul_Vend',              'Ultima Vendita'),
    ('G_PDV',                'Giac. PDV'),
    ('G_Dep',                'Giac. Dep.'),
    ('Ul_Ric',               'Ultimo Ric.'),
    ('Gestione',             'Gestione'),
]


def _esegui_query(opzione_key):
    """
    Esegue la query SQL per l'opzione specificata.

    Esegue un LEFT JOIN tra t_invendutiTot (dati invenduti) e t_masterData
    (dati anagrafici articolo, per ottenere CCOM e descrizione CCOM).
    La clausola WHERE viene costruita dinamicamente dalla configurazione OPZIONI.
    Restituisce una lista di dizionari (un dict per riga).
    """
    where = OPZIONI[opzione_key]['where']
    sql = f"""
        SELECT
            t.Dtaaggio, t.Cor, t.Camp, t.rep,
            m.CCOM, m.DESCRCCOM,
            t.CodArticolo, t.Descrizione_Articolo,
            t.St, t.Ul_Vend, t.G_PDV, t.G_Dep, t.Ul_Ric, t.Gestione
        FROM t_invendutiTot t
        LEFT JOIN t_masterData m ON t.CodArticolo = m.CODART
        WHERE {where}
        ORDER BY t.Cor, t.Camp, t.rep
    """
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql)
        cols = [c[0] for c in cursor.description]
        rows = [dict(zip(cols, row)) for row in cursor.fetchall()]
    return rows


def main(request):
    """
    Pagina principale dell'app Invenduti.
    Mostra le quattro opzioni disponibili come card selezionabili.
    """
    return render(request, 'invenduti/main.html', {'opzioni': OPZIONI})


def anteprima(request, opzione):
    """
    Visualizza i dati dell'opzione selezionata in una tabella HTML.
    Valida che l'opzione esista nel dizionario OPZIONI prima di procedere.
    Passa al template anche il numero totale di righe.
    """
    if opzione not in OPZIONI:
        return HttpResponse('Opzione non valida', status=400)
    righe = _esegui_query(opzione)
    ctx = {
        'titolo': OPZIONI[opzione]['titolo'],
        'opzione': opzione,
        'colonne': COLONNE,
        'righe': righe,
        'totale': len(righe),
    }
    return render(request, 'invenduti/anteprima.html', ctx)


def export_excel(request, opzione):
    """
    Esporta i dati dell'opzione selezionata in un file Excel (.xlsx).

    Formattazione:
    - Intestazione con sfondo blu scuro (#1F4E79) e testo bianco in grassetto.
    - Larghezze colonne calcolate automaticamente (max 40 caratteri).
    - Risposta HTTP con content-type Excel e nome file descrittivo.
    """
    if opzione not in OPZIONI:
        return HttpResponse('Opzione non valida', status=400)

    righe = _esegui_query(opzione)
    titolo = OPZIONI[opzione]['titolo']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Opzione {opzione}'

    # Stile intestazione
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, (_, label) in enumerate(COLONNE, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Popolamento righe dati
    for row_idx, riga in enumerate(righe, 2):
        for col_idx, (campo, _) in enumerate(COLONNE, 1):
            ws.cell(row=row_idx, column=col_idx, value=riga.get(campo))

    # Calcola larghezza colonne in base al contenuto (massimo 40 per leggibilità)
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="invenduti_opzione{opzione}.xlsx"'
    wb.save(response)
    return response
