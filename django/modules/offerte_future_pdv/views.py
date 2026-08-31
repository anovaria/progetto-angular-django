"""
Offerte Future PDV - Views
Modulo per l'estrazione della griglia articoli in promo futura ad uso del PDV.

Legge i dati dalla vista v_ArtPromoFuture (join con v_AllArticolo) sul database
'goldreport'. Filtra solo gli articoli con EAN principale (EANPRINC = 1) e recupera
la data di consegna prevista dall'ordine fornitore aperto legato allo stesso piano
promo (vista dbo.V_OrdiniGenerale).

Funzionalità:
- index: pagina di ricerca con filtri a tendina (piano promo, reparto, sottoreparto, fornitore)
- anteprima: API JSON che restituisce i primi 100 risultati per l'anteprima nel browser
- export_excel: esportazione di tutti i risultati filtrati in formato .xlsx formattato
- report_pdf: stampa PDF dei risultati filtrati, apribile in una nuova finestra del browser
"""
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import connections
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.graphics.barcode import createBarcodeDrawing
from datetime import datetime


def get_current_user(request):
    """Recupera lo username dalla sessione portale."""
    session_user = request.session.get('user', {})
    return session_user.get('username', 'anonymous').lower()


def require_auth(view_func):
    """Inietta request.username dalla sessione portale."""
    def wrapper(request, *args, **kwargs):
        request.username = get_current_user(request)
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def _get_queryset(codpromo=None, reparto=None, sottoreparto=None, codforn=None):
    """Esegue la query sulla griglia offerte future PDV con i filtri opzionali specificati.

    Filtra solo gli articoli con EAN principale (EANPRINC = 1). La data di consegna
    viene recuperata dall'ordine fornitore aperto (stato diverso da 'Evaso') collegato
    allo stesso articolo/piano promo su dbo.V_OrdiniGenerale.

    Argomenti:
        codpromo     (str): codice piano promo (OPLCEXOPR)
        reparto      (str): codice reparto
        sottoreparto (str): codice sottoreparto
        codforn      (str): codice fornitore

    Restituisce:
        list[dict]: lista di articoli promo ordinati per reparto, sottoreparto, CCOM, articolo.
    """
    sql = """
        SELECT DISTINCT
            p.CCOM,
            p.DESCRCCOM,
            p.ARVCEXR          AS CodArt,
            a.STATO,
            p.Descrizione,
            a.EAN,
            p.GIACENZA_DEPOSITO,
            p.GIACENZA_PDV,
            p.QTA_IN_ORDINE,
            p.REPARTO,
            p.SOTTOREPARTO,
            p.CODFORN,
            p.OPLCEXOPR        AS PIANOB,
            dc.DataConsegna
        FROM v_ArtPromoFuture p
        INNER JOIN v_AllArticolo a ON p.ARVCEXR = a.CODART
        OUTER APPLY (
            -- Prima consegna prevista tra gli ordini ancora aperti (STATO 7 = Evaso)
            -- per lo stesso articolo/piano promo
            SELECT MIN(CONVERT(date, o.DATA_CONSEGNA, 103)) AS DataConsegna
            FROM dbo.V_OrdiniGenerale o
            WHERE o.CODART = TRY_CAST(p.ARVCEXR AS INT)
              AND o.CODPROMO = p.OPLCEXOPR
              AND o.STATO <> 7
        ) dc
        WHERE a.EANPRINC = 1
    """
    params = []

    if codpromo:
        sql += " AND p.OPLCEXOPR = %s"
        params.append(codpromo)
    if reparto:
        sql += " AND p.REPARTO = %s"
        params.append(reparto)
    if sottoreparto:
        sql += " AND p.SOTTOREPARTO = %s"
        params.append(sottoreparto)
    if codforn:
        sql += " AND p.CODFORN = %s"
        params.append(codforn)

    sql += " ORDER BY p.REPARTO, p.SOTTOREPARTO, p.CCOM, p.ARVCEXR"

    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql, params)
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # Dedup: stesso articolo può comparire con fornitori diversi → tieni una riga per (CodArt, PIANOB)
    seen = set()
    deduped = []
    for r in rows:
        key = (r.get('CodArt'), r.get('PIANOB'))
        if key not in seen:
            seen.add(key)
            deduped.append(r)
    return deduped


def _get_filtri_disponibili():
    """Carica le liste distinte per popolare le tendine di filtro (piano promo, reparto,
    sottoreparto, fornitore)."""
    with connections['goldreport'].cursor() as cursor:
        cursor.execute("""
            SELECT DISTINCT OPLCEXOPR, MIN(DTAINI) AS DataIni, MAX(DTAFINE) AS DataFin
            FROM v_ArtPromoFuture
            WHERE OPLCEXOPR IS NOT NULL AND OPLCEXOPR <> ''
            GROUP BY OPLCEXOPR
            ORDER BY OPLCEXOPR ASC
        """)
        piani = [
            {
                'codice': row[0],
                'dtaini': str(row[1])[:10] if row[1] else '',
                'dtafine': str(row[2])[:10] if row[2] else '',
            }
            for row in cursor.fetchall()
        ]

        cursor.execute("""
            SELECT DISTINCT REPARTO, DESCREP FROM v_ArtPromoFuture
            WHERE REPARTO IS NOT NULL AND REPARTO <> ''
            ORDER BY DESCREP
        """)
        reparti = [{'codice': row[0], 'descrizione': row[1]} for row in cursor.fetchall()]

        cursor.execute("""
            SELECT DISTINCT SOTTOREPARTO, DESCSREP FROM v_ArtPromoFuture
            WHERE SOTTOREPARTO IS NOT NULL AND SOTTOREPARTO <> ''
            ORDER BY DESCSREP
        """)
        sottoreparti = [{'codice': row[0], 'descrizione': row[1]} for row in cursor.fetchall()]

        cursor.execute("""
            SELECT DISTINCT CODFORN, DESCFORN FROM v_ArtPromoFuture
            WHERE CODFORN IS NOT NULL AND CODFORN <> ''
            ORDER BY DESCFORN
        """)
        fornitori = [{'codice': row[0], 'descrizione': row[1]} for row in cursor.fetchall()]

    return piani, reparti, sottoreparti, fornitori


@csrf_exempt
@require_auth
def index(request):
    """Vista principale del modulo Offerte Future PDV.

    Carica le liste per le tendine di filtro (piano promo, reparto, sottoreparto,
    fornitore) e renderizza la pagina di ricerca.
    """
    try:
        piani, reparti, sottoreparti, fornitori = _get_filtri_disponibili()
    except Exception:
        # In caso di errore (es. DB non raggiungibile), mostra la pagina con liste vuote
        piani, reparti, sottoreparti, fornitori = [], [], [], []

    context = {
        'username': request.username,
        'piani': piani,
        'reparti': reparti,
        'sottoreparti': sottoreparti,
        'fornitori': fornitori,
    }
    return render(request, 'offerte_future_pdv/index.html', context)


@csrf_exempt
@require_auth
def fornitori_per_piano(request):
    """API (GET): restituisce i soli fornitori presenti nel piano promo indicato.

    Usata dal checkbox 'Abilita Fornitore' per restringere la tendina Fornitore
    ai soli fornitori facenti parte del piano promo selezionato (come nell'Access originale).

    Query params: codpromo
    Risposta JSON: { fornitori: [{codice, descrizione}, ...] }
    """
    codpromo = request.GET.get('codpromo', '').strip()
    if not codpromo:
        return JsonResponse({'fornitori': []})

    try:
        with connections['goldreport'].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT CODFORN, DESCFORN FROM v_ArtPromoFuture
                WHERE OPLCEXOPR = %s AND CODFORN IS NOT NULL AND CODFORN <> ''
                ORDER BY DESCFORN
            """, [codpromo])
            fornitori = [{'codice': row[0], 'descrizione': row[1]} for row in cursor.fetchall()]
        return JsonResponse({'fornitori': fornitori})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_auth
def anteprima(request):
    """API (GET): restituisce i primi 100 risultati della query per l'anteprima nel browser.

    Query params: codpromo, reparto, sottoreparto, codforn
    Risposta JSON: { count: N, rows: [...] }
    """
    codpromo     = request.GET.get('codpromo', '').strip() or None
    reparto      = request.GET.get('reparto', '').strip() or None
    sottoreparto = request.GET.get('sottoreparto', '').strip() or None
    codforn      = request.GET.get('codforn', '').strip() or None

    try:
        rows = _get_queryset(codpromo, reparto, sottoreparto, codforn)
        # Converte i valori date in stringa ISO per la serializzazione JSON
        for r in rows:
            for k, v in r.items():
                if hasattr(v, 'isoformat'):
                    r[k] = v.isoformat() if v else None
        # Restituisce solo i primi 100 per l'anteprima; l'Excel conterrà tutti i record
        return JsonResponse({'count': len(rows), 'rows': rows[:100]})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_auth
def export_excel(request):
    """Esporta tutti i risultati filtrati in un file Excel (.xlsx) formattato.

    Query params: stessi di anteprima.
    """
    codpromo     = request.GET.get('codpromo', '').strip() or None
    reparto      = request.GET.get('reparto', '').strip() or None
    sottoreparto = request.GET.get('sottoreparto', '').strip() or None
    codforn      = request.GET.get('codforn', '').strip() or None

    try:
        rows = _get_queryset(codpromo, reparto, sottoreparto, codforn)
    except Exception as e:
        return HttpResponse(f'Errore query: {e}', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Offerte Future PDV'

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', start_color='1A5276')
    data_font   = Font(name='Arial', size=9)
    alt_fill    = PatternFill('solid', start_color='D6EAF8')
    thin        = Side(style='thin', color='BDC3C7')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal='center', vertical='center')

    headers = [
        ('CCOM',          10),
        ('Desc. CCOM',    30),
        ('Cod. Art.',     12),
        ('Stato',          7),
        ('Descrizione',   40),
        ('Cod. EAN',      16),
        ('Giac. Dep.',    11),
        ('Giac. PDV',     11),
        ('Qta Ordine',    11),
        ('Data Consegna', 14),
    ]

    for col_idx, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    num_cols  = {7, 8, 9}
    date_cols = {10}

    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None

        ean_raw = row.get('EAN') or ''
        ean_val = ean_raw.zfill(13) if ean_raw else ''

        ccom = row.get('CCOM')

        values = [
            int(ccom) if ccom is not None else None,
            row.get('DESCRCCOM'),
            row.get('CodArt'),
            row.get('STATO'),
            row.get('Descrizione'),
            ean_val,
            row.get('GIACENZA_DEPOSITO'),
            row.get('GIACENZA_PDV'),
            row.get('QTA_IN_ORDINE'),
            row.get('DataConsegna'),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border
            if fill:
                cell.fill = fill
            if col_idx in num_cols and val is not None:
                cell.number_format = '#,##0'
            if col_idx in date_cols and val is not None:
                cell.number_format = 'DD/MM/YYYY'

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.append([])
    ws.append([f'Totale: {len(rows)} righe'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'offerte_pdv_{codpromo}.xlsx' if codpromo else 'offerte_future_pdv.xlsx'

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _tronca(testo, lunghezza=40):
    """Tronca una stringa aggiungendo '...' se supera la lunghezza indicata."""
    if not testo:
        return ''
    testo = str(testo)
    return testo[:lunghezza] + '...' if len(testo) > lunghezza else testo


def _genera_barcode_ean13(ean_code, bar_width=1.8, bar_height=48, font_size=9):
    """Genera il barcode EAN13 come disegno vettoriale nativo di reportlab (grande e
    leggibile, con cifre stampate sotto), da inserire come flowable nella tabella PDF.

    Essendo vettoriale (non un'immagine raster) resta sempre nitido e scansionabile a
    qualsiasi dimensione o risoluzione di stampa — a differenza di un PNG generato a bassa
    risoluzione, che produce barre di larghezza incoerente e quindi illeggibili.

    Restituisce None se l'EAN non è valido (meno di 12 cifre), in modo da lasciare
    la cella vuota nel PDF senza interrompere la generazione.
    """
    if not ean_code or len(str(ean_code)) < 12:
        return None
    try:
        ean_12 = str(ean_code)[:12]
        return createBarcodeDrawing(
            'EAN13',
            value=ean_12,
            barWidth=bar_width,
            barHeight=bar_height,
            humanReadable=True,
            fontSize=font_size,
        )
    except Exception as e:
        print(f"Errore generazione barcode per {ean_code}: {e}")
        return None


@csrf_exempt
@require_auth
def report_pdf(request):
    """Genera il PDF di stampa (formato A4 orizzontale) dei risultati filtrati,
    da aprire in una nuova finestra del browser (bottone 'Visualizza per Ordine').

    Query params: stessi di anteprima/export_excel.
    """
    codpromo     = request.GET.get('codpromo', '').strip() or None
    reparto      = request.GET.get('reparto', '').strip() or None
    sottoreparto = request.GET.get('sottoreparto', '').strip() or None
    codforn      = request.GET.get('codforn', '').strip() or None

    try:
        rows = _get_queryset(codpromo, reparto, sottoreparto, codforn)
    except Exception as e:
        return HttpResponse(f'Errore query: {e}', status=500)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('cell', fontName='Helvetica', fontSize=8, leading=9.5)
    elementi = []
    titolo = 'Offerte Future PDV'
    if codpromo:
        titolo += f' - Piano {codpromo}'
    elementi.append(Paragraph(titolo, styles['Title']))
    elementi.append(Spacer(1, 12))

    dati_tabella = [['CCOM', 'Desc. CCOM', 'Cod. Art.', 'Stato', 'Descrizione',
                      'EAN', 'Giac. Dep.', 'Giac. PDV', 'Qta Ordine', 'Data Consegna']]
    for row in rows:
        ccom = row.get('CCOM')
        ean_raw = row.get('EAN') or ''
        ean_val = ean_raw.zfill(13) if ean_raw else ''
        barcode_cell = _genera_barcode_ean13(ean_val) or ''
        dati_tabella.append([
            str(int(ccom)) if ccom is not None else '',
            Paragraph(_tronca(row.get('DESCRCCOM'), 60), cell_style),
            row.get('CodArt') or '',
            row.get('STATO') or '',
            Paragraph(_tronca(row.get('Descrizione'), 90), cell_style),
            barcode_cell,
            str(int(row['GIACENZA_DEPOSITO'])) if row.get('GIACENZA_DEPOSITO') is not None else '',
            str(int(row['GIACENZA_PDV'])) if row.get('GIACENZA_PDV') is not None else '',
            str(int(row['QTA_IN_ORDINE'])) if row.get('QTA_IN_ORDINE') is not None else '',
            row['DataConsegna'].strftime('%d/%m/%Y') if row.get('DataConsegna') else '',
        ])

    tabella = Table(dati_tabella, colWidths=[
        32,   # CCOM
        105,  # Desc. CCOM
        42,   # Cod. Art.
        24,   # Stato
        175,  # Descrizione
        205,  # EAN (barcode, largo per scansione veloce con pistola)
        42,   # Giac. Dep.
        42,   # Giac. PDV
        44,   # Qta Ordine
        54,   # Data Consegna
    ], repeatRows=1)
    tabella.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        # Spazio extra sopra/sotto il barcode (colonna EAN, indice 5): evita che la
        # pistola scannerizzata legga per errore il codice della riga sopra/sotto
        ('TOPPADDING', (5, 1), (5, -1), 26),
        ('BOTTOMPADDING', (5, 1), (5, -1), 26),
        # Riga divisoria più marcata tra un articolo e l'altro, come ulteriore delimitatore visivo
        ('LINEBELOW', (0, 1), (-1, -1), 1.2, colors.black),
    ]))
    elementi.append(tabella)
    elementi.append(Spacer(1, 24))

    data_generazione = datetime.now().strftime('%d/%m/%Y %H:%M')
    elementi.append(Paragraph(f"Generato il {data_generazione} da {request.username} — {len(rows)} righe", styles['Normal']))

    doc.build(elementi)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')
