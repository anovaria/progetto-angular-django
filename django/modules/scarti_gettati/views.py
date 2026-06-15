from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum
from datetime import date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import EanArticolo, ScartoGiornaliero


def _estrai_peso_kg(ean_str):
    """Estrae il peso in kg dalle cifre 9-12 del codice EAN (logica foglio Banco Taglio)."""
    if len(ean_str) >= 12:
        try:
            return round(int(ean_str[8:12]) / 1000, 3)
        except ValueError:
            pass
    return 0.0


def index(request):
    data_oggi = date.today()
    scarti = ScartoGiornaliero.objects.filter(data=data_oggi).order_by('-creato_il')
    totale_kg = scarti.aggregate(t=Sum('peso_kg'))['t'] or 0
    context = {
        'scarti': scarti,
        'data_oggi': data_oggi,
        'totale_kg': totale_kg,
    }
    return render(request, 'scarti_gettati/index.html', context)


def api_lookup_ean(request):
    """GET ?ean=... → cerca l'articolo per i primi 8 caratteri dell'EAN."""
    ean_str = request.GET.get('ean', '').strip()
    if len(ean_str) < 8:
        return JsonResponse({'errore': 'EAN troppo corto (minimo 8 cifre)'}, status=400)

    peso_kg = _estrai_peso_kg(ean_str)
    prefix = ean_str[:8]

    articolo = (
        EanArticolo.objects
        .using('goldreport')
        .filter(ean__startswith=prefix, princ=1)
        .first()
    )
    if articolo is None:
        # fallback: cerca senza vincolo PRINC (es. EAN con prefisso 0 nel DB)
        articolo = (
            EanArticolo.objects
            .using('goldreport')
            .filter(ean__contains=prefix)
            .first()
        )
    if articolo is None:
        return JsonResponse({'errore': f'Articolo non trovato per EAN "{ean_str}"'}, status=404)

    return JsonResponse({
        'codart': articolo.codart,
        'descrart': articolo.descrart,
        'ean_principale': articolo.ean,
        'peso_kg': peso_kg,
    })


def salva_scarto(request):
    """POST — salva uno scarto e torna all'index."""
    if request.method != 'POST':
        return redirect('scarti_gettati_index')

    codice_ean = request.POST.get('codice_ean', '').strip()
    codart = request.POST.get('codart', '').strip()
    descrart = request.POST.get('descrart', '').strip()
    ean_principale = request.POST.get('ean_principale', '0').strip()
    peso_kg = request.POST.get('peso_kg', '0').strip()

    if not codice_ean or not codart:
        return redirect('scarti_gettati_index')

    try:
        ScartoGiornaliero.objects.create(
            data=date.today(),
            codice_ean_letto=codice_ean,
            codart=int(codart),
            descrart=descrart,
            ean_principale=int(ean_principale) if ean_principale else 0,
            peso_kg=float(peso_kg) if peso_kg else 0,
            utente=request.user.username if request.user.is_authenticated else 'anonimo',
        )
    except (ValueError, TypeError):
        pass

    return redirect('scarti_gettati_index')


def elimina_scarto(request, pk):
    """POST — elimina uno scarto e torna all'index."""
    if request.method == 'POST':
        scarto = get_object_or_404(ScartoGiornaliero, pk=pk)
        scarto.delete()
    return redirect('scarti_gettati_index')


def riepilogo(request):
    """Riepilogo totali per articolo, filtrabile per data."""
    data_str = request.GET.get('data', '')
    try:
        data_sel = date.fromisoformat(data_str) if data_str else date.today()
    except ValueError:
        data_sel = date.today()

    righe = (
        ScartoGiornaliero.objects
        .filter(data=data_sel)
        .values('codart', 'descrart', 'ean_principale')
        .annotate(tot_kg=Sum('peso_kg'))
        .order_by('codart')
    )

    context = {
        'righe': righe,
        'data_sel': data_sel,
    }
    return render(request, 'scarti_gettati/riepilogo.html', context)


def export_riepilogo(request):
    """Export Excel del riepilogo per data (struttura foglio BancoTaglioEanGold)."""
    data_str = request.GET.get('data', '')
    try:
        data_sel = date.fromisoformat(data_str) if data_str else date.today()
    except ValueError:
        data_sel = date.today()

    righe = (
        ScartoGiornaliero.objects
        .filter(data=data_sel)
        .values('codart', 'descrart', 'ean_principale')
        .annotate(tot_kg=Sum('peso_kg'))
        .order_by('codart')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ScartiGettati'

    titolo_fill = PatternFill('solid', fgColor='2E5090')
    titolo_font = Font(bold=True, color='FFFFFF')

    ws.append([f'BANCO TAGLIO totale scarti causale 12 — {data_sel.strftime("%d/%m/%Y")}'])
    ws['A1'].font = Font(bold=True, size=12)
    ws.append([])

    intestazioni = ['CodArticolo', 'Descrizione Articolo', 'Ean', 'tot Kg']
    ws.append(intestazioni)
    for col_idx, titolo in enumerate(intestazioni, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = titolo_fill
        cell.font = titolo_font
        cell.alignment = Alignment(horizontal='center')

    for riga in righe:
        ws.append([
            riga['codart'],
            riga['descrart'],
            riga['ean_principale'],
            float(riga['tot_kg'] or 0),
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="scarti_gettati_{data_sel}.xlsx"'
    return response
