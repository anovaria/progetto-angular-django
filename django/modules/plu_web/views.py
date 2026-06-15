"""
Modulo views per l'app PLU Web.

Gestisce la visualizzazione, la ricerca e l'export degli articoli PLU
(Price Look-Up) con codice a barre EAN13.

La query principale estrae gli articoli che hanno un codice PLU (tipo EAN = 5,
EANPRINC = 0) con il relativo EAN commerciale. I filtri disponibili sono:
reparto, banco bilancia, fornitore (CCOM), testo libero e filtro Cervinia.

L'export genera un file Excel con due fogli:
1. 'Dati': tabella dati filtrabili senza immagini.
2. 'Stampa Barcode': immagini barcode EAN13 generate con la libreria python-barcode.
"""
from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
from django.utils import timezone
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
import barcode
from barcode.writer import ImageWriter


def _encode_ean13_font(ean13):
    """
    Codifica un EAN13 per il font Code EAN13 (Grand Zebu).

    Utilizza la codifica A/B/C standard per i barcode EAN13:
    - A-encoding (sinistra): lettere 'A'-'J' (cifre 0-9)
    - B-encoding (sinistra, parità inversa): lettere 'K'-'T' (cifre 0-9)
    - C-encoding (destra): caratteri '0'-'9' (offset 107)
    - Guardie: Start='*', Center='[', End='\\'

    La prima cifra (system digit) determina la sequenza di parità A/B
    per le 6 cifre del gruppo sinistro tramite la tabella PARITY.
    Restituisce None se l'input non è un EAN13 valido (13 cifre).
    """
    if not ean13 or len(str(ean13)) != 13:
        return None
    ean13 = str(ean13)

    # Tabella parità: indica se usare encoding A (0) o B (1) per ogni cifra sinistra
    PARITY = [
        [0,0,0,0,0,0],  # 0: AAAAAA
        [0,0,1,0,1,1],  # 1: AABABB
        [0,0,1,1,0,1],  # 2: AABBAB
        [0,0,1,1,1,0],  # 3: AABBBA
        [0,1,0,0,1,1],  # 4: ABAABB
        [0,1,1,0,0,1],  # 5: ABBAAB
        [0,1,1,1,0,0],  # 6: ABBBAA
        [0,1,0,1,0,1],  # 7: ABABAB
        [0,1,0,1,1,0],  # 8: ABABBA
        [0,1,1,0,1,0],  # 9: ABBABA
    ]

    first = int(ean13[0])
    parity = PARITY[first]

    result = '*'  # start guard
    for i, ch in enumerate(ean13[1:7]):
        d = int(ch)
        if parity[i] == 0:  # A encoding
            result += chr(65 + d)   # 'A'-'J'
        else:               # B encoding
            result += chr(75 + d)   # 'K'-'T'
    result += '['           # center guard
    for ch in ean13[7:]:    # C encoding per le 6 cifre destre
        result += chr(107 + int(ch))
    result += '\\'          # end guard
    return result


def execute_plu_query(filters=None):
    """
    Esegue la query PLU su GoldReport con la stessa logica del modulo plu.

    La query usa due CTE:
    - codvel: articoli con EAN tipo 5 (codice bilancia/PLU) e EANPRINC=0.
    - codean: articoli con EAN principale (PRINC=1, tipo non 5, EANPRINC=1).

    Il JOIN tra le due CTE tramite CODART consente di ottenere sia il PLU
    (SUBSTRING dell'EAN tipo 5, caratteri 3-7) sia l'EAN commerciale.

    LEFT JOIN con t_masterData per CCOM/descrizione e t_Reparti per nome reparto.
    Se il filtro 'cervinia' è attivo, aggiunge INNER JOIN con t_assoCervinia.

    Supporta ordinamento dinamico tramite SORT_MAP.
    Aggiunge ean_formatted (EAN zero-padded a 13 caratteri) e reparto_nome.
    """
    base_query = """
    WITH codvel AS (
        SELECT DISTINCT
            dbo.t_AnagArticoli.REP,
            dbo.t_t_Ean.CODART,
            dbo.t_AnagArticoli.DESCRART,
            dbo.t_t_Ean.EAN
        FROM dbo.t_t_Ean
        INNER JOIN dbo.t_AnagArticoli ON dbo.t_t_Ean.CODART = dbo.t_AnagArticoli.CODART
        WHERE dbo.t_t_Ean.TIPO = 5 AND dbo.t_AnagArticoli.EANPRINC = 0
    ),
    codean AS (
        SELECT DISTINCT
            t_t_Ean_1.CODART,
            t_t_Ean_1.EAN
        FROM dbo.t_t_Ean AS t_t_Ean_1
        INNER JOIN dbo.t_AnagArticoli AS t_AnagArticoli_1 ON t_t_Ean_1.CODART = t_AnagArticoli_1.CODART
        WHERE t_t_Ean_1.TIPO NOT IN (5) AND t_t_Ean_1.PRINC = 1 AND t_AnagArticoli_1.EANPRINC = 1
    )
    SELECT DISTINCT
        m.CCOM as ccom,
        m.DESCRCCOM as descrccom,
        a.REP as rep,
        r.RepDescrizione as reparto_nome,
        a.CODART AS codArticolo,
        a.DESCRART AS descrizione,
        SUBSTRING(a.EAN, 1, 2) AS bancobilancia,
        SUBSTRING(a.EAN, 3, 7) AS plu,
        b.EAN as ean
    FROM codvel AS a
    INNER JOIN codean AS b ON a.CODART = b.CODART
    LEFT OUTER JOIN dbo.t_masterData AS m ON a.CODART = m.CODART
    LEFT OUTER JOIN dbo.t_Reparti AS r ON a.REP = r.NrReparto
    {cervinia_join}
    WHERE 1=1
    """

    # Aggiunge INNER JOIN con t_assoCervinia se il filtro Cervinia è attivo
    cervinia_join = ""
    if filters and filters.get('cervinia'):
        cervinia_join = "INNER JOIN dbo.t_assoCervinia ac ON a.CODART = ac.CODART"
    base_query = base_query.format(cervinia_join=cervinia_join)

    params = []

    if filters:
        if filters.get('reparto'):
            base_query += " AND a.REP = %s"
            params.append(filters['reparto'])

        if filters.get('banco'):
            # Il banco bilancia corrisponde ai primi 2 caratteri dell'EAN tipo 5
            base_query += " AND SUBSTRING(a.EAN, 1, 2) = %s"
            params.append(filters['banco'])

        if filters.get('fornitore'):
            base_query += " AND m.CCOM = %s"
            params.append(filters['fornitore'])

        if filters.get('search'):
            # Ricerca testo libero su descrizione, codice, PLU e descrizione CCOM
            search_term = f"%{filters['search']}%"
            base_query += """ AND (
                a.DESCRART LIKE %s
                OR a.CODART LIKE %s
                OR SUBSTRING(a.EAN, 3, 7) LIKE %s
                OR m.DESCRCCOM LIKE %s
            )"""
            params.extend([search_term, search_term, search_term, search_term])

    # Mappa campi disponibili per l'ordinamento dinamico
    SORT_MAP = {
        'plu':         'SUBSTRING(a.EAN, 3, 7)',
        'codArticolo': 'a.CODART',
        'descrizione': 'a.DESCRART',
        'reparto':     'a.REP',
        'fornitore':   'm.DESCRCCOM',
        'ean':         'b.EAN',
    }
    sort_col = SORT_MAP.get(filters.get('sort') if filters else None, 'a.REP, SUBSTRING(a.EAN, 3, 7)')
    sort_dir = 'DESC' if filters and filters.get('dir') == 'desc' else 'ASC'
    if filters and filters.get('sort'):
        base_query += f" ORDER BY {sort_col} {sort_dir}"
    else:
        base_query += " ORDER BY a.REP, SUBSTRING(a.EAN, 3, 7)"

    with connections['goldreport'].cursor() as cursor:
        cursor.execute(base_query, params)
        columns = [col[0] for col in cursor.description]
        results = []

        for row in cursor.fetchall():
            row_dict = dict(zip(columns, row))
            # Formatta l'EAN con zero-padding a 13 caratteri per la lettura barcode
            row_dict['ean_formatted'] = str(row_dict.get('ean', '')).zfill(13) if row_dict.get('ean') else ''
            # Fallback per il nome reparto se non trovato nella tabella t_Reparti
            if not row_dict.get('reparto_nome'):
                row_dict['reparto_nome'] = f"Reparto {row_dict.get('rep', '')}"
            results.append(row_dict)

        return results


def get_filter_options():
    """
    Recupera le opzioni per i dropdown di filtro (reparti, banchi, fornitori).
    Esegue una singola query leggera su t_AnagArticoli con JOIN opzionali
    per evitare di caricare tutti i dati PLU solo per popolare i filtri.
    Deduplica i valori usando dizionari indicizzati per codice.
    Restituisce un dizionario con tre liste di tuple (codice, descrizione).
    """
    query = """
    SELECT DISTINCT
        a.REP,
        r.RepDescrizione as reparto_nome,
        SUBSTRING(a.EAN, 1, 2) as banco,
        m.CCOM,
        m.DESCRCCOM
    FROM dbo.t_AnagArticoli a
    LEFT OUTER JOIN dbo.t_masterData m ON a.CODART = m.CODART
    LEFT OUTER JOIN dbo.t_Reparti r ON a.REP = r.NrReparto
    WHERE a.EANPRINC = 1
    ORDER BY a.REP
    """
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()

    reparti, banchi, fornitori = {}, {}, {}
    for rep, rep_nome, banco, ccom, descrccom in rows:
        if rep and rep not in reparti:
            reparti[rep] = rep_nome or f"Reparto {rep}"
        if banco and banco not in banchi:
            banchi[banco] = banco
        if ccom and ccom not in fornitori:
            fornitori[ccom] = descrccom

    return {
        'reparti': sorted(reparti.items()),
        'banchi': sorted(banchi.items()),
        # Ordina i fornitori per descrizione (gestendo None)
        'fornitori': sorted(fornitori.items(), key=lambda x: x[1] or ''),
    }


def plu_stampa(request):
    """
    Vista per la stampa dei barcode PLU.
    Raccoglie i filtri GET, esegue la query e renderizza il template
    di stampa (apertura in nuova finestra tramite target="_blank").
    Se non sono presenti filtri, carica comunque tutti gli articoli.
    """
    filters = {
        'reparto': request.GET.get('reparto'),
        'banco': request.GET.get('banco'),
        'fornitore': request.GET.get('fornitore'),
        'search': request.GET.get('search'),
        'cervinia': request.GET.get('cervinia'),
    }
    # Rimuove i filtri con valore None o vuoto
    filters = {k: v for k, v in filters.items() if v}
    filters['sort'] = 'descrizione'
    articoli = execute_plu_query(filters)
    return render(request, 'plu_web/stampa.html', {'articoli': articoli, 'filters': filters})


def plu_list(request):
    """
    Vista lista PLU con filtri e ordinamento colonne.
    Se non viene applicato alcun filtro, articoli=None per mostrare
    il messaggio "Applica un filtro" senza eseguire la query costosa.
    Passa al template le opzioni per i dropdown e la configurazione
    delle colonne ordinabili (con direzione corrente).
    """
    filters = {
        'reparto': request.GET.get('reparto'),
        'banco': request.GET.get('banco'),
        'fornitore': request.GET.get('fornitore'),
        'search': request.GET.get('search'),
        'cervinia': request.GET.get('cervinia'),
        'sort': request.GET.get('sort'),
        'dir': request.GET.get('dir'),
    }
    # Rimuove i filtri vuoti prima di passarli alla query
    filters = {k: v for k, v in filters.items() if v}

    filter_options = get_filter_options()
    # Esegui la query solo se almeno un filtro è stato specificato
    articoli = execute_plu_query(filters) if filters else None

    context = {
        'articoli': articoli,
        'totale': len(articoli) if articoli else 0,
        'filters': filters,
        'filter_options': filter_options,
        # Configurazione colonne ordinabili: (chiave_sort, etichetta)
        # None come chiave indica colonna non ordinabile (es. Barcode)
        'columns_sortable': [
            ('plu',         'PLU'),
            ('codArticolo', 'Cod. Articolo'),
            ('descrizione', 'Descrizione'),
            ('ean',         'EAN'),
            (None,          'Barcode'),
            ('reparto',     'Reparto'),
            ('banco',       'Banco'),
            ('fornitore',   'Fornitore'),
        ],
    }

    return render(request, 'plu_web/list.html', context)



def plu_export(request):
    """
    Export Excel con barcode EAN13 come immagini PNG.

    Genera un workbook con due fogli:
    1. 'Dati': tabella dati con header colorato, filtri automatici e
       larghezze ottimizzate. Nessuna immagine per velocità.
    2. 'Stampa Barcode': per ogni articolo inserisce il PLU, la descrizione,
       l'EAN e il barcode come immagine PNG generata con python-barcode.
       Le righe hanno altezza fissa (ROW_H=45pt) per allineare le immagini.

    Se la generazione del barcode fallisce per un articolo, viene saltato
    silenziosamente (except Exception: pass).
    """
    filters = {
        'reparto': request.GET.get('reparto'),
        'banco': request.GET.get('banco'),
        'fornitore': request.GET.get('fornitore'),
        'search': request.GET.get('search'),
        'cervinia': request.GET.get('cervinia'),
    }
    filters = {k: v for k, v in filters.items() if v}

    results = execute_plu_query(filters) if filters else execute_plu_query()

    wb = openpyxl.Workbook()

    # ── Foglio 1: Dati (filtrabile, senza immagini) ──────────────────────────
    ws_dati = wb.active
    ws_dati.title = "Dati"

    header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ['PLU', 'Cod.Articolo', 'Descrizione', 'EAN', 'Reparto', 'Banco', 'Cod.Forn', 'Fornitore']
    for col_num, header in enumerate(headers, 1):
        cell = ws_dati.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Dati del foglio Dati
    for row_num, item in enumerate(results, 2):
        ws_dati.cell(row=row_num, column=1, value=item.get('plu'))
        ws_dati.cell(row=row_num, column=2, value=item.get('codArticolo'))
        ws_dati.cell(row=row_num, column=3, value=item.get('descrizione'))
        ws_dati.cell(row=row_num, column=4, value=item.get('ean_formatted'))
        ws_dati.cell(row=row_num, column=5, value=item.get('reparto_nome'))
        ws_dati.cell(row=row_num, column=6, value=item.get('bancobilancia'))
        ws_dati.cell(row=row_num, column=7, value=item.get('ccom'))
        ws_dati.cell(row=row_num, column=8, value=item.get('descrccom'))

    # Larghezze automatiche capped a 50 caratteri
    for col in ws_dati.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws_dati.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    # ── Foglio 2: Stampa Barcode (immagini PNG, per la stampa) ───────────────
    ws_bc = wb.create_sheet(title="Stampa Barcode")

    bc_header_font = Font(bold=True, size=9)
    for col_num, header in enumerate(['PLU', 'Descrizione', 'EAN', 'Barcode'], 1):
        cell = ws_bc.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal='center')

    # Larghezze fisse per il foglio stampa
    ws_bc.column_dimensions['A'].width = 8
    ws_bc.column_dimensions['B'].width = 35
    ws_bc.column_dimensions['C'].width = 16
    ws_bc.column_dimensions['D'].width = 28

    # Dimensioni immagine barcode in punti/pixel
    ROW_H = 45
    BARCODE_W = 180
    BARCODE_H = 52

    for row_num, item in enumerate(results, 2):
        ws_bc.cell(row=row_num, column=1, value=item.get('plu')).font = Font(size=9)
        ws_bc.cell(row=row_num, column=2, value=item.get('descrizione')).font = Font(size=9)
        ws_bc.cell(row=row_num, column=3, value=item.get('ean_formatted')).font = Font(size=9)
        ws_bc.row_dimensions[row_num].height = ROW_H

        ean = item.get('ean_formatted')
        if ean and len(str(ean)) == 13:
            try:
                # Genera il barcode EAN13 come immagine PNG in memoria
                buf = io.BytesIO()
                ean_obj = barcode.get('ean13', str(ean), writer=ImageWriter())
                ean_obj.write(buf, options={
                    'write_text': True,
                    'module_width': 0.9,
                    'module_height': 12.0,
                    'quiet_zone': 3.0,
                    'font_size': 8,
                    'text_distance': 3.0,
                })
                buf.seek(0)
                img = XLImage(buf)
                img.width = BARCODE_W
                img.height = BARCODE_H
                img.anchor = f'D{row_num}'
                ws_bc.add_image(img)
            except Exception:
                # Se la generazione del barcode fallisce, la cella rimane vuota
                pass

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'plu_articoli_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response
