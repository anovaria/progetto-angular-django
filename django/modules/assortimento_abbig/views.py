"""
Modulo views per l'app Assortimento Abbig.

Visualizza l'assortimento abbigliamento dalla view v_abbigliamento del
database GoldReport, con filtri a cascata Reparto -> Fornitore -> CCOM.
L'anteprima di stampa inserisce N righe vuote dopo ogni articolo per
replicare il report stampabile che gli utenti preparavano a mano con la
macro Inser_Row.xlsm (riga vuota = spazio per annotazioni sulla stampa).
"""

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Colonne mostrate in anteprima ed esportate (campo DB, etichetta).
# Sono le stesse 11 colonne che gli utenti estraevano a mano in Excel.
COLONNE = [
    ('REP',          'Rep.'),
    ('CODFORN',      'Cod. Forn.'),
    ('CCOM',         'CCOM'),
    ('DESCR_LINEA',  'Linea'),
    ('tcol',         'Collezione'),
    ('CODARTFO',     'Cod. Art. Forn.'),
    ('CODART',       'Cod. Articolo'),
    ('DESCRART',     'Descrizione'),
    ('giacenza_pdv', 'Giac. PDV'),
    ('pv_std',       'PV Std'),
    ('PRLIST',       'Pr. Listino'),
]


def _query_goldreport(sql, params=None):
    """Esegue una query sul database GoldReport e restituisce le righe."""
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql, params or [])
        return cursor.fetchall()


def _lista_reparti():
    """Restituisce i reparti distinti presenti nella view."""
    rows = _query_goldreport(
        "SELECT DISTINCT REP FROM v_abbigliamento WHERE REP IS NOT NULL ORDER BY REP"
    )
    return [r[0] for r in rows]


def _lista_fornitori(rep):
    """Restituisce i codici fornitore distinti per il reparto indicato."""
    rows = _query_goldreport(
        "SELECT DISTINCT CODFORN FROM v_abbigliamento "
        "WHERE REP = %s AND CODFORN IS NOT NULL ORDER BY CODFORN",
        [rep],
    )
    return [r[0] for r in rows]


def _lista_ccom(rep, codforn=None):
    """
    Restituisce i CCOM distinti (codice + descrizione) per reparto
    ed eventualmente fornitore.
    """
    sql = (
        "SELECT DISTINCT CCOM, DESCRCCOM FROM v_abbigliamento "
        "WHERE REP = %s AND CCOM IS NOT NULL"
    )
    params = [rep]
    if codforn:
        sql += " AND CODFORN = %s"
        params.append(codforn)
    sql += " ORDER BY CCOM"
    return [{'cod': r[0], 'descr': r[1] or ''} for r in _query_goldreport(sql, params)]


def _where_filtri(rep, codforn=None, ccom=None, linea=None, tcol=None):
    """Costruisce clausola WHERE e parametri dai filtri (reparto obbligatorio)."""
    sql = " WHERE REP = %s"
    params = [rep]
    if codforn:
        sql += " AND CODFORN = %s"
        params.append(codforn)
    if ccom:
        sql += " AND CCOM = %s"
        params.append(ccom)
    if linea:
        sql += " AND DESCR_LINEA = %s"
        params.append(linea)
    if tcol:
        sql += " AND tcol = %s"
        params.append(tcol)
    return sql, params


def _lista_distinct(campo, rep, codforn=None, ccom=None):
    """Valori distinti di un campo per i filtri reparto/fornitore/CCOM correnti."""
    where, params = _where_filtri(rep, codforn, ccom)
    rows = _query_goldreport(
        f"SELECT DISTINCT {campo} FROM v_abbigliamento{where} "
        f"AND {campo} IS NOT NULL ORDER BY {campo}",
        params,
    )
    return [r[0] for r in rows]


def _esegui_query(rep, codforn=None, ccom=None, linea=None, tcol=None):
    """
    Estrae gli articoli filtrati dalla view v_abbigliamento.
    Il reparto è obbligatorio; gli altri sono filtri opzionali.
    Restituisce una lista di tuple nell'ordine definito da COLONNE,
    ordinata per linea (raggruppamento) e poi CCOM/articolo.
    """
    campi = ', '.join(campo for campo, _ in COLONNE)
    where, params = _where_filtri(rep, codforn, ccom, linea, tcol)
    # La view ha piu righe per articolo (un EAN per riga): teniamo una sola
    # riga per CODART preferendo l'EAN principale (EANPRINC=1). ROW_NUMBER
    # gestisce anche gli articoli senza EAN principale o con righe identiche,
    # evitando di scartarli o di duplicarli nello staging/stampa.
    sql = (
        f"SELECT {campi} FROM ("
        f"SELECT {campi}, ROW_NUMBER() OVER ("
        f"PARTITION BY CODART ORDER BY EANPRINC DESC) AS rn "
        f"FROM v_abbigliamento{where}) t WHERE rn = 1 "
        f"ORDER BY DESCR_LINEA, CCOM, CODART"
    )
    return _query_goldreport(sql, params)


# Indici dei campi DESCR_LINEA e CODART nelle righe (posizione in COLONNE)
IDX_LINEA = 3
IDX_CODART = 6


def _raggruppa_per_linea(righe):
    """Raggruppa righe consecutive per linea: [{'linea': ..., 'righe': [...]}]."""
    gruppi = []
    for r in righe:
        linea = r[IDX_LINEA] or '(senza linea)'
        if not gruppi or gruppi[-1]['linea'] != linea:
            gruppi.append({'linea': linea, 'righe': []})
        gruppi[-1]['righe'].append(r)
    return gruppi


def _formatta_riga(riga):
    """Formatta i valori di una riga per la visualizzazione HTML."""
    out = []
    for (campo, _), val in zip(COLONNE, riga):
        if val is None:
            out.append('')
        elif campo == 'giacenza_pdv':
            # Giacenza: mostra come intero (in DB è float)
            out.append(int(val))
        elif campo in ('pv_std', 'PRLIST'):
            # Prezzi: due decimali con virgola
            out.append(f'{val:.2f}'.replace('.', ','))
        else:
            out.append(val)
    return out


def main(request):
    """
    Pagina principale: filtri a cascata Reparto -> Fornitore -> CCOM.
    I dati vengono mostrati solo dopo il click su "Visualizza" (parametro
    'mostra'), per evitare di caricare l'intero reparto a ogni cambio filtro.
    """
    rep = request.GET.get('rep', '')
    codforn = request.GET.get('codforn', '')
    ccom = request.GET.get('ccom', '')
    linea = request.GET.get('linea', '')
    tcol = request.GET.get('tcol', '')

    ctx = {
        'reparti': _lista_reparti(),
        'fornitori': _lista_fornitori(rep) if rep else [],
        'ccom_list': _lista_ccom(rep, codforn or None) if rep else [],
        'linee': _lista_distinct('DESCR_LINEA', rep, codforn or None, ccom or None) if rep else [],
        'collezioni': _lista_distinct('tcol', rep, codforn or None, ccom or None) if rep else [],
        'sel': {'rep': rep, 'codforn': codforn, 'ccom': ccom, 'linea': linea, 'tcol': tcol},
        'colonne': COLONNE,
        'righe': None,
    }
    if rep and request.GET.get('mostra'):
        righe = [_formatta_riga(r) for r in
                 _esegui_query(rep, codforn or None, ccom or None, linea or None, tcol or None)]
        ctx['righe'] = righe
        ctx['gruppi'] = _raggruppa_per_linea(righe)
        ctx['totale'] = len(righe)
    return render(request, 'assortimento_abbig/main.html', ctx)


def api_fornitori(request):
    """Endpoint JSON: fornitori distinti per il reparto selezionato (cascata filtri)."""
    rep = request.GET.get('rep', '')
    if not rep:
        return JsonResponse({'fornitori': []})
    return JsonResponse({'fornitori': _lista_fornitori(rep)})


def api_ccom(request):
    """Endpoint JSON: CCOM distinti per reparto ed eventuale fornitore (cascata filtri)."""
    rep = request.GET.get('rep', '')
    codforn = request.GET.get('codforn', '')
    if not rep:
        return JsonResponse({'ccom': []})
    return JsonResponse({'ccom': _lista_ccom(rep, codforn or None)})


def api_attributi(request):
    """
    Endpoint JSON: linee (DESCR_LINEA) e collezioni (tcol) distinte
    per i filtri reparto/fornitore/CCOM correnti (cascata filtri).
    """
    rep = request.GET.get('rep', '')
    codforn = request.GET.get('codforn', '') or None
    ccom = request.GET.get('ccom', '') or None
    if not rep:
        return JsonResponse({'linee': [], 'collezioni': []})
    return JsonResponse({
        'linee': _lista_distinct('DESCR_LINEA', rep, codforn, ccom),
        'collezioni': _lista_distinct('tcol', rep, codforn, ccom),
    })


def _costruisci_report(params):
    """
    Logica condivisa da anteprima stampa ed export Excel: estrae le righe
    filtrate, applica l'ordine "sposta in fondo" e costruisce la lista
    piatta di blocchi (intestazioni linea + articoli con flag alternanza).
    Restituisce None se manca il reparto, altrimenti un dict con blocchi e
    metadati (filtri, descrizione CCOM, n. righe note, totale).
    """
    rep = params.get('rep', '')
    if not rep:
        return None
    codforn = params.get('codforn', '')
    ccom = params.get('ccom', '')
    linea = params.get('linea', '')
    tcol = params.get('tcol', '')

    try:
        n_note = max(0, min(int(params.get('righe_note', 1)), 10))
    except (TypeError, ValueError):
        n_note = 1

    righe = [_formatta_riga(r) for r in
             _esegui_query(rep, codforn or None, ccom or None, linea or None, tcol or None)]

    # Articoli spostati in fondo dallo staging: vanno in coda alla lista,
    # nell'ordine in cui sono stati spostati
    fondo = [c for c in params.get('fondo', '').split(',') if c]
    coda = []
    if fondo:
        per_codart = {str(r[IDX_CODART]): r for r in righe}
        coda = [per_codart[c] for c in fondo if c in per_codart]
        in_coda = set(fondo)
        righe = [r for r in righe if str(r[IDX_CODART]) not in in_coda]

    # Struttura piatta: intestazioni di gruppo (linea) e blocchi articolo
    # con flag di alternanza colore progressiva
    blocchi = []
    alt = 0
    for gruppo in _raggruppa_per_linea(righe):
        blocchi.append({'tipo': 'linea', 'label': gruppo['linea']})
        for r in gruppo['righe']:
            blocchi.append({'tipo': 'art', 'valori': r, 'alt': alt % 2 == 1})
            alt += 1
    # Coda: come nello staging, l'intestazione di linea compare solo se
    # l'intera linea è stata spostata in fondo (cioè non ha più articoli
    # nella parte principale)
    linee_principali = {r[IDX_LINEA] or '(senza linea)' for r in righe}
    for gruppo in _raggruppa_per_linea(coda):
        if gruppo['linea'] not in linee_principali:
            blocchi.append({'tipo': 'linea', 'label': gruppo['linea']})
        for r in gruppo['righe']:
            blocchi.append({'tipo': 'art', 'valori': r, 'alt': alt % 2 == 1})
            alt += 1

    # Descrizione CCOM per l'intestazione del report
    descr_ccom = ''
    if ccom:
        rows = _query_goldreport(
            "SELECT TOP 1 DESCRCCOM FROM v_abbigliamento WHERE CCOM = %s", [ccom]
        )
        descr_ccom = rows[0][0] if rows else ''

    return {
        'blocchi': blocchi,
        'rep': rep,
        'codforn': codforn,
        'ccom': ccom,
        'linea': linea,
        'tcol': tcol,
        'descr_ccom': descr_ccom,
        'n_note': n_note,
        'totale': len(righe) + len(coda),
    }


def stampa(request):
    """
    Anteprima di stampa: pagina HTML standalone (senza menu del portale)
    con la tabella articoli raggruppata per linea, N righe vuote per
    annotazioni dopo ogni articolo e colori alternati per blocco.
    Accetta anche POST per ricevere l'elenco degli articoli spostati in
    fondo dallo staging (parametro 'fondo', codici articolo separati da
    virgola). Il pulsante "Stampa" richiama la stampa del browser.
    """
    params = request.POST if request.method == 'POST' else request.GET
    data = _costruisci_report(params)
    if data is None:
        return HttpResponse('Parametro reparto mancante', status=400)

    ctx = {
        'colonne': COLONNE,
        'blocchi': data['blocchi'],
        'totale': data['totale'],
        'rep': data['rep'],
        'codforn': data['codforn'],
        'ccom': data['ccom'],
        'linea': data['linea'],
        'tcol': data['tcol'],
        'descr_ccom': data['descr_ccom'],
        'range_note': range(data['n_note']),
        'n_colonne': len(COLONNE),
    }
    return render(request, 'assortimento_abbig/stampa.html', ctx)


def export_excel(request):
    """
    Export Excel del report, identico all'anteprima di stampa: stessa
    intestazione, raggruppamento per linea, ordine "sposta in fondo",
    N righe vuote per annotazioni, colori alternati per blocco e riga di
    separazione in grassetto tra gli articoli. Accetta POST (come la
    stampa) per ricevere i parametri e l'elenco 'fondo'.
    """
    params = request.POST if request.method == 'POST' else request.GET
    data = _costruisci_report(params)
    if data is None:
        return HttpResponse('Parametro reparto mancante', status=400)

    rep = data['rep']
    n_note = data['n_note']
    ncol = len(COLONNE)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Rep {rep}'

    # Stili allineati ai colori dell'anteprima HTML
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')
    linea_fill = PatternFill('solid', fgColor='E3ECF6')
    linea_font = Font(bold=True, color='1F4E79')
    alt_fill = PatternFill('solid', fgColor='DCE8F5')
    thin = Side(style='thin', color='B0B0B0')
    thick = Side(style='medium', color='444444')
    bordo = Border(left=thin, right=thin, top=thin, bottom=thin)
    bordo_sep = Border(left=thin, right=thin, top=thin, bottom=thick)

    # Intestazione del report (titolo + sottotitolo con i filtri), come l'HTML
    sotto = []
    if data['codforn']:
        sotto.append(f"Fornitore: {data['codforn']}")
    if data['ccom']:
        s = f"CCOM: {data['ccom']}"
        if data['descr_ccom']:
            s += f" — {data['descr_ccom']}"
        sotto.append(s)
    if data['linea']:
        sotto.append(f"Linea: {data['linea']}")
    if data['tcol']:
        sotto.append(f"Collezione: {data['tcol']}")
    sotto.append(f"Articoli: {data['totale']}")

    ws.cell(1, 1, f'Assortimento Abbigliamento — Reparto {rep}').font = Font(bold=True, size=13, color='1F4E79')
    ws.cell(2, 1, '   |   '.join(sotto)).font = Font(size=9, color='495057')

    # Riga intestazione colonne (riga 4, ripetuta a ogni pagina di stampa)
    HDR_ROW = 4
    for c, (_, label) in enumerate(COLONNE, 1):
        cell = ws.cell(HDR_ROW, c, label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = bordo

    # Corpo: intestazioni linea + blocchi articolo con righe note
    r = HDR_ROW + 1
    for b in data['blocchi']:
        if b['tipo'] == 'linea':
            for c in range(1, ncol + 1):
                cell = ws.cell(r, c)
                cell.fill = linea_fill
                cell.border = bordo_sep
            cell = ws.cell(r, 1, b['label'])
            cell.font = linea_font
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            r += 1
        else:
            fill = alt_fill if b['alt'] else None
            for c, val in enumerate(b['valori'], 1):
                cell = ws.cell(r, c, val)
                cell.border = bordo
                if fill:
                    cell.fill = fill
            for n in range(1, n_note + 1):
                for c in range(1, ncol + 1):
                    cell = ws.cell(r + n, c)
                    cell.border = bordo
                    if fill:
                        cell.fill = fill
            # Riga di separazione in grassetto: bordo inferiore spesso
            # sull'ultima riga del blocco (dati o ultima riga nota)
            ultima = r + n_note
            for c in range(1, ncol + 1):
                cell = ws.cell(ultima, c)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)
                if fill:
                    cell.fill = fill
            r = ultima + 1

    # Larghezza colonne in base al contenuto (max 40)
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    # Impostazioni di stampa: orizzontale, scala 85% (come il vecchio layout
    # Excel), intestazione colonne ripetuta, piè di pagina solo n. pagina
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.scale = 85
    ws.print_title_rows = f'{HDR_ROW}:{HDR_ROW}'
    ws.oddFooter.center.text = '&P di &N'
    ws.evenFooter.center.text = '&P di &N'

    # Nome file descrittivo in base ai filtri applicati
    nome = f'assortimento_abbig_rep{rep}'
    if data['codforn']:
        nome += f"_forn{data['codforn']}"
    if data['ccom']:
        nome += f"_ccom{data['ccom']}"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome}.xlsx"'
    wb.save(response)
    return response
