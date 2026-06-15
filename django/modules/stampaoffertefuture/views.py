"""
Stampa Offerte Future - Views
Modulo per la consultazione e l'esportazione delle offerte promozionali future.

Legge i dati dalla vista v_ArtPromoFuture (join con v_AllArticolo) sul database
'goldreport'. Filtra solo gli articoli con EAN principale (EANPRINC = 1).

Funzionalità:
- index: pagina di ricerca con filtri (piano promo, date, codice articolo, EAN)
- anteprima: API JSON che restituisce i primi 100 risultati per l'anteprima nel browser
- export_excel: esportazione di tutti i risultati filtrati in formato .xlsx formattato
"""
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import connections
from io import BytesIO
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_current_user(request):
    """Recupera lo username dalla sessione portale."""
    session_user = request.session.get('user', {})
    return session_user.get('username', 'anonymous').lower()


def require_auth(view_func):
    """Inietta request.username dalla sessione portale."""
    def wrapper(request, *args, **kwargs):
        request.username = get_current_user(request)
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def _get_queryset(codpromo=None, dtaini=None, dtafine=None, codart=None, ean=None):
    """Esegue la query sulle offerte future con i filtri opzionali specificati.

    Filtra solo gli articoli con EAN principale (EANPRINC = 1).
    I parametri date sono in formato YYYY-MM-DD (come prodotti da input[type=date]).
    Per EAN, effettua il padding a 13 cifre e cerca sia il valore originale che quello paddato.

    Argomenti:
        codpromo  (str): codice piano promo (OPLCEXOPR)
        dtaini    (str): data inizio >= (YYYY-MM-DD)
        dtafine   (str): data fine <= (YYYY-MM-DD)
        codart    (str): codice articolo esatto
        ean       (str): codice EAN (cerca con e senza padding a 13 cifre)

    Restituisce:
        list[dict]: lista di articoli promo ordinati per data, reparto, fornitore, articolo.
    """
    sql = """
        SELECT DISTINCT
            p.REPARTO,
            p.DESCREP,
            p.SOTTOREPARTO,
            p.DESCSREP,
            p.CCOM,
            p.DESCRCCOM,
            p.CODFORN,
            p.DESCFORN,
            p.CODPIANO,
            p.OPLCEXOPR        AS PIANOB,
            CONVERT(date, p.DTAINI, 105)   AS DataInizio,
            CONVERT(date, p.DTAFINE, 105)  AS DataFine,
            p.ARVCEXR          AS CodArt,
            p.Descrizione,
            a.STATO,
            p.PRZ_OFF          AS PV_OFF,
            a.PREZZOVEND       AS PV_STD,
            p.GIACENZA_PDV,
            p.GIACENZA_DEPOSITO,
            p.QTA_IN_ORDINE,
            p.ULTIMO_ORDINE,
            a.EAN,
            p.BUYER,
            p.CODARTFO
        FROM v_ArtPromoFuture p
        INNER JOIN v_AllArticolo a ON p.ARVCEXR = a.CODART
        WHERE a.EANPRINC = 1
    """
    params = []

    if codpromo:
        sql += " AND p.OPLCEXOPR = %s"
        params.append(codpromo)
    if dtaini:
        # La data arriva nel formato YYYY-MM-DD dal form HTML
        sql += " AND CONVERT(date, p.DTAINI, 105) >= %s"
        params.append(dtaini)
    if dtafine:
        sql += " AND CONVERT(date, p.DTAFINE, 105) <= %s"
        params.append(dtafine)
    if codart:
        sql += " AND p.ARVCEXR = %s"
        params.append(codart)
    if ean:
        # Cerca sia il valore originale che quello con padding a 13 cifre (standard EAN-13)
        ean_padded = ean.zfill(13)
        sql += " AND (a.EAN = %s OR a.EAN = %s)"
        params.append(ean)
        params.append(ean_padded)

    sql += " ORDER BY DataInizio, p.REPARTO, p.SOTTOREPARTO, p.CODFORN, p.ARVCEXR"

    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql, params)
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # Dedup: stesso articolo può comparire con fornitori diversi → tieni una riga per (CodArt, PIANOB)
    seen = set()
    deduped = []
    for r in rows:
        key = (r.get('CodArt'), r.get('PIANOB'))
        if key not in seen:
            seen.add(key)
            deduped.append(r)
    return deduped


@csrf_exempt
@require_auth
def index(request):
    """Vista principale del modulo Stampa Offerte Future.

    Carica la lista dei piani promo disponibili per la datalist del form
    e renderizza la pagina di ricerca. I piani sono ordinati per data di inizio
    decrescente (i più recenti prima).
    """
    # Carica i codici promo disponibili per la dropdown/datalist
    try:
        with connections['goldreport'].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT OPLCEXOPR, MIN(DTAINI) AS DataIni, MAX(DTAFINE) AS DataFin
                FROM v_ArtPromoFuture
                WHERE OPLCEXOPR IS NOT NULL AND OPLCEXOPR <> ''
                GROUP BY OPLCEXOPR
                ORDER BY MIN(DTAINI) DESC
            """)
            piani = [
                {
                    'codice': row[0],
                    # Tronca la data al solo giorno (senza ora) per la visualizzazione
                    'dtaini': str(row[1])[:10] if row[1] else '',
                    'dtafine': str(row[2])[:10] if row[2] else '',
                }
                for row in cursor.fetchall()
            ]
    except Exception:
        # In caso di errore (es. DB non raggiungibile), mostra la pagina con lista vuota
        piani = []

    context = {
        'oggi': date.today().isoformat(),
        'username': request.username,
        'piani': piani,
    }
    return render(request, 'stampaoffertefuture/index.html', context)


@csrf_exempt
@require_auth
def anteprima(request):
    """API (GET): restituisce i primi 100 risultati della query per l'anteprima nel browser.

    Query params: codpromo, dtaini, dtafine, codart, ean
    Risposta JSON: { count: N, rows: [...] }

    I valori datetime vengono convertiti in stringa ISO 8601 per la serializzazione JSON.
    """
    codpromo = request.GET.get('codpromo', '').strip() or None
    dtaini   = request.GET.get('dtaini', '').strip() or None
    dtafine  = request.GET.get('dtafine', '').strip() or None
    codart   = request.GET.get('codart', '').strip() or None
    ean      = request.GET.get('ean', '').strip() or None

    try:
        rows = _get_queryset(codpromo, dtaini, dtafine, codart, ean)
        # Converte i valori datetime/date in stringa ISO per la serializzazione JSON
        for r in rows:
            for k, v in r.items():
                if hasattr(v, 'isoformat'):
                    r[k] = v.isoformat() if v else None
        # Restituisce solo i primi 100 per l'anteprima; l'Excel conterrà tutti i record
        return JsonResponse({'count': len(rows), 'rows': rows[:100]})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_auth
def export_excel(request):
    """Esporta tutti i risultati filtrati in un file Excel (.xlsx) formattato.

    Query params: stessi di anteprima.
    Il file generato ha:
    - Intestazioni con sfondo blu #1A5276, testo bianco, bordi
    - Righe alternate con sfondo azzurro chiaro
    - Prezzi formattati a 2 decimali, date formattate DD/MM/YYYY
    - Colonna 'Data' (col 8) vuota, da compilare manualmente
    - Filtri automatici attivati sulla riga di intestazione
    - Riga di totale in fondo
    """
    codpromo = request.GET.get('codpromo', '').strip() or None
    dtaini   = request.GET.get('dtaini', '').strip() or None
    dtafine  = request.GET.get('dtafine', '').strip() or None
    codart   = request.GET.get('codart', '').strip() or None
    ean      = request.GET.get('ean', '').strip() or None

    try:
        rows = _get_queryset(codpromo, dtaini, dtafine, codart, ean)
    except Exception as e:
        return HttpResponse(f'Errore query: {e}', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Offerte Future'

    # Stili per intestazione, dati e righe alternate
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', start_color='1A5276')
    data_font   = Font(name='Arial', size=9)
    alt_fill    = PatternFill('solid', start_color='D6EAF8')  # Azzurro chiaro per righe pari
    thin        = Side(style='thin', color='BDC3C7')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal='center', vertical='center')

    # Definizione colonne con larghezza
    headers = [
        ('CodArt',       12),
        ('Stato',         7),
        ('Descrizione',  40),
        ('Marca',        25),
        ('PV STD',       10),
        ('PV OFF',       10),
        ('EAN',          16),
        ('Data',         13),   # Colonna vuota da compilare
        ('Piano',        12),
        ('Data Inizio',  13),
        ('Data Fine',    13),
    ]

    for col_idx, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    # Blocca la prima riga per lo scorrimento verticale
    ws.freeze_panes = 'A2'

    # Indici delle colonne con formato speciale
    # col_idx: 1=CodArt, 2=Stato, 3=Descr, 4=Marca, 5=PV_STD, 6=PV_OFF,
    #          7=EAN, 8=Data(vuota), 9=Piano, 10=DataInizio, 11=DataFine
    price_cols = {5, 6}
    date_cols  = {10, 11}

    for row_idx, row in enumerate(rows, 2):
        # Righe pari con sfondo azzurro alternato
        fill = alt_fill if row_idx % 2 == 0 else None

        # Padding EAN a 13 cifre per conformità standard EAN-13
        ean_raw = row.get('EAN') or ''
        ean_val = ean_raw.zfill(13) if ean_raw else ''

        values = [
            row.get('CodArt'),
            row.get('STATO'),
            row.get('Descrizione'),
            row.get('DESCFORN'),
            row.get('PV_STD'),
            row.get('PV_OFF'),
            ean_val,
            None,  # Colonna Data: lasciata vuota per compilazione manuale
            row.get('PIANOB'),
            row.get('DataInizio'),
            row.get('DataFine'),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border
            if fill:
                cell.fill = fill
            # Applica formato numero per i prezzi
            if col_idx in price_cols and val is not None:
                cell.number_format = '#,##0.00'
            # Applica formato data per le colonne data
            if col_idx in date_cols and val is not None:
                cell.number_format = 'DD/MM/YYYY'

    # Aggiunge filtri automatici sull'intestazione
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    # Riga di totale in fondo
    ws.append([])
    ws.append([f'Totale: {len(rows)} righe'])

    # Usa BytesIO per creare il file in memoria prima di inviarlo
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Il nome del file include il codice promo se disponibile
    filename = f'{codpromo}.xlsx' if codpromo else 'offerte_future.xlsx'

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
