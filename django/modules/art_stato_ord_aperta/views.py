from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

COLONNE = [
    ('DTAAGGIO',         'Data Agg.'),
    ('SETT',             'Sett.'),
    ('REP',              'Rep.'),
    ('SREP',             'SRep.'),
    ('FAM',              'Fam.'),
    ('CCOM',             'CCom'),
    ('DESCRCCOM',        'Fornitore'),
    ('CODART',           'Cod. Art.'),
    ('DESCRARTICOLO',    'Descrizione'),
    ('ST',               'Stato'),
    ('ean',              'EAN'),
    ('GIACENZA_PDV',     'Giac. PDV'),
    ('GIACENZA_DEPOSITO','Giac. Dep.'),
    ('BLOCK901',         'Block 901'),
    ('BLOCK1001',        'Block 1001'),
    ('ULTIMA_VENDITA',   'Ultima Vendita'),
]

LIMIT_DEFAULT = 500


def _build_query(sett=None, ricerca=None, limit=None):
    where, params = [], []

    if sett:
        where.append("SETT = %s")
        params.append(sett)

    if ricerca:
        like = f"%{ricerca}%"
        where.append("(CODART LIKE %s OR DESCRARTICOLO LIKE %s OR DESCRCCOM LIKE %s OR CCOM LIKE %s)")
        params.extend([like, like, like, like])

    where_clause = f"WHERE {' AND '.join(where)}" if where else ""
    top_clause   = f"TOP {limit}" if limit else ""

    sql = f"""
        SELECT {top_clause}
            DTAAGGIO, SETT, REP, SREP, FAM,
            CCOM, DESCRCCOM, CODART, DESCRARTICOLO, ST,
            ean, GIACENZA_PDV, GIACENZA_DEPOSITO,
            BLOCK901, BLOCK1001, ULTIMA_VENDITA
        FROM t_ArtEOrdAperta
        {where_clause}
        ORDER BY SETT, REP, SREP, FAM, CODART
    """
    return sql, params


def _esegui_query(sett=None, ricerca=None, limit=None):
    sql, params = _build_query(sett=sett, ricerca=ricerca, limit=limit)
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql, params)
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]


def _get_settori():
    """Valori distinti di SETT per la select del filtro."""
    with connections['goldreport'].cursor() as cursor:
        cursor.execute("SELECT DISTINCT SETT FROM t_ArtEOrdAperta ORDER BY SETT")
        return [r[0] for r in cursor.fetchall()]


def main(request):
    sett    = request.GET.get('sett', '').strip()
    ricerca = request.GET.get('ricerca', '').strip()
    filtrato = bool(sett or ricerca)

    limit = None if filtrato else LIMIT_DEFAULT
    righe = _esegui_query(sett=sett or None, ricerca=ricerca or None, limit=limit)

    ctx = {
        'colonne':   COLONNE,
        'righe':     righe,
        'totale':    len(righe),
        'troncato':  (not filtrato and len(righe) == LIMIT_DEFAULT),
        'sett':      sett,
        'ricerca':   ricerca,
        'settori':   _get_settori(),
    }
    return render(request, 'art_stato_ord_aperta/main.html', ctx)


def export_excel(request):
    sett    = request.GET.get('sett', '').strip()
    ricerca = request.GET.get('ricerca', '').strip()
    righe   = _esegui_query(sett=sett or None, ricerca=ricerca or None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Art Stato E Ord Aperta'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, (_, label) in enumerate(COLONNE, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, riga in enumerate(righe, 2):
        for col_idx, (campo, _) in enumerate(COLONNE, 1):
            ws.cell(row=row_idx, column=col_idx, value=riga.get(campo))

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="art_stato_ord_aperta.xlsx"'
    wb.save(response)
    return response
