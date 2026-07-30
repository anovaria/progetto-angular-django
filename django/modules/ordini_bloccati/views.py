"""
Modulo views per l'app Ordini Bloccati.

Sostituisce i fogli "Bidone Dash" / "Bidone Ord Dettaglio Art" del file
Excel "lista controlli vari.xlsx". Quei fogli leggevano t_BidoneOrdiniDash
e la vista v_CDOrdiniBloccati, entrambi ormai morti: v_CDOrdiniBloccati fa
un RIGHT OUTER JOIN su t_CD_testataOrdiniBloccati, tabella che non viene
più popolata da nessuna SP, quindi restituisce sempre 0 righe.

I dati veri sono in t_CD_DettaglioOrdiniBloccati (popolata ogni giorno alle
05:50 da Db_GoldReport.dbo.sp_Create_t_CD_DettaglioOrdiniBloccati via
OPENQUERY su Oracle GOLDPROD) e contengono già tutte le info di
fornitore/ordine: non serve la testata. Arricchiamo solo con t_masterData
per descrizione articolo, pezzi per cartone e descrizione CCOM.
"""

from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from modules.bidone.views import carica_annotazioni

COLONNE = [
    ('DTAAGGIO',         'Data Agg.'),
    ('DTAORD',           'Data Ordine'),
    ('FornitoreFattura', 'Cod. Forn. Fattura'),
    ('RagioneSociale',   'Ragione Sociale'),
    ('CCom',             'CCom'),
    ('NrOrdine',         'Nr Ordine'),
    ('Articolo',         'Cod. Articolo'),
    ('DESCRART',         'Descrizione'),
    ('QtaOrd',           'Qta Ord.'),
    ('VL',               'VL'),
    ('PZXCART',          'Pz x Cart.'),
    ('DESCRCCOM',        'Descr. CCom'),
    ('Errore',           'Errore'),
]

APP_NAME = 'ordini_bloccati'


def _esegui_query():
    sql = """
        SELECT
            d.DTAAGGIO, d.DTAORD,
            d.[Fornitore Fattura] AS FornitoreFattura,
            d.[Ragione Sociale]   AS RagioneSociale,
            d.[Fornitore Ordine]  AS CCom,
            d.[Nr Ordine]         AS NrOrdine,
            d.Articolo, m.DESCRART, d.Qta_Ord AS QtaOrd, d.VL,
            m.PZXCART, m.DESCRCCOM, d.Errore
        FROM t_CD_DettaglioOrdiniBloccati d
        LEFT JOIN t_masterData m ON d.Articolo = m.CODART
        ORDER BY d.DTAORD DESC, d.[Nr Ordine]
    """
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql)
        cols = [c[0] for c in cursor.description]
        rows = [dict(zip(cols, row)) for row in cursor.fetchall()]
    return rows


def _build_key(r):
    return f"{r.get('NrOrdine', '')}|{r.get('Articolo', '')}"


def _merge_annotazioni(righe, annotazioni):
    for r in righe:
        key = _build_key(r)
        ann = annotazioni.get(key, {})
        r['record_key']    = key
        r['ann_gestito']   = ann.get('gestito', False)
        r['ann_nota']      = ann.get('nota', '')
        r['ann_utente']    = ann.get('utente', '')
        r['ann_data']      = ann.get('aggiornato_il')


def main(request):
    righe = _esegui_query()
    _merge_annotazioni(righe, carica_annotazioni(APP_NAME))
    return render(request, 'ordini_bloccati/main.html', {
        'colonne':   COLONNE,
        'righe':     righe,
        'totale':    len(righe),
        'app_name':  APP_NAME,
    })


def export_excel(request):
    righe = _esegui_query()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ordini Bloccati'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, (_, label) in enumerate(COLONNE, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, riga in enumerate(righe, 2):
        for col_idx, (campo, _) in enumerate(COLONNE, 1):
            ws.cell(row=row_idx, column=col_idx, value=riga.get(campo))

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ordini_bloccati.xlsx"'
    wb.save(response)
    return response
