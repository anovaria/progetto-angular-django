"""
Scarico Promo Reparto - Business Logic (services.py)
Flusso: selezione CCOM → carica articoli → parametri promo → accodamento a perexport.
Tutte le operazioni su tabelle di staging filtrano per utente (concorrenza multi-utente).
"""
from datetime import datetime
from django.db import connections

from .models import (
    ArtFreFase1, ArtDaVerificare,
    PerExport, Meccanica, TipoSconto,
)


def get_gold_cursor():
    return connections['goldreport'].cursor()


# ============================================================
# DROPDOWN CCOM
# ============================================================

def get_ccom_list():
    sql = """
        SELECT DISTINCT CCOM, RAGIONESOCIALE, DIVISIONE
        FROM t_artfrepromo
        WHERE CCOM IS NOT NULL
          AND CCOM <> ''
        ORDER BY CCOM
    """
    with get_gold_cursor() as cursor:
        cursor.execute(sql)
        rows = cursor.fetchall()
    return [{'ccom': str(r[0] or '').strip(),
             'descrccom': str(r[2] or '').strip()} for r in rows]


def get_piani_promo():
    sql = """
        SELECT PIANO, DESCRIZIONE, INIZIO_SELLOUT, FINE_SELLOUT, INIZIO_SELLIN, FINE_SELLIN
        FROM t_PianoPromo
        ORDER BY PIANO DESC
    """
    with get_gold_cursor() as cursor:
        cursor.execute(sql)
        rows = cursor.fetchall()
    return [{'piano': r[0].strip(), 'descrizione': r[1].strip(),
             'inizio': r[2].strip() if r[2] else '',
             'fine': r[3].strip() if r[3] else '',
             'inizio_sellin': r[4].strip() if r[4] else '',
             'fine_sellin': r[5].strip() if r[5] else ''} for r in rows]


def piano_esiste(piano_codice):
    """Verifica che il codice piano sia effettivamente presente in t_PianoPromo su Gold
    al momento dell'accodamento, per evitare di accodare/esportare promozioni legate
    a piani non (più) reali/autorizzati lato Gold."""
    piano_codice = (piano_codice or '').strip()
    if not piano_codice:
        return False
    with get_gold_cursor() as cursor:
        cursor.execute(
            "SELECT TOP 1 1 FROM t_PianoPromo WHERE PIANO = %s",
            [piano_codice]
        )
        return cursor.fetchone() is not None


# ============================================================
# FASE 1: CARICA ARTICOLI PER CCOM
# ============================================================

def carica_articoli_fase1(ccom, utente):
    """Carica articoli da t_artfrepromo per CCOM. Filtra per utente."""
    ArtFreFase1.objects.filter(utente=utente).delete()

    sql_con_prezzo = """
        SELECT DISTINCT
            REP, DESCREP, CCOM, CODFO, DIVISIONE,
            DESC_CEXR, LINEA_PRODOTTO, TIPO_RIORDINO, ARTFO,
            DESCR_LINEA, PRZ_VEND, CEXR, VL,
            (SELECT TOP 1 CAST(PRZ_OFF AS VARCHAR(50))
             FROM v_ArtPromoFuture
             WHERE ARVCEXR = t_artfrepromo.CEXR
             ORDER BY DTAFINE DESC) AS PRZ_OFF
        FROM t_artfrepromo
        WHERE CCOM = %s
        ORDER BY LINEA_PRODOTTO, TIPO_RIORDINO
    """
    sql_senza_prezzo = """
        SELECT DISTINCT
            REP, DESCREP, CCOM, CODFO, DIVISIONE,
            DESC_CEXR, LINEA_PRODOTTO, TIPO_RIORDINO, ARTFO,
            DESCR_LINEA, PRZ_VEND, CEXR, VL
        FROM t_artfrepromo
        WHERE CCOM = %s
        ORDER BY LINEA_PRODOTTO, TIPO_RIORDINO
    """
    try:
        with get_gold_cursor() as cursor:
            cursor.execute(sql_con_prezzo, [ccom])
            rows = cursor.fetchall()
        has_prz_off = True
    except Exception:
        # v_ArtPromoFuture non disponibile: carica senza prezzo offerta
        with get_gold_cursor() as cursor:
            cursor.execute(sql_senza_prezzo, [ccom])
            rows = cursor.fetchall()
        has_prz_off = False

    objs = []
    for r in rows:
        objs.append(ArtFreFase1(
            SOBCEXT=str(r[0] or '').strip(),
            TSOBDESC=str(r[1] or '').strip(),
            CNUF=str(r[2] or '').strip(),
            CNUM=str(r[3] or '').strip(),
            DESC_CNUF=str(r[4] or '').strip(),
            DESC_CEXR=str(r[5] or '').strip(),
            LINEA_PRODOTTO=str(r[6] or '').strip(),
            TIPO_RIORDINO=str(r[7] or '').strip(),
            ARTFO=str(r[8] or '').strip(),
            Desc_Linea=str(r[9] or '').strip(),
            PrezzoV=str(r[10] or '').strip(),
            CEXR=str(r[11] or '').strip(),
            VL=str(r[12] or '').strip(),
            STATO='',
            PrezzoVOff=str(r[13] or '').strip() if has_prz_off else '',
            scelta=False,
            utente=utente,
        ))

    if objs:
        ArtFreFase1.objects.bulk_create(objs)

    return len(objs)


# ============================================================
# SELEZIONE ARTICOLI
# ============================================================

def seleziona_tutti(utente):
    return ArtFreFase1.objects.filter(utente=utente).update(scelta=True)


def deseleziona_tutti(utente):
    return ArtFreFase1.objects.filter(utente=utente).update(scelta=False)


def seleziona_per_linea(linea_prodotto, utente):
    return ArtFreFase1.objects.filter(
        utente=utente, LINEA_PRODOTTO=linea_prodotto
    ).update(scelta=True)


def seleziona_per_tipo_riordino(tipo, utente):
    return ArtFreFase1.objects.filter(
        utente=utente, TIPO_RIORDINO=tipo
    ).update(scelta=True)


def seleziona_per_fascia(prezzo, utente):
    return ArtFreFase1.objects.filter(
        utente=utente, PrezzoV=prezzo
    ).update(scelta=True)


def toggle_selezione(art_id, utente):
    try:
        art = ArtFreFase1.objects.get(id=art_id, utente=utente)
        art.scelta = not art.scelta
        art.save()
        return art.scelta
    except ArtFreFase1.DoesNotExist:
        return None


# ============================================================
# FILTRI DROPDOWN
# ============================================================

def get_linee_prodotto(utente):
    return list(
        ArtFreFase1.objects.filter(utente=utente)
        .exclude(LINEA_PRODOTTO='')
        .values('LINEA_PRODOTTO', 'Desc_Linea')
        .distinct().order_by('LINEA_PRODOTTO')
    )


def get_tipi_riordino(utente):
    return list(
        ArtFreFase1.objects.filter(utente=utente)
        .exclude(TIPO_RIORDINO='')
        .values_list('TIPO_RIORDINO', flat=True)
        .distinct().order_by('TIPO_RIORDINO')
    )


def get_fasce_prezzo(utente):
    return list(
        ArtFreFase1.objects.filter(utente=utente)
        .exclude(PrezzoV='')
        .values_list('PrezzoV', flat=True)
        .distinct().order_by('PrezzoV')
    )


# ============================================================
# FASE 2: ACCODA A PEREXPORT
# ============================================================

def accoda_a_artdaverificare(params, utente, fallback_prezzo=False):
    ArtDaVerificare.objects.filter(utente=utente).delete()

    selezionati = ArtFreFase1.objects.filter(utente=utente, scelta=True)
    if not selezionati.exists():
        return 0, []

    valore_form = (params.get('valore') or '').strip()
    objs = []
    senza_prezzo = []
    for art in selezionati:
        if fallback_prezzo and not valore_form and not (art.PrezzoVOff or '').strip():
            senza_prezzo.append({'codart': art.CEXR, 'desc': art.DESC_CEXR})
            continue
        objs.append(ArtDaVerificare(
            Promozioni='',
            FornitoreAmministrativo=art.CNUM,
            ContrattoCommerciale=art.CNUF,
            RagioneSociale=art.DESC_CNUF,
            CodiceProdotto=art.CEXR,
            DescrizioneProdotto=art.DESC_CEXR,
            SelezionePromozione=params.get('selezione_promo', ''),
            DataInizio=params.get('data_inizio', ''),
            DataFine=params.get('data_fine', ''),
            DataInizioSellin=params.get('sellin_inizio', ''),
            DataFineSellin=params.get('sellin_fine', ''),
            ScontoExtra=params.get('sconto_extra', ''),
            TipoSconto1=params.get('sconto1', ''),
            TipoSconto=params.get('tipo_sconto', ''),
            Meccanica=params.get('meccanica', ''),
            Meccanicav=params.get('meccanicav', ''),
            Valore=params.get('valore') or (art.PrezzoVOff if fallback_prezzo else ''),
            Valore1=params.get('sconto2', '') or params.get('valore1', ''),
            export='',
            QtaOmaggio=params.get('qta_omaggio', ''),
            VL=art.VL,
            promor='',
            utente=utente,
        ))

    if objs:
        ArtDaVerificare.objects.bulk_create(objs)

    return len(objs), senza_prezzo


def verifica_duplicati(utente, modulo=''):
    codici_nuovi = set(
        ArtDaVerificare.objects.filter(utente=utente)
        .values_list('CodiceProdotto', flat=True)
    )
    qs = PerExport.objects.filter(utenteWind=modulo) if modulo else PerExport.objects.all()
    codici_esistenti = set(qs.values_list('CodiceProdotto', flat=True))
    return list(codici_nuovi & codici_esistenti)


def conferma_accodamento(utente, modulo='', operatore=''):
    verificati = ArtDaVerificare.objects.filter(utente=utente)
    if not verificati.exists():
        return 0

    objs = []
    for v in verificati:
        objs.append(PerExport(
            Promozioni=v.Promozioni,
            FornitoreAmministrativo=v.FornitoreAmministrativo,
            ContrattoCommerciale=v.ContrattoCommerciale,
            RagioneSociale=v.RagioneSociale,
            CodiceProdotto=v.CodiceProdotto,
            DescrizioneProdotto=v.DescrizioneProdotto,
            SelezionePromozione=v.SelezionePromozione,
            DataInizio=v.DataInizio,
            DataFine=v.DataFine,
            DataInizioSellin=v.DataInizioSellin,
            DataFineSellin=v.DataFineSellin,
            ScontoExtra=v.ScontoExtra,
            TipoSconto1=v.TipoSconto1,
            TipoSconto=v.TipoSconto,
            Meccanica=v.Meccanica,
            Meccanicav=v.Meccanicav,
            Valore=v.Valore,
            Valore1=v.Valore1,
            export=v.export,
            QtaOmaggio=v.QtaOmaggio,
            pianoB=operatore or v.utente,
            vl=v.VL,
            DATAEXPORT=datetime.now().strftime('%d/%m/%Y %H:%M'),
            utenteWind=modulo,
        ))

    if objs:
        PerExport.objects.bulk_create(objs)

    count = len(objs)
    ArtDaVerificare.objects.filter(utente=utente).delete()
    return count


def svuota_fase1(utente):
    ArtFreFase1.objects.filter(utente=utente).delete()
    ArtDaVerificare.objects.filter(utente=utente).delete()


# ============================================================
# CONTEGGI
# ============================================================

def conta_fase1(utente):
    qs = ArtFreFase1.objects.filter(utente=utente)
    tot = qs.count()
    selezionati = qs.filter(scelta=True).count()
    return {'totale': tot, 'selezionati': selezionati}


def conta_export(modulo=''):
    qs = PerExport.objects.filter(utenteWind=modulo) if modulo else PerExport.objects.all()
    tot = qs.count()
    condacq = qs.exclude(ScontoExtra='').exclude(ScontoExtra='0').count()
    return {'tutti': tot, 'condacq': condacq}


# ============================================================
# LOOKUP
# ============================================================

def get_meccaniche():
    return list(Meccanica.objects.all().values('id', 'codice', 'descrizione'))


def get_tipi_sconto():
    return list(TipoSconto.objects.all().values('id', 'codice', 'descrizione'))


# ============================================================
# CARICAMENTO MANUALE (barcode)
# ============================================================

def cerca_articolo_barcode(barcode):
    """Cerca articolo per CODART o EAN in t_artfrepromo.
    EAN viene risolto tramite join con v_AllArticolo."""
    barcode = barcode.strip()
    sql_con_prezzo = """
        SELECT DISTINCT TOP 1
            f.REP, f.DESCREP, f.CCOM, f.CODFO, f.DIVISIONE,
            f.DESC_CEXR, f.LINEA_PRODOTTO, f.TIPO_RIORDINO, f.ARTFO,
            f.DESCR_LINEA, f.PRZ_VEND, f.CEXR, f.VL,
            (SELECT TOP 1 CAST(PRZ_OFF AS VARCHAR(50))
             FROM v_ArtPromoFuture
             WHERE ARVCEXR = f.CEXR
             ORDER BY DTAFINE DESC) AS PRZ_OFF
        FROM t_artfrepromo f
        LEFT JOIN v_AllArticolo a ON a.CODART = f.CEXR AND a.EANPRINC = 1
        WHERE f.CEXR = %s OR a.EAN = %s
        ORDER BY f.CEXR
    """
    sql_senza_prezzo = """
        SELECT DISTINCT TOP 1
            f.REP, f.DESCREP, f.CCOM, f.CODFO, f.DIVISIONE,
            f.DESC_CEXR, f.LINEA_PRODOTTO, f.TIPO_RIORDINO, f.ARTFO,
            f.DESCR_LINEA, f.PRZ_VEND, f.CEXR, f.VL
        FROM t_artfrepromo f
        LEFT JOIN v_AllArticolo a ON a.CODART = f.CEXR AND a.EANPRINC = 1
        WHERE f.CEXR = %s OR a.EAN = %s
        ORDER BY f.CEXR
    """
    try:
        with get_gold_cursor() as cursor:
            cursor.execute(sql_con_prezzo, [barcode, barcode])
            row = cursor.fetchone()
        prz_off = str(row[13] or '').strip() if row else ''
    except Exception:
        # v_ArtPromoFuture non disponibile: cerca senza prezzo offerta
        with get_gold_cursor() as cursor:
            cursor.execute(sql_senza_prezzo, [barcode, barcode])
            row = cursor.fetchone()
        prz_off = ''

    if not row:
        return None

    return {
        'SOBCEXT':        str(row[0] or '').strip(),
        'TSOBDESC':       str(row[1] or '').strip(),
        'CNUF':           str(row[2] or '').strip(),
        'CNUM':           str(row[3] or '').strip(),
        'DESC_CNUF':      str(row[4] or '').strip(),
        'DESC_CEXR':      str(row[5] or '').strip(),
        'LINEA_PRODOTTO': str(row[6] or '').strip(),
        'TIPO_RIORDINO':  str(row[7] or '').strip(),
        'ARTFO':          str(row[8] or '').strip(),
        'Desc_Linea':     str(row[9] or '').strip(),
        'PrezzoV':        str(row[10] or '').strip(),
        'CEXR':           str(row[11] or '').strip(),
        'VL':             str(row[12] or '').strip(),
        'PrezzoVOff':     prz_off,
    }


def aggiungi_articoli_manuali(articoli, utente):
    """
    Aggiunge articoli scansionati manualmente in ArtFreFase1 (scelta=True).
    Non cancella gli articoli esistenti — si aggiungono a quelli già caricati da CCOM.
    Evita duplicati per CEXR+utente.
    """
    esistenti = set(
        ArtFreFase1.objects.filter(utente=utente).values_list('CEXR', flat=True)
    )
    objs = []
    for a in articoli:
        if a['CEXR'] in esistenti:
            continue
        objs.append(ArtFreFase1(
            SOBCEXT=a['SOBCEXT'],
            TSOBDESC=a['TSOBDESC'],
            CNUF=a['CNUF'],
            CNUM=a['CNUM'],
            DESC_CNUF=a['DESC_CNUF'],
            DESC_CEXR=a['DESC_CEXR'],
            LINEA_PRODOTTO=a['LINEA_PRODOTTO'],
            TIPO_RIORDINO=a['TIPO_RIORDINO'],
            ARTFO=a['ARTFO'],
            Desc_Linea=a['Desc_Linea'],
            PrezzoV=a['PrezzoV'],
            CEXR=a['CEXR'],
            VL=a['VL'],
            STATO='',
            PrezzoVOff=a.get('PrezzoVOff', ''),
            scelta=True,
            utente=utente,
        ))
        esistenti.add(a['CEXR'])

    if objs:
        ArtFreFase1.objects.bulk_create(objs)


def importa_articoli_excel(file_obj, utente):
    """
    Importa articoli da file Excel in ArtFreFase1 (scelta=True).
    Formato atteso: riga 1 = intestazioni, dati da riga 2.
    Colonne chiave: CODART, DESCRART, CODFORN, CCOM, DESCRCCOM, REP,
                    CODARTFO, pv_std (prezzo vendita), PVOFF ARR (prezzo offerta).
    Svuota la staging area solo se ci sono articoli da importare.
    """
    import openpyxl

    if hasattr(file_obj, 'seek'):
        file_obj.seek(0)

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    print(f"[importa_excel] righe lette: {len(rows)}, utente: {utente}")

    if len(rows) < 2:
        raise ValueError(f"File vuoto o con sola intestazione (righe lette: {len(rows)})")

    headers = [str(h).strip().upper() if h is not None else '' for h in rows[0]]

    def idx(*names):
        for name in names:
            try:
                return headers.index(name.upper())
            except ValueError:
                pass
        return None

    i_codart    = idx('CODART')
    i_descrart  = idx('DESCRART')
    i_codforn   = idx('CODFORN')
    i_ccom      = idx('CCOM')
    i_descrccom = idx('DESCRCCOM')
    i_rep       = idx('REP')
    i_codartfo  = idx('CODARTFO')
    i_pvstd     = idx('PV_STD', 'PRZ_VEND')
    i_pvoff     = idx('PVOFF ARR', 'PVOFF')

    if i_codart is None:
        nomi = ', '.join(h for h in headers if h) or '(nessuna)'
        raise ValueError(f"Colonna CODART non trovata. Intestazioni rilevate: {nomi}")

    def v(row, i):
        if i is None or i >= len(row) or row[i] is None:
            return ''
        val = row[i]
        if hasattr(val, 'year'):
            return ''
        return str(val).strip()

    visti = set()
    objs = []
    for row in rows[1:]:
        cexr = v(row, i_codart)
        if not cexr or cexr in visti:
            continue
        visti.add(cexr)
        objs.append(ArtFreFase1(
            SOBCEXT=v(row, i_rep),
            TSOBDESC='',
            CNUF=v(row, i_ccom),
            CNUM=v(row, i_codforn),
            DESC_CNUF=v(row, i_descrccom),
            DESC_CEXR=v(row, i_descrart),
            LINEA_PRODOTTO='',
            TIPO_RIORDINO='',
            ARTFO=v(row, i_codartfo),
            Desc_Linea='',
            PrezzoV=v(row, i_pvstd),
            CEXR=cexr,
            VL='',
            STATO='',
            PrezzoVOff=v(row, i_pvoff),
            scelta=True,
            utente=utente,
        ))

    if not objs:
        campioni = [str(v(row, i_codart) or '(vuoto)') for row in rows[1:6]]
        raise ValueError(
            f"Colonna CODART trovata ma nessun valore valido. "
            f"Primi valori CODART: {', '.join(campioni)}"
        )

    # Svuota lo staging solo dopo aver verificato che ci sono articoli
    ArtFreFase1.objects.filter(utente=utente).delete()
    ArtDaVerificare.objects.filter(utente=utente).delete()
    # batch_size=100: SQL Server limita a 2100 parametri per query (~20 campi × 100 = 2000)
    ArtFreFase1.objects.bulk_create(objs, batch_size=100)
    return len(objs)


def importa_articoli_excel_svendita(file_obj, utente):
    """
    Importa articoli da file Excel svendita in ArtFreFase1 (scelta=True).
    Formato atteso: Codice | Descrizione | Prezzo svendita [€]
    Non cancella gli articoli esistenti: si aggiungono, i duplicati per CEXR
    vengono ignorati (stesso comportamento del caricamento barcode).
    """
    import openpyxl

    if hasattr(file_obj, 'seek'):
        file_obj.seek(0)

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

    esistenti = set(ArtFreFase1.objects.filter(utente=utente).values_list('CEXR', flat=True))
    visti = set(esistenti)
    objs = []

    def idx(headers, *names):
        for name in names:
            try:
                return headers.index(name.upper())
            except ValueError:
                pass
        for name in names:
            for i, h in enumerate(headers):
                if h.startswith(name.upper()):
                    return i
        return None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h).strip().upper() if h is not None else '' for h in rows[0]]
        i_codice = idx(headers, 'CODICE')
        i_descr  = idx(headers, 'DESCRIZIONE')
        i_prezzo = idx(headers, 'PREZZO SVENDITA', 'PREZZO')

        if i_codice is None:
            continue

        for row in rows[1:]:
            val_codice = row[i_codice] if i_codice < len(row) else None
            if val_codice is None:
                continue
            cexr = str(val_codice).strip()
            if not cexr or cexr in visti:
                continue
            visti.add(cexr)

            descr = ''
            if i_descr is not None and i_descr < len(row) and row[i_descr] is not None:
                descr = str(row[i_descr]).strip()

            prezzo = ''
            if i_prezzo is not None and i_prezzo < len(row) and row[i_prezzo] is not None:
                val = row[i_prezzo]
                if not hasattr(val, 'year'):
                    prezzo = str(val).strip()

            objs.append(ArtFreFase1(
                SOBCEXT='', TSOBDESC='', CNUF='', CNUM='', DESC_CNUF='',
                DESC_CEXR=descr, LINEA_PRODOTTO='', TIPO_RIORDINO='',
                ARTFO='', Desc_Linea='', PrezzoV='',
                CEXR=cexr, VL='', STATO='',
                PrezzoVOff=prezzo,
                scelta=True, utente=utente,
            ))

    wb.close()

    if not objs:
        raise ValueError("Nessun articolo nuovo da importare (file vuoto o tutti già presenti)")

    # Arricchimento dati da Gold: recupera in batch i campi mancanti (CNUM, CNUF,
    # SOBCEXT, VL, ecc.) necessari per l'accodamento promo.
    codici = [o.CEXR for o in objs]
    placeholders = ','.join(['%s'] * len(codici))
    gold_data = {}
    try:
        with get_gold_cursor() as cursor:
            # Prima passata: t_artfrepromo (articoli attivi, ha LINEA_PRODOTTO e TIPO_RIORDINO)
            cursor.execute(f"""
                SELECT CEXR, REP, DESCREP, CCOM, CODFO, DIVISIONE,
                       LINEA_PRODOTTO, TIPO_RIORDINO, ARTFO, DESCR_LINEA, PRZ_VEND, VL
                FROM t_artfrepromo
                WHERE CEXR IN ({placeholders})
            """, codici)
            for row in cursor.fetchall():
                cexr = str(row[0] or '').strip()
                if cexr and cexr not in gold_data:
                    gold_data[cexr] = {
                        'SOBCEXT':        str(row[1] or '').strip(),
                        'TSOBDESC':       str(row[2] or '').strip(),
                        'CNUF':           str(row[3] or '').strip(),
                        'CNUM':           str(row[4] or '').strip(),
                        'DESC_CNUF':      str(row[5] or '').strip(),
                        'LINEA_PRODOTTO': str(row[6] or '').strip(),
                        'TIPO_RIORDINO':  str(row[7] or '').strip(),
                        'ARTFO':          str(row[8] or '').strip(),
                        'Desc_Linea':     str(row[9] or '').strip(),
                        'PrezzoV':        str(row[10] or '').strip(),
                        'VL':             str(row[11] or '').strip(),
                    }

            # Seconda passata: t_masterData per gli articoli fine serie non in t_artfrepromo
            mancanti = [c for c in codici if c not in gold_data]
            if mancanti:
                mp = ','.join(['%s'] * len(mancanti))
                cursor.execute(f"""
                    SELECT CODART, REPARTO, DESCREP, CCOM, CODFORN, DESCRCCOM,
                           CODARTFO, PRZ_VEND, VL
                    FROM t_masterData
                    WHERE CODART IN ({mp})
                """, mancanti)
                for row in cursor.fetchall():
                    cexr = str(row[0] or '').strip()
                    if cexr and cexr not in gold_data:
                        gold_data[cexr] = {
                            'SOBCEXT':        str(row[1] or '').strip(),
                            'TSOBDESC':       str(row[2] or '').strip(),
                            'CNUF':           str(row[3] or '').strip(),
                            'CNUM':           str(row[4] or '').strip(),
                            'DESC_CNUF':      str(row[5] or '').strip(),
                            'LINEA_PRODOTTO': '',
                            'TIPO_RIORDINO':  '',
                            'ARTFO':          str(row[6] or '').strip(),
                            'Desc_Linea':     '',
                            'PrezzoV':        str(row[7] or '').strip(),
                            'VL':             str(row[8] or '').strip(),
                        }
    except Exception:
        pass  # Se Gold non risponde, procede con i soli campi Excel

    incompleti = []
    for o in objs:
        g = gold_data.get(o.CEXR, {})
        o.SOBCEXT        = g.get('SOBCEXT', '')
        o.TSOBDESC       = g.get('TSOBDESC', '')
        o.CNUF           = g.get('CNUF', '')
        o.CNUM           = g.get('CNUM', '')
        o.DESC_CNUF      = g.get('DESC_CNUF', '')
        o.LINEA_PRODOTTO = g.get('LINEA_PRODOTTO', '')
        o.TIPO_RIORDINO  = g.get('TIPO_RIORDINO', '')
        o.ARTFO          = g.get('ARTFO', '')
        o.Desc_Linea     = g.get('Desc_Linea', '')
        o.PrezzoV        = g.get('PrezzoV', '')
        o.VL             = g.get('VL', '')
        if not o.DESC_CEXR:
            o.DESC_CEXR = g.get('DESC_CEXR', '')
        if not o.CNUF or not o.CNUM:
            incompleti.append({'codart': o.CEXR, 'descr': o.DESC_CEXR})

    ArtFreFase1.objects.bulk_create(objs, batch_size=100)
    return len(objs), incompleti


# ============================================================
# DUPLICA PROMO DA STORICO GOLD (t_recuperapromo)
# ============================================================

def get_promos_recupera():
    """Restituisce le promo distinte da t_recuperapromo, ordinate per data decrescente."""
    sql = """
        SELECT DISTINCT
            CAST(CODGOLD AS BIGINT) AS CODGOLD,
            OPLCEXOPR,
            DESCPIANO,
            DTAINI,
            DTAFINE
        FROM t_recuperapromo
        WHERE CODGOLD IS NOT NULL
        ORDER BY DTAINI DESC
    """
    with get_gold_cursor() as cursor:
        cursor.execute(sql)
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]


def get_ccom_by_codgold(codgold):
    """Restituisce i CCOM disponibili per un CODGOLD in t_recuperapromo."""
    sql = """
        SELECT DISTINCT CCOM, DESCRCCOM
        FROM t_recuperapromo
        WHERE CAST(CODGOLD AS BIGINT) = %s
          AND CCOM IS NOT NULL AND CCOM <> ''
        ORDER BY CCOM
    """
    with get_gold_cursor() as cursor:
        cursor.execute(sql, [codgold])
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]


def get_articoli_recupera(codgold, ccom):
    """Restituisce gli articoli per CODGOLD + CCOM da t_recuperapromo."""
    if ccom == '__TUTTI__':
        return get_articoli_recupera_tutti(codgold)
    sql = """
        SELECT
            CODART, DESCRARTICOLO, CODARTFO,
            PRZ_VEND, PRZ_OFF, VL, REP, ST, CODFORN, DESCRCCOM
        FROM t_recuperapromo
        WHERE CAST(CODGOLD AS BIGINT) = %s AND CCOM = %s
        ORDER BY CCOM, CODART
    """
    with get_gold_cursor() as cursor:
        cursor.execute(sql, [codgold, ccom])
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]


def get_articoli_recupera_tutti(codgold):
    """Restituisce tutti gli articoli di una promo (tutti i CCOM) da t_recuperapromo."""
    sql = """
        SELECT
            CODART, DESCRARTICOLO, CODARTFO,
            PRZ_VEND, PRZ_OFF, VL, REP, ST, CODFORN, DESCRCCOM, CCOM
        FROM t_recuperapromo
        WHERE CAST(CODGOLD AS BIGINT) = %s
        ORDER BY CCOM, CODART
    """
    with get_gold_cursor() as cursor:
        cursor.execute(sql, [codgold])
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]


def carica_da_recupera(codgold, ccom, utente):
    """
    Copia gli articoli da t_recuperapromo in ArtFreFase1 per l'utente corrente.
    Salta articoli già presenti (evita duplicati per CEXR).
    Restituisce il numero di articoli inseriti.
    """
    articoli = get_articoli_recupera(codgold, ccom)
    if not articoli:
        return 0

    esistenti = set(ArtFreFase1.objects.filter(utente=utente).values_list('CEXR', flat=True))

    def fmt_prezzo(v):
        if v is None:
            return ''
        try:
            f = float(v)
            return str(int(f)) if f == int(f) else str(f)
        except (ValueError, TypeError):
            return ''

    objs = []
    for a in articoli:
        cexr = (a['CODART'] or '').strip()
        if not cexr or cexr in esistenti:
            continue
        try:
            vl_str = str(int(float(a['VL']))) if a['VL'] else ''
        except (ValueError, TypeError):
            vl_str = str(a['VL']) if a['VL'] else ''

        cnuf = a.get('CCOM') or ccom  # per __TUTTI__ usa il CCOM dell'articolo
        objs.append(ArtFreFase1(
            SOBCEXT=a['REP'] or '',
            TSOBDESC='',
            CNUM=a['CODFORN'] or '',
            CNUF=cnuf,
            DESC_CNUF=a['DESCRCCOM'] or '',
            ARTFO=a['CODARTFO'] or '',
            CEXR=cexr,
            DESC_CEXR=a['DESCRARTICOLO'] or '',
            PrezzoVOff=fmt_prezzo(a['PRZ_OFF']),
            PrezzoV=fmt_prezzo(a['PRZ_VEND']),
            VL=vl_str,
            TIPO_RIORDINO='',
            LINEA_PRODOTTO='',
            Desc_Linea='',
            STATO=a['ST'] or '',
            scelta=False,
            ok='',
            SOBCINT='',
            utente=utente,
        ))
        esistenti.add(cexr)

    if objs:
        ArtFreFase1.objects.bulk_create(objs)
    return len(objs)

    return len(objs)