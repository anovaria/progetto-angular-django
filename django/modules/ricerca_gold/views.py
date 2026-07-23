import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from .models import MasterdataAll
from openpyxl.styles import Font, PatternFill


def _costruisci_righe(codiciclean, tipo):
    """Esegue la ricerca su Gold e ricostruisce le righe NELL'ORDINE
    in cui l'utente ha incollato i codici.

    Ogni elemento della lista restituita è un dizionario:
      {"codice": <codice cercato>, "trovato": True/False, "art": <MasterdataAll o None>}

    Un codice può produrre più righe (stesso codice fornitore usato da
    fornitori diversi): in quel caso le righe compaiono tutte, consecutive,
    nella posizione del codice. Un codice non trovato produce una riga
    "vuota" con trovato=False, che schermo ed export evidenziano in rosso.

    Usata sia da ricerca che da export_excel: un solo posto dove
    correggere, nessun rischio che schermo e file differiscano.
    """
    # Il tipo di ricerca determina il campo del modello su cui filtrare
    if tipo == "gold":
        campo = "codart"
    elif tipo == "ean":
        campo = "ean"
    else:
        campo = "codartfo"

    # Filtro dinamico: {"codart__in": [...]} — **filtro lo spacchetta
    # come argomento nominato della query
    filtro = {campo + "__in": codiciclean}

    # Query sul DB Gold (read-only) — .using() forza il database goldreport
    risultati_db = MasterdataAll.objects.using('goldreport').filter(**filtro)

    # Raggruppa le righe trovate per codice cercato:
    # {"100008": [riga1, riga2], "361975": [riga]}
    per_codice = {}
    for r in risultati_db:
        chiave = getattr(r, campo)
        per_codice.setdefault(chiave, []).append(r)

    # Ricostruisce la lista nell'ordine di incollaggio dell'utente
    righe = []
    for c in codiciclean:
        if c in per_codice:
            for r in per_codice[c]:
                righe.append({"codice": c, "trovato": True, "art": r})
        else:
            righe.append({"codice": c, "trovato": False, "art": None})
    return righe


def ricerca(request):
    """Ricerca articoli su t_masterdataall (DB Gold) per codice Gold, EAN
    o codice articolo fornitore.

    Replica i tre fogli del vecchio Excel Ricerca_Ean_CodGold_CodFor:
    l'utente incolla una lista di codici nella textarea, sceglie il tipo
    di ricerca con i tab e ottiene i dati anagrafici da Gold in tempo reale.
    Le righe escono nello stesso ordine dei codici incollati; i codici
    non presenti in Gold compaiono come riga rossa nella loro posizione.
    """
    # None = prima visita: il template non mostra la tabella
    righe = None
    # Default "gold" — serve anche al template per evidenziare il tab attivo
    tipo = "gold"

    if request.method == "POST":
        # Tipo di ricerca dal campo hidden valorizzato dai tab (gold/ean/fornitore)
        tipo = request.POST.get("tipo_ricerca", "gold")
        # Testo grezzo della textarea: un codice per riga
        codici = request.POST.get("codici", "")
        # Pulizia: spezza per riga, toglie gli spazi, scarta le righe vuote
        codiciclean = [c.strip() for c in codici.split("\n") if c.strip()]

        # Gli EAN in tabella sono a 13 cifre: padding con zeri iniziali
        # (es. "80703853" -> "0000080703853")
        if tipo == "ean":
            codiciclean = [c.zfill(13) for c in codiciclean]

        # Costruzione righe nell'ordine di incollaggio (logica condivisa con l'export)
        righe = _costruisci_righe(codiciclean, tipo)

        # Salva in sessione per l'export Excel (che arriva in GET, senza form).
        # Si salvano solo codici e tipo: l'export ricostruisce le righe
        # con la stessa funzione (gli oggetti modello non sono serializzabili).
        request.session['ricerca_gold_codici'] = codiciclean
        request.session['ricerca_gold_tipo'] = tipo
    else:
        # GET = prima visita o "Pulisci": svuota la sessione della ricerca
        request.session.pop('ricerca_gold_codici', None)
        request.session.pop('ricerca_gold_tipo', None)

    return render(request, "ricerca_gold/ricerca.html", {
        "righe": righe,
        "tipo": tipo,
    })


def export_excel(request):
    """Esporta in Excel i risultati dell'ultima ricerca effettuata.

    Arriva in GET (è un link), quindi codici e tipo vengono riletti dalla
    sessione. Le righe sono ricostruite con la stessa funzione della
    ricerca: stesso ordine di incollaggio, con i codici non trovati
    evidenziati in rosso nella colonna del campo cercato.
    """
    # Codici già puliti (e già paddati se EAN) dalla view ricerca
    codiciclean = request.session.get('ricerca_gold_codici', [])
    tipo = request.session.get('ricerca_gold_tipo', 'gold')

    # Stesse righe della schermata — logica condivisa
    righe = _costruisci_righe(codiciclean, tipo)

    # In quale colonna scrivere il codice non trovato: quella del campo cercato
    # (1=Cod.Art.Fornitore, 2=Cod.Gold, 8=EAN — segue l'ordine delle intestazioni)
    if tipo == "gold":
        col_codice = 2
    elif tipo == "ean":
        col_codice = 8
    else:
        col_codice = 1

    # Costruzione del file Excel in memoria
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ricerca Gold"

    # Stili per le righe "non trovato": testo rosso su sfondo rosato
    font_errore = Font(color='C0392B', bold=True)
    fill_errore = PatternFill('solid', start_color='FDECEA')

    # Riga di intestazione — stesso ordine delle colonne della tabella HTML
    ws.append(['Cod. Art. Fornitore', 'Cod. Art. Gold', 'Cod. Fornitore',
               'Descrizione', 'Fornitore', 'CCOM', 'Descr. CCOM',
               'EAN', 'Stato', 'IVA', 'Prezzo Acq.', 'Prezzo Vend.'])

    # Una riga per ogni elemento, nell'ordine di incollaggio
    for riga in righe:
        if riga["trovato"]:
            r = riga["art"]
            ws.append([r.codartfo, r.codart, r.codforn, r.descrart, r.descforn,
                       r.ccom, r.descrccom, r.ean, r.stato, r.iva,
                       r.pracq, r.przvend])
            # Prezzi a 2 decimali sulla riga appena scritta
            ws.cell(row=ws.max_row, column=11).number_format = '#,##0.00'
            ws.cell(row=ws.max_row, column=12).number_format = '#,##0.00'
        else:
            # Codice non trovato: riga di 12 celle vuote con il codice
            # nella colonna del campo cercato e la dicitura in Descrizione
            valori = [''] * 12
            valori[col_codice - 1] = riga["codice"]   # -1: liste da 0, colonne da 1
            ws.append(valori)
            ws.cell(row=ws.max_row, column=4, value='NON PRESENTE IN GOLD')
            for col in range(1, 13):
                cella = ws.cell(row=ws.max_row, column=col)
                cella.font = font_errore
                cella.fill = fill_errore

    # Risposta HTTP con content_type Excel: il browser avvia il download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # attachment = forza il download con questo nome file
    response['Content-Disposition'] = 'attachment; filename="ricerca_gold.xlsx"'
    wb.save(response)
    return response