"""
Stock Picking - Views
Modulo per la consultazione paginata e l'esportazione dello stock picking articoli.

Legge i dati dalla vista V_StockArticoli sul database 'goldreport'.
La ricerca richiede almeno un filtro attivo (campo cerca=1 + ha_filtri=True)
per evitare il caricamento incontrollato di milioni di righe.

Funzionalità:
- filtri_json: API JSON che restituisce i valori distinti per magazzino, stato e VL
- main: pagina di consultazione paginata (PAGE_SIZE righe per pagina)
- export_excel: esportazione di tutti i record filtrati in formato .xlsx
"""
import json
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Numero massimo di righe per pagina nella vista principale
PAGE_SIZE = 200

# Definizione delle colonne: tuple (campo_db, etichetta_display)
# Usate sia per la tabella HTML che per le intestazioni Excel
COLONNE = [
    ('MAG',              'Mag'),
    ('PICKSTATICO',      'Pick Statico'),
    ('CODART',           'Cod. Art.'),
    ('DESCR',            'Descrizione'),
    ('SOGLIA',           'Soglia'),
    ('QTASAT',           'Qta Sat.'),
    ('DATAINV',          'Data Inv.'),
    ('CODATTRI',         'Cod. Attri.'),
    ('STATO',            'Stato'),
    ('VL',               'VL'),
    ('GIACENZADEPOSITO', 'Giac. Dep.'),
]

# Insieme dei parametri GET che costituiscono un filtro attivo.
# Usato dalla view main per decidere se eseguire la query o mostrare lo stato vuoto.
FILTRI_ATTIVI = {'mag', 'stato', 'codart', 'descr', 'vl', 'giac_min', 'giac_max'}


def _build_where(params, stati=None):
    """Costruisce la clausola WHERE e la lista di parametri per la query principale.

    Gestisce tutti i filtri disponibili: magazzino (match esatto), stato (IN list
    da checkbox multipli), codice articolo e descrizione (LIKE), VL (match esatto),
    giacenza deposito minima e massima (range numerico).

    Argomenti:
        params (dict): dizionario dei parametri GET (senza 'stato' e 'page').
        stati  (list): lista dei valori stato selezionati (da checkbox multipli).

    Restituisce:
        tuple: (where_clause str, args list) pronti per cursor.execute.
    """
    conditions = []
    args = []

    mag    = params.get('mag', '').strip()
    codart = params.get('codart', '').strip()
    descr  = params.get('descr', '').strip()
    vl     = params.get('vl', '').strip()

    if mag:
        conditions.append('MAG = %s')
        args.append(mag)
    if stati:
        # Costruisce IN (%s, %s, ...) con tanti placeholder quanti sono gli stati selezionati
        placeholders = ','.join(['%s'] * len(stati))
        conditions.append(f'STATO IN ({placeholders})')
        args.extend(stati)
    if codart:
        # Ricerca parziale: cerca il codice articolo come sottostringa
        conditions.append('CODART LIKE %s')
        args.append(f'%{codart}%')
    if descr:
        # Ricerca parziale: cerca la descrizione come sottostringa
        conditions.append('DESCR LIKE %s')
        args.append(f'%{descr}%')
    if vl:
        conditions.append('VL = %s')
        args.append(vl)

    giac_min = params.get('giac_min', '').strip()
    giac_max = params.get('giac_max', '').strip()
    if giac_min:
        conditions.append('GIACENZADEPOSITO >= %s')
        args.append(giac_min)
    if giac_max:
        conditions.append('GIACENZADEPOSITO <= %s')
        args.append(giac_max)

    # Se non ci sono condizioni, where è stringa vuota (nessun WHERE nella query)
    where = 'WHERE ' + ' AND '.join(conditions) if conditions else ''
    return where, args


def _esegui_query(params, stati=None, limit=None, offset=None):
    """Esegue la query su V_StockArticoli con filtri opzionali e paginazione.

    Esegue prima una COUNT(*) per il totale, poi la query dati con eventuale
    paginazione tramite OFFSET...FETCH NEXT (sintassi SQL Server 2012+).
    Se limit è None, restituisce tutti i record (usato dall'export Excel).

    Argomenti:
        params (dict): parametri di filtro (senza 'stato' e 'page').
        stati  (list): lista valori stato da filtrare.
        limit  (int|None): numero massimo di righe da restituire (None = tutti).
        offset (int): numero di righe da saltare per la paginazione.

    Restituisce:
        tuple: (rows list[dict], totale int)
    """
    where, args = _build_where(params, stati=stati)

    count_sql = f"SELECT COUNT(*) FROM V_StockArticoli {where}"
    data_sql = f"""
        SELECT MAG, PICKSTATICO, CODART, DESCR, SOGLIA, QTASAT,
               DATAINV, CODATTRI, STATO, VL, GIACENZADEPOSITO
        FROM V_StockArticoli
        {where}
        ORDER BY MAG, PICKSTATICO
    """
    # Aggiunge la clausola di paginazione solo quando richiesta (vista principale)
    if limit is not None:
        data_sql += f" OFFSET {offset} ROWS FETCH NEXT {limit} ROWS ONLY"

    with connections['goldreport'].cursor() as cursor:
        cursor.execute(count_sql, args)
        totale = cursor.fetchone()[0]
        cursor.execute(data_sql, args)
        cols = [c[0] for c in cursor.description]
        rows = [dict(zip(cols, row)) for row in cursor.fetchall()]

    return rows, totale


def filtri_json(request):
    """API (GET): restituisce i valori distinti di magazzino, stato e VL.

    Usata dal JavaScript del template per popolare dinamicamente i filtri
    (select magazzino, checkbox stati, select VL) senza ricaricare la pagina.
    In caso di errore di connessione restituisce liste vuote invece di un errore 500.

    Risposta JSON: { mags: [...], stati: [...], vls: [...] }
    """
    try:
        with connections['goldreport'].cursor() as cursor:
            cursor.execute("SELECT DISTINCT MAG FROM V_StockArticoli WHERE MAG IS NOT NULL ORDER BY MAG")
            mags = [r[0] for r in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT STATO FROM V_StockArticoli WHERE STATO IS NOT NULL ORDER BY STATO")
            stati = [r[0] for r in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT VL FROM V_StockArticoli WHERE VL IS NOT NULL ORDER BY VL")
            vls = [r[0] for r in cursor.fetchall()]
        return JsonResponse({'mags': mags, 'stati': stati, 'vls': vls})
    except Exception:
        # In caso di errore restituisce struttura vuota per non bloccare l'interfaccia
        return JsonResponse({'mags': [], 'stati': [], 'vls': []})


def main(request):
    """Vista principale del modulo Stock Picking.

    La query viene eseguita solo se sono soddisfatte entrambe le condizioni:
    1. Il parametro GET 'cerca' è uguale a '1' (bottone Cerca premuto)
    2. Almeno uno dei campi in FILTRI_ATTIVI ha un valore non vuoto

    Questo evita il caricamento automatico di tutti i record al primo accesso.
    La paginazione usa OFFSET/FETCH NEXT con PAGE_SIZE righe per pagina.
    """
    # Estrae la lista degli stati selezionati (checkbox multipli con lo stesso name)
    stati = [s for s in request.GET.getlist('stato') if s.strip()]
    params = request.GET.dict()
    # Rimuove 'stato' dal dict perché già estratto come lista sopra
    params.pop('stato', None)
    # Estrae e rimuove il numero di pagina (default 1 se assente o vuoto)
    page_num = int(params.pop('page', 1) or 1)
    offset = (page_num - 1) * PAGE_SIZE

    # cerca=True solo se l'utente ha esplicitamente premuto il bottone Cerca
    cerca = request.GET.get('cerca') == '1'
    # ha_filtri=True se almeno uno dei campi filtro (escluso stato) è valorizzato, oppure se ci sono stati selezionati
    ha_filtri = any(params.get(k, '').strip() for k in FILTRI_ATTIVI - {'stato'}) or bool(stati)

    # Costruisce la query string per i link di paginazione ed export (senza il parametro page)
    q = request.GET.copy()
    q.pop('page', None)
    query_string = q.urlencode()

    errore = None
    if cerca and ha_filtri:
        try:
            righe, totale = _esegui_query(params, stati=stati, limit=PAGE_SIZE, offset=offset)
            num_pages = (totale + PAGE_SIZE - 1) // PAGE_SIZE
            # Mostra solo le pagine nell'intervallo ±3 rispetto alla pagina corrente
            page_range = range(max(1, page_num - 3), min(num_pages + 1, page_num + 4))
        except Exception as e:
            righe, totale, num_pages, page_range = [], None, 1, range(1, 2)
            errore = str(e)
    else:
        # Nessuna ricerca eseguita: stato iniziale o filtri non attivi
        righe, totale, num_pages, page_range = [], None, 1, range(1, 2)

    ctx = {
        'colonne': COLONNE,
        'righe': righe,
        'totale': totale,
        'params': params,
        'stati': stati,
        # stati_json è passato al template per ripristinare i checkbox selezionati dopo submit
        'stati_json': json.dumps(stati),
        'query_string': query_string,
        'page_num': page_num,
        'num_pages': num_pages,
        'page_range': page_range,
        'cerca': cerca,
        'ha_filtri': ha_filtri,
        'errore': errore,
    }
    return render(request, 'stock_picking/main.html', ctx)


def export_excel(request):
    """Esporta tutti i record filtrati in un file Excel (.xlsx).

    Usa gli stessi parametri GET della vista main ma senza paginazione
    (limit=None), quindi restituisce tutti i record che corrispondono ai filtri.
    I parametri 'page' e 'cerca' vengono ignorati per l'export.

    Il file generato ha:
    - Intestazioni con sfondo blu #1F4E79, testo bianco grassetto, centrate
    - Larghezza colonne auto-adattata al contenuto (max 40 caratteri)
    """
    stati = [s for s in request.GET.getlist('stato') if s.strip()]
    params = request.GET.dict()
    # Rimuove parametri non pertinenti alla query dati
    params.pop('page', None)
    params.pop('cerca', None)
    params.pop('stato', None)

    # Esegue la query senza limite di righe per ottenere il dataset completo
    righe, _ = _esegui_query(params, stati=stati)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stock Picking'

    # Stile intestazione: sfondo blu scuro, testo bianco grassetto
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')

    # Scrive le intestazioni usando le etichette display da COLONNE
    for col_idx, (_, label) in enumerate(COLONNE, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Scrive i dati: per ogni riga usa il campo_db per prelevare il valore dal dizionario
    for row_idx, riga in enumerate(righe, 2):
        for col_idx, (campo, _) in enumerate(COLONNE, 1):
            ws.cell(row=row_idx, column=col_idx, value=riga.get(campo))

    # Auto-dimensionamento colonne (max 40 caratteri per non allargare eccessivamente)
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stock_picking.xlsx"'
    wb.save(response)
    return response
