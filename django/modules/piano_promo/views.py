import json
from datetime import date, datetime as _datetime

from django.contrib import messages
from django.db import connections
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from .models import PromoSessione, PromoSessioneRiga

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

"""
Modulo views per l'app Piano Promo.

Gestisce la creazione, visualizzazione e gestione delle sessioni di promozione
(Piano Gold). Ogni sessione contiene un elenco di articoli con giacenze e stato
(incluso / escluso / neutro). Gli utenti editor possono creare sessioni e
modificare gli stati; gli utenti reparto possono solo segnare gli articoli
come "Ordinato".
"""


# ── Auth helpers ──────────────────────────────────────────────────────────────

def _username(request):
    """Restituisce il nome utente dalla sessione; 'anonimo' se non autenticato."""
    return request.session.get('user', {}).get('username', 'anonimo')

def _is_editor(request):
    """
    Verifica se l'utente ha permessi di editor (gruppo 'depositi' o 'itd').
    Gli editor possono creare sessioni, modificare stati e cancellare righe.
    """
    groups = request.session.get('user', {}).get('groups', [])
    return any(g.lower() in ('depositi', 'itd') for g in groups)

def _is_logged(request):
    """Restituisce True se esiste un utente in sessione (utente autenticato)."""
    return bool(request.session.get('user'))


# ── DB helpers ────────────────────────────────────────────────────────────────


def _righe_da_db(piano_gold):
    """
    Legge le righe di un singolo Piano Gold dalla vista r_pianopromofuturo
    sul database GoldReport. Usata quando viene specificato un solo piano
    senza altri filtri, per ottimizzare la query.
    """
    sql = """
        SELECT
            [Cod. Piano Gold]      AS cod_piano_gold,
            [Inizio SellOut]       AS inizio_sellout,
            codForn                AS cod_forn,
            descForn               AS desc_forn,
            [Cod.Art.]             AS cod_art,
            [Descrizione Articolo] AS desc_art,
            Giac_PDV               AS giac_pdv,
            Giac_Dep               AS giac_dep,
            [Cod.Art.Expo]         AS cod_art_expo,
            [Descrizione Expo]     AS desc_expo,
            [Prezzo Promo]         AS prezzo_promo,
            [Ordine in Corso]      AS ordine_in_corso,
            [Data di Consegna]     AS data_consegna
        FROM r_pianopromofuturo
        WHERE [Cod. Piano Gold] = %s
        ORDER BY [Cod. Piano Gold], descForn, [Cod.Art.]
    """
    with connections['goldreport'].cursor() as cur:
        cur.execute(sql, [piano_gold])
        cols = [c[0] for c in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


def _safe_date(val):
    """
    Converte un valore in oggetto date in modo sicuro.
    Gestisce: None, datetime, date, stringhe ISO (YYYY-MM-DD),
    stringhe italiane (DD/MM/YYYY) e datetime come stringa (YYYY-MM-DD HH:MM:SS).
    """
    if val is None:
        return None
    # datetime prima di date: datetime è sottoclasse di date
    if isinstance(val, _datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    # ISO: YYYY-MM-DD (eventualmente con orario)
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        pass
    # Formato italiano: DD/MM/YYYY
    try:
        parts = s[:10].split('/')
        if len(parts) == 3:
            return date(int(parts[2]), int(parts[1]), int(parts[0]))
    except Exception:
        pass
    return None


# Campi aggiornabili tramite bulk_update (usato in _upsert_righe).
# cod_piano_gold, cod_art_expo e ordine_in_corso non ci sono: fanno parte
# della chiave (cod_art, cod_piano_gold, cod_art_expo, ordine_in_corso).
_UPSERT_FIELDS = [
    'desc_art', 'cod_forn', 'desc_forn',
    'giac_pdv', 'giac_dep', 'desc_expo',
    'prezzo_promo', 'data_consegna',
]

def _kstr(val):
    """Normalizza un valore a stringa pulita per usarlo come parte della chiave riga."""
    return str(val or '').strip()

def _piano_str(val):
    """cod_piano_gold arriva dalla vista come numero (Decimal): normalizza a stringa,
    nello stesso formato con cui le righe sono già salvate su PromoSessioneRiga."""
    return str(val or '').strip()

def _riga_key_db(r):
    """Chiave identità di una riga letta dal DB:
    (cod_art, cod_piano_gold, cod_art_expo, ordine_in_corso).
    Lo stesso articolo viene mostrato su righe distinte per ogni piano, per ogni
    testa expo e per ogni tipo di ordine (EXPO, COLLI, ...), così si vede tutto."""
    return (
        _kstr(r.get('cod_art')),
        _piano_str(r.get('cod_piano_gold')),
        _kstr(r.get('cod_art_expo')),
        _kstr(r.get('ordine_in_corso')),
    )

def _riga_key_model(r):
    """Stessa chiave identità di _riga_key_db ma per un'istanza PromoSessioneRiga."""
    return (_kstr(r.cod_art), _piano_str(r.cod_piano_gold), _kstr(r.cod_art_expo), _kstr(r.ordine_in_corso))

def _deduplicarighe(righe_db):
    """
    Una riga per terna+expo (cod_art, cod_piano_gold, cod_art_expo, ordine_in_corso):
    lo stesso articolo viene mostrato su righe distinte per ogni piano, per ogni
    testa expo e per ogni tipo di ordine (EXPO, COLLI, ...), così si vede tutto.
    Rimuove solo i doppioni esatti restituiti dalla vista.
    """
    seen = {}
    for r in righe_db:
        if not _kstr(r.get('cod_art')):
            continue
        key = _riga_key_db(r)
        if key not in seen:
            seen[key] = r
    return list(seen.values())

def _upsert_righe(sessione, righe_db):
    """
    Sincronizza le righe di una sessione con i dati appena letti dal DB.
    La chiave è (cod_art, cod_piano_gold, cod_art_expo, ordine_in_corso): lo
    stesso articolo occupa una riga per ogni piano, per ogni testa expo e per
    ogni tipo di ordine (EXPO, COLLI, ...), così si vede tutto.

    Per ogni riga del DB:
    - Se la chiave esiste già nella sessione, aggiorna i campi
      (giacenze, descrizioni, ecc.) senza toccare stato e note.
    - Se non esiste, crea una nuova riga nella sessione.

    Utilizza bulk_update e bulk_create per efficienza.
    Restituisce (n_nuove, n_aggiornate).
    """
    righe_db = _deduplicarighe(righe_db)

    # Rimuove i veri doppioni pre-esistenti nella sessione: per ogni chiave
    # (cod_art, cod_piano_gold, cod_art_expo, ordine_in_corso) tieni solo la riga
    # con id minore. Le righe dello stesso articolo su piani, expo o ordini
    # diversi NON sono doppioni e restano.
    # .order_by() rimuove l'ordinamento di Meta che causa errore GROUP BY su SQL Server.
    from django.db.models import Min
    keep_ids = (
        sessione.righe.values('cod_art', 'cod_piano_gold', 'cod_art_expo', 'ordine_in_corso')
        .annotate(min_id=Min('id'))
        .order_by()
        .values_list('min_id', flat=True)
    )
    PromoSessioneRiga.objects.filter(sessione=sessione).exclude(id__in=list(keep_ids)).delete()

    existing = {_riga_key_model(r): r for r in sessione.righe.all()}
    to_create, to_update = [], []
    for r in righe_db:
        cod_art = _kstr(r.get('cod_art'))
        if not cod_art:
            continue
        key = _riga_key_db(r)
        if key in existing:
            # Aggiorna riga esistente mantenendo stato e nota invariati
            riga = existing[key]
            riga.desc_art        = r.get('desc_art') or ''
            riga.cod_forn        = r.get('cod_forn') or ''
            riga.desc_forn       = r.get('desc_forn') or ''
            riga.giac_pdv        = r.get('giac_pdv')
            riga.giac_dep        = r.get('giac_dep')
            riga.desc_expo       = r.get('desc_expo') or ''
            riga.prezzo_promo    = r.get('prezzo_promo')
            riga.data_consegna   = _safe_date(r.get('data_consegna'))
            to_update.append(riga)
        else:
            # Crea nuova riga con stato predefinito ('escluso' / 'neutro' da modello)
            to_create.append(PromoSessioneRiga(
                sessione=sessione,
                cod_piano_gold=_piano_str(r.get('cod_piano_gold')),
                cod_art=cod_art,
                desc_art=r.get('desc_art') or '',
                cod_forn=r.get('cod_forn') or '',
                desc_forn=r.get('desc_forn') or '',
                giac_pdv=r.get('giac_pdv'),
                giac_dep=r.get('giac_dep'),
                cod_art_expo=r.get('cod_art_expo') or '',
                desc_expo=r.get('desc_expo') or '',
                prezzo_promo=r.get('prezzo_promo'),
                ordine_in_corso=str(r.get('ordine_in_corso') or ''),
                data_consegna=_safe_date(r.get('data_consegna')),
            ))
    if to_update:
        PromoSessioneRiga.objects.bulk_update(to_update, _UPSERT_FIELDS)
    if to_create:
        PromoSessioneRiga.objects.bulk_create(to_create)
    return len(to_create), len(to_update)


def _salva_query(sessione, piani_gold, data_da, data_a, sett):
    """
    Accumula i parametri di ricerca usati per popolare la sessione.
    I piani vengono aggiunti a quelli già salvati (non sostituiti), così
    'aggiorna_lista' può rieseguire la query su tutti i piani mai aggiunti.
    """
    esistenti = [p for p in sessione.query_piani.split(',') if p.strip()]
    tutti = list(dict.fromkeys(esistenti + piani_gold))  # preserva ordine, elimina doppioni
    sessione.query_piani   = ','.join(tutti)
    sessione.query_data_da = _safe_date(data_da) if data_da else sessione.query_data_da
    sessione.query_data_a  = _safe_date(data_a)  if data_a  else sessione.query_data_a
    sessione.query_sett    = sett or sessione.query_sett or ''
    sessione.save(update_fields=['query_piani', 'query_data_da', 'query_data_a', 'query_sett'])


# ── Vista dati live (madre) ───────────────────────────────────────────────────

def _righe_live(piani=None, fornitore=None, data_da=None, data_a=None, search=None, sett=None):
    """
    Interroga la vista r_pianopromofuturo con filtri dinamici.
    Costruisce la query SQL aggiungendo clausole WHERE solo per i parametri
    non vuoti. Supporta filtro per piano gold, fornitore, intervallo date
    SellOut, settore e testo libero (codice, descrizione, fornitore).
    Restituisce una lista di dizionari.
    """
    sql = """
        SELECT
            [Cod. Piano Gold]      AS cod_piano_gold,
            [Inizio SellOut]       AS inizio_sellout,
            [Fine SellOut]         AS fine_sellout,
            Sett                   AS sett,
            codForn                AS cod_forn,
            descForn               AS desc_forn,
            [Cod.Art.]             AS cod_art,
            [Descrizione Articolo] AS desc_art,
            Giac_PDV               AS giac_pdv,
            Giac_Dep               AS giac_dep,
            [Cod.Art.Expo]         AS cod_art_expo,
            [Descrizione Expo]     AS desc_expo,
            [Prezzo Promo]         AS prezzo_promo,
            [Ordine in Corso]      AS ordine_in_corso,
            [Data di Consegna]     AS data_consegna
        FROM r_pianopromofuturo
        WHERE 1=1
    """
    params = []
    if piani:
        # Filtro multi-piano con placeholders dinamici
        placeholders = ','.join(['%s'] * len(piani))
        sql += f" AND [Cod. Piano Gold] IN ({placeholders})"
        params.extend(piani)
    if fornitore:
        sql += " AND codForn = %s"
        params.append(fornitore)
    if data_da:
        # TRY_CONVERT evita errori su date non valide nel DB (SQL Server)
        sql += " AND TRY_CONVERT(DATE, [Inizio SellOut], 103) >= %s"
        params.append(data_da)
    if data_a:
        sql += " AND TRY_CONVERT(DATE, [Inizio SellOut], 103) <= %s"
        params.append(data_a)
    if sett:
        sql += " AND Sett = %s"
        params.append(sett)
    if search:
        # Ricerca testo libero su codice articolo, descrizione e fornitore
        like = f"%{search}%"
        sql += " AND ([Cod.Art.] LIKE %s OR [Descrizione Articolo] LIKE %s OR descForn LIKE %s OR codForn LIKE %s)"
        params.extend([like, like, like, like])
    sql += " ORDER BY [Cod. Piano Gold], descForn, [Cod.Art.]"

    with connections['goldreport'].cursor() as cur:
        cur.execute(sql, params)
        cols = [c[0] for c in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


def dati_live(request):
    """
    Vista che mostra i dati live del Piano Promo direttamente dal DB GoldReport,
    senza passare per le sessioni salvate. Accessibile solo agli editor.
    Permette di filtrare per piano gold, date SellOut, settore e testo libero.
    """
    if not _is_editor(request):
        return redirect('piano_promo:index')

    piani_sel = [p for p in request.GET.getlist('piano') if p.strip()]
    data_da   = request.GET.get('data_da', '').strip() or None
    data_a    = request.GET.get('data_a', '').strip() or None
    search    = request.GET.get('search', '').strip() or None
    sett      = request.GET.get('sett', '').strip() or None

    piani, setti, rows, error = [], [], [], None
    try:
        piani, _, setti = _get_filter_options()
    except Exception as e:
        error = str(e)

    # Esegui la query solo se almeno un filtro è stato specificato
    if piani_sel or data_da or data_a or search or sett:
        try:
            rows = _righe_live(piani=piani_sel or None, data_da=data_da, data_a=data_a, search=search, sett=sett)
        except Exception as e:
            error = str(e)

    return render(request, 'piano_promo/dati_live.html', {
        'rows': rows,
        'totale': len(rows),
        'piani': piani,
        'setti': setti,
        'filters': {
            'piani': piani_sel,
            'data_da': data_da or '',
            'data_a': data_a or '',
            'search': search or '',
            'sett': sett or '',
        },
        'error': error,
    })


def _get_filter_options():
    """
    Recupera le opzioni per i dropdown di filtro (piano gold, fornitori, settori).
    Esegue tre query separate su r_pianopromofuturo.
    Restituisce (piani, fornitori, setti).
    """
    sql_p = """
        SELECT [Cod. Piano Gold], MIN(TRY_CONVERT(DATE, [Inizio SellOut], 103)) AS inizio
        FROM r_pianopromofuturo
        WHERE [Cod. Piano Gold] IS NOT NULL
        GROUP BY [Cod. Piano Gold]
        ORDER BY MIN(TRY_CONVERT(DATE, [Inizio SellOut], 103)) DESC
    """
    sql_f = """
        SELECT DISTINCT codForn, descForn
        FROM r_pianopromofuturo
        WHERE codForn IS NOT NULL
        ORDER BY descForn
    """
    sql_s = """
        SELECT DISTINCT Sett
        FROM r_pianopromofuturo
        WHERE Sett IS NOT NULL
        ORDER BY Sett
    """
    with connections['goldreport'].cursor() as cur:
        cur.execute(sql_p)
        piani = [{'cod': r[0], 'inizio': r[1]} for r in cur.fetchall()]
    with connections['goldreport'].cursor() as cur:
        cur.execute(sql_f)
        fornitori = [{'cod': r[0], 'desc': r[1]} for r in cur.fetchall()]
    with connections['goldreport'].cursor() as cur:
        cur.execute(sql_s)
        setti = [r[0] for r in cur.fetchall()]
    return piani, fornitori, setti


# ── Views ─────────────────────────────────────────────────────────────────────

def index(request):
    """
    Pagina principale dell'app Piano Promo.
    Mostra tutte le sessioni salvate come schede (card).
    Gli editor vedono anche il pulsante per creare nuove sessioni.
    """
    if not _is_logged(request):
        return redirect('/portal/')

    editor = _is_editor(request)
    sessioni = PromoSessione.objects.prefetch_related('righe').all()

    piani, setti = [], []
    error_piani = None
    if editor:
        # Carica le opzioni per il modal "Nuova Sessione" solo se è editor
        try:
            piani, _, setti = _get_filter_options()
        except Exception as e:
            error_piani = str(e)

    return render(request, 'piano_promo/index.html', {
        'sessioni': sessioni,
        'editor': editor,
        'piani': piani,
        'setti': setti,
        'error_piani': error_piani,
    })


def nuova(request):
    """
    Crea una nuova sessione di promo importando gli articoli dal DB.
    Accetta solo POST; accessibile solo agli editor.

    Logica di selezione:
    - Se viene fornito un solo piano gold senza altri filtri, usa _righe_da_db
      (query ottimizzata su singolo piano).
    - In tutti gli altri casi usa _righe_live con i filtri specificati.

    L'etichetta della sessione viene costruita dai piani selezionati o
    dall'intervallo di date se non sono stati specificati piani.
    """
    if not _is_editor(request):
        messages.error(request, 'Permesso negato.')
        return redirect('piano_promo:index')

    if request.method != 'POST':
        return redirect('piano_promo:index')

    piani_gold = [p for p in request.POST.getlist('piano_gold') if p.strip()]
    titolo  = request.POST.get('titolo', '').strip()
    note    = request.POST.get('note', '').strip()
    data_da = request.POST.get('data_da', '').strip() or None
    data_a  = request.POST.get('data_a', '').strip() or None
    sett    = request.POST.get('sett', '').strip() or None

    # Almeno un criterio di filtro è obbligatorio
    if not piani_gold and not data_da and not data_a and not sett:
        messages.error(request, 'Seleziona un Piano Gold o specifica almeno un filtro.')
        return redirect('piano_promo:index')

    try:
        if len(piani_gold) == 1 and not data_da and not data_a and not sett:
            righe_db = _righe_da_db(piani_gold[0])
        else:
            righe_db = _righe_live(piani=piani_gold or None, data_da=data_da, data_a=data_a, sett=sett)
    except Exception as e:
        messages.error(request, f'Errore lettura DB: {e}')
        return redirect('piano_promo:index')

    # Costruisce l'etichetta della sessione
    if piani_gold:
        piano_label = ', '.join(piani_gold)
    else:
        parts = []
        if data_da:
            parts.append(f'dal {data_da}')
        if data_a:
            parts.append(f'al {data_a}')
        piano_label = 'SellOut ' + ' '.join(parts)

    try:
        # Ricava la data di inizio sellout dalla prima riga se disponibile
        inizio = _safe_date(righe_db[0]['inizio_sellout']) if righe_db else None
        sessione = PromoSessione.objects.create(
            piano_gold=piano_label,
            inizio_sellout=inizio,
            titolo=titolo,
            note=note,
            creato_da=_username(request),
            query_piani=','.join(piani_gold),
            query_data_da=_safe_date(data_da) if data_da else None,
            query_data_a=_safe_date(data_a)  if data_a  else None,
            query_sett=sett or '',
        )
        # Deduplica per (cod_art, piano) prima dell'inserimento
        righe_db = _deduplicarighe(righe_db)
        # Inserimento massivo di tutte le righe in un'unica operazione
        PromoSessioneRiga.objects.bulk_create([
            PromoSessioneRiga(
                sessione=sessione,
                cod_piano_gold=r.get('cod_piano_gold') or '',
                cod_art=r.get('cod_art') or '',
                desc_art=r.get('desc_art') or '',
                cod_forn=r.get('cod_forn') or '',
                desc_forn=r.get('desc_forn') or '',
                giac_pdv=r.get('giac_pdv'),
                giac_dep=r.get('giac_dep'),
                cod_art_expo=r.get('cod_art_expo') or '',
                desc_expo=r.get('desc_expo') or '',
                prezzo_promo=r.get('prezzo_promo'),
                ordine_in_corso=str(r.get('ordine_in_corso') or ''),
                data_consegna=_safe_date(r.get('data_consegna')),
            )
            for r in righe_db
        ])
    except Exception as e:
        messages.error(request, f'Errore creazione sessione: {e}')
        return redirect('piano_promo:index')

    messages.success(request, f'Sessione "{piano_label}" creata con {len(righe_db)} articoli (una riga per articolo e piano).')
    return redirect('piano_promo:dettaglio', pk=sessione.pk)


def dettaglio(request, pk):
    """
    Visualizza il dettaglio di una sessione con tutte le righe articolo.
    Accessibile a tutti gli utenti autenticati (editor e reparto).

    Gestisce anche il POST per aggiungere manualmente un articolo non presente
    nella lista (solo per editor), tramite action='aggiungi_riga'.
    """
    if not _is_logged(request):
        return redirect('/portal/')

    editor = _is_editor(request)
    sessione = get_object_or_404(PromoSessione, pk=pk)

    # POST: aggiungi riga manuale (solo editor)
    if request.method == 'POST' and editor:
        action = request.POST.get('action')
        if action == 'aggiungi_riga':
            PromoSessioneRiga.objects.create(
                sessione=sessione,
                cod_art=request.POST.get('cod_art', '').strip(),
                desc_art=request.POST.get('desc_art', '').strip(),
                cod_forn=request.POST.get('cod_forn', '').strip(),
                desc_forn=request.POST.get('desc_forn', '').strip(),
                nota=request.POST.get('nota', '').strip(),
                stato='incluso',
                aggiunta_manuale=True,
            )
        return redirect('piano_promo:dettaglio', pk=pk)

    righe = sessione.righe.all()
    piani, setti = [], []
    if editor:
        try:
            piani, _, setti = _get_filter_options()
        except Exception:
            pass
    return render(request, 'piano_promo/dettaglio.html', {
        'sessione': sessione,
        'righe': righe,
        'editor': editor,
        'n_inclusi': sessione.n_inclusi,
        'n_esclusi': sessione.n_esclusi,
        'n_neutri': sessione.n_neutri,
        'piani': piani,
        'setti': setti,
    })


def aggiorna_riga(request, pk):
    """
    API JSON per aggiornare stato e/o nota di una singola riga.
    Accessibile via AJAX dal template dettaglio.html.

    Permessi differenziati:
    - Editor: può impostare qualsiasi stato (neutro/incluso/escluso) e nota.
    - Reparto: può solo portare da 'escluso'/'neutro' a 'incluso' (segnare come Ordinato).
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'method'}, status=405)
    if not _is_logged(request):
        return JsonResponse({'error': 'forbidden'}, status=403)

    riga = get_object_or_404(PromoSessioneRiga, pk=pk)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'json invalid'}, status=400)

    if _is_editor(request):
        if 'stato' in data and data['stato'] in ('neutro', 'incluso', 'escluso'):
            riga.stato = data['stato']
        if 'nota' in data:
            riga.nota = str(data['nota'])[:500]
    else:
        # reparto: può segnare Ordinato (da neutro/escluso) o tornare Neutro (da incluso)
        nuovo = data.get('stato')
        if nuovo == 'incluso' and riga.stato in ('escluso', 'neutro'):
            riga.stato = 'incluso'
        elif nuovo == 'neutro' and riga.stato == 'incluso':
            riga.stato = 'neutro'
        else:
            return JsonResponse({'error': 'forbidden'}, status=403)

    riga.save()
    return JsonResponse({'ok': True, 'stato': riga.stato, 'nota': riga.nota})


@require_POST
def segna_tutti_ordinato(request, pk):
    """
    Segna come 'incluso' (Ordinato) tutte le righe in stato 'escluso'
    per la sessione indicata. Usato dal pulsante di azione rapida nel reparto.
    """
    if not _is_logged(request):
        return JsonResponse({'error': 'forbidden'}, status=403)
    sessione = get_object_or_404(PromoSessione, pk=pk)
    updated = sessione.righe.filter(stato='escluso').update(stato='incluso')
    return JsonResponse({'ok': True, 'updated': updated})


@require_POST
def aggiungi_articoli(request, pk):
    """
    Aggiunge articoli a una sessione esistente utilizzando i filtri specificati.
    Utilizza _upsert_righe per aggiornare le righe già presenti e creare le nuove.
    Salva anche i parametri della query per usi futuri (aggiorna_lista).
    Accessibile solo agli editor.
    """
    if not _is_editor(request):
        messages.error(request, 'Permesso negato.')
        return redirect('piano_promo:index')

    sessione = get_object_or_404(PromoSessione, pk=pk)
    piani_gold = [p for p in request.POST.getlist('piano_gold') if p.strip()]
    data_da = request.POST.get('data_da', '').strip() or None
    data_a  = request.POST.get('data_a', '').strip() or None
    sett    = request.POST.get('sett', '').strip() or None

    if not piani_gold and not data_da and not data_a and not sett:
        messages.error(request, 'Seleziona almeno un filtro.')
        return redirect('piano_promo:dettaglio', pk=pk)

    try:
        if len(piani_gold) == 1 and not data_da and not data_a and not sett:
            righe_db = _righe_da_db(piani_gold[0])
        else:
            righe_db = _righe_live(piani=piani_gold or None, data_da=data_da, data_a=data_a, sett=sett)
    except Exception as e:
        messages.error(request, f'Errore lettura DB: {e}')
        return redirect('piano_promo:dettaglio', pk=pk)

    n_new, n_upd = _upsert_righe(sessione, righe_db)
    _salva_query(sessione, piani_gold, data_da, data_a, sett)
    messages.success(request, f'Aggiunti {n_new} articoli nuovi, aggiornati {n_upd} esistenti.')
    return redirect('piano_promo:dettaglio', pk=pk)


@require_POST
def aggiorna_lista(request, pk):
    """
    Riesegue la query salvata nella sessione e aggiorna le giacenze
    di tutte le righe presenti. Le righe nuove vengono aggiunte;
    gli stati e le note delle righe esistenti restano invariati.
    Accessibile a tutti gli utenti autenticati: non modifica stati né note,
    quindi anche il reparto può rinfrescare i dati.
    """
    if not _is_logged(request):
        return redirect('/portal/')

    sessione = get_object_or_404(PromoSessione, pk=pk)
    # Ricostruisce i parametri dalla query salvata nella sessione
    piani_gold = [p for p in sessione.query_piani.split(',') if p.strip()]
    data_da    = str(sessione.query_data_da) if sessione.query_data_da else None
    data_a     = str(sessione.query_data_a)  if sessione.query_data_a  else None
    sett       = sessione.query_sett or None

    if not piani_gold and not data_da and not data_a and not sett:
        # Sessioni vecchie senza query salvata: ricava i piani dalle righe esistenti
        piani_gold = list(
            sessione.righe.exclude(cod_piano_gold='')
            .values_list('cod_piano_gold', flat=True)
            .distinct()
        )
        if not piani_gold:
            messages.error(request, 'Nessuna query salvata e nessun piano rilevabile dalle righe.')
            return redirect('piano_promo:dettaglio', pk=pk)

    try:
        if len(piani_gold) == 1 and not data_da and not data_a and not sett:
            righe_db = _righe_da_db(piani_gold[0])
        else:
            righe_db = _righe_live(piani=piani_gold or None, data_da=data_da, data_a=data_a, sett=sett)
    except Exception as e:
        messages.error(request, f'Errore lettura DB: {e}')
        return redirect('piano_promo:dettaglio', pk=pk)

    n_new, n_upd = _upsert_righe(sessione, righe_db)
    messages.success(request, f'Lista aggiornata: {n_new} nuovi, {n_upd} aggiornati.')
    return redirect('piano_promo:dettaglio', pk=pk)


@require_POST
def elimina_riga(request, pk):
    """
    Elimina una singola riga da una sessione. Accessibile solo agli editor.
    Dopo l'eliminazione reindirizza al dettaglio della sessione padre.
    """
    if not _is_editor(request):
        return redirect('piano_promo:index')
    riga = get_object_or_404(PromoSessioneRiga, pk=pk)
    sessione_pk = riga.sessione_id
    riga.delete()
    return redirect('piano_promo:dettaglio', pk=sessione_pk)


@require_POST
def elimina(request, pk):
    """
    Elimina una sessione completa (con tutte le righe associate tramite cascade).
    Accessibile solo agli editor.
    """
    if not _is_editor(request):
        return redirect('piano_promo:index')
    sessione = get_object_or_404(PromoSessione, pk=pk)
    sessione.delete()
    return redirect('piano_promo:index')


def export_sessione(request, pk):
    """
    Esporta la sessione in un file Excel (.xlsx) con formattazione.

    Colori per stato:
    - Verde (D5F5E3) = incluso (Ordinato)
    - Rosso (FDEDEC) = escluso (Da Ordinare)
    - Bianco = neutro

    Include intestazioni fisse, colori header blu scuro, bordi sottili,
    filtri automatici e colonne con larghezze ottimizzate.
    Accessibile a tutti gli utenti autenticati.
    """
    if not _is_logged(request):
        return redirect('/portal/')

    sessione = get_object_or_404(PromoSessione, pk=pk)

    righe = sessione.righe.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    # Il titolo del foglio è limitato a 31 caratteri (limite Excel)
    ws.title = sessione.piano_gold[:31]

    # Stili per intestazione e dati
    hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', start_color='1A5276')
    data_font = Font(name='Arial', size=9)
    thin      = Side(style='thin', color='BDC3C7')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_inc  = PatternFill('solid', start_color='A9DFBF')   # Verde = incluso
    fill_exc  = PatternFill('solid', start_color='F1948A')   # Rosso = escluso
    fill_neu  = None                                          # Nessun fill = neutro
    fill_nota = PatternFill('solid', start_color='FEF9C3')   # Giallo = nota su neutro
    center    = Alignment(horizontal='center', vertical='center')

    headers = [
        ('Cod. Piano Gold',   14),
        ('Cod. Art.',         12),
        ('Descrizione Art.',  40),
        ('Cod. Forn.',        12),
        ('Fornitore',         30),
        ('Giac. PDV',         12),
        ('Giac. Dep.',        12),
        ('Cod. Art. Expo',    14),
        ('Descrizione Expo',  30),
        ('Prezzo Promo',      12),
        ('Ordine in Corso',   16),
        ('Data Consegna',     14),
        ('Stato',             12),
        ('Nota',              35),
    ]

    # Riga intestazione con stile
    for i, (label, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = center
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'   # Blocca la riga di intestazione durante lo scroll
    ws.row_dimensions[1].height = 20

    # Righe dati con colore in base allo stato
    for r_idx, riga in enumerate(righe, 2):
        fill = fill_inc if riga.stato == 'incluso' else fill_exc if riga.stato == 'escluso' else (fill_nota if riga.nota else fill_neu)
        values = [
            riga.cod_piano_gold,
            riga.cod_art, riga.desc_art, riga.cod_forn, riga.desc_forn,
            riga.giac_pdv, riga.giac_dep,
            riga.cod_art_expo, riga.desc_expo,
            riga.prezzo_promo,
            riga.ordine_in_corso, riga.data_consegna,
            riga.get_stato_display(), riga.nota,
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = border
            if fill:
                cell.fill = fill
            # Formato numerico per giacenze (col. 6 e 7)
            if c_idx in (6, 7) and val is not None:
                cell.number_format = '#,##0'
            # Formato valuta per prezzo promo (col. 10)
            if c_idx == 10 and val is not None:
                cell.number_format = '€ #,##0.00'
            # Formato data per data consegna (col. 12)
            if c_idx == 12 and val is not None:
                cell.number_format = 'DD/MM/YYYY'

    # Filtri automatici sull'intestazione
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"promo_{sessione.piano_gold}.xlsx"
    resp = HttpResponse(buf.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
