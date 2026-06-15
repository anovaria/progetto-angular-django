"""
Picking Negativi - Views
Modulo per la visualizzazione e l'esportazione dei record di picking con quantità negative.

Legge i dati dalla tabella t_PickingNegativi sul database 'goldreport'.
I picking negativi indicano situazioni in cui la quantità movimentata
è inferiore a zero (es. rettifiche, errori di inventario).

Funzionalità:
- main: pagina di visualizzazione tabellare di tutti i record
- export_excel: esportazione degli stessi dati in formato .xlsx
"""
from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Definizione delle colonne: tuple (campo_db, etichetta_display)
# Usate sia per la tabella HTML che per le intestazioni Excel
COLONNE = [
    ('ID',                'ID'),
    ('DTAAGGIO',          'Data Agg.'),
    ('DataCreazione',     'Data Creazione'),
    ('DEPOSITO',          'Deposito'),
    ('PRODOTTO',          'Prodotto'),
    ('VARIANTE',          'Variante'),
    ('DESCRIZIONE',       'Descrizione'),
    ('IMBALLO',           'Imballo'),
    ('QTA_PEZZI',         'Qta Pezzi'),
    ('QTA_COLLI',         'Qta Colli'),
    ('INDIRIZZO_PICKING', 'Indirizzo Pick'),
]


def _esegui_query():
    """Esegue la query sulla tabella t_PickingNegativi e restituisce tutti i record.

    I record sono ordinati per DataCreazione decrescente (i più recenti prima)
    e per ID decrescente come criterio di tiebreak.

    Restituisce:
        list[dict]: lista di dizionari con i campi della tabella.
    """
    sql = """
        SELECT ID, DTAAGGIO, DataCreazione, DEPOSITO, PRODOTTO, VARIANTE,
               DESCRIZIONE, IMBALLO, QTA_PEZZI, QTA_COLLI, INDIRIZZO_PICKING
        FROM t_PickingNegativi
        ORDER BY DataCreazione DESC, ID DESC
    """
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql)
        # Costruisce dizionari {colonna: valore} per ogni riga
        cols = [c[0] for c in cursor.description]
        rows = [dict(zip(cols, row)) for row in cursor.fetchall()]
    return rows


def main(request):
    """Vista principale del modulo Picking Negativi.

    Esegue la query e passa i risultati al template. In caso di errore
    di connessione o query, mostra il messaggio di errore nel template
    senza interrompere la pagina.
    """
    errore = None
    righe = []
    try:
        righe = _esegui_query()
    except Exception as e:
        # Cattura qualsiasi eccezione (connessione, timeout, ecc.) e la passa al template
        errore = str(e)

    ctx = {
        'colonne': COLONNE,
        'righe': righe,
        'totale': len(righe),
        'errore': errore,
    }
    return render(request, 'picking_negativi/main.html', ctx)


def export_excel(request):
    """Esporta tutti i record di picking negativi in un file Excel (.xlsx).

    Il file generato ha:
    - Intestazioni con sfondo blu #1F4E79, testo bianco grassetto, centrate
    - Larghezza colonne auto-adattata al contenuto (max 40 caratteri)
    """
    righe = _esegui_query()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Picking Negativi'

    # Stile intestazione: sfondo blu scuro, testo bianco
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')

    # Scrive le intestazioni usando le etichette display da COLONNE
    for col_idx, (_, label) in enumerate(COLONNE, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Scrive i dati: per ogni riga usa il campo_db per prelevare il valore dal dizionario
    for row_idx, riga in enumerate(righe, 2):
        for col_idx, (campo, _) in enumerate(COLONNE, 1):
            ws.cell(row=row_idx, column=col_idx, value=riga.get(campo))

    # Auto-dimensionamento colonne (max 40 caratteri per non allargare eccessivamente)
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="picking_negativi.xlsx"'
    wb.save(response)
    return response
