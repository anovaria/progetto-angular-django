from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# V_QC110: controllo legami kit/bundle - verifica coerenza prezzi componenti vs kit
COLONNE = [
    ('DTAAGGIO',            'Data Agg.'),
    ('PROMO',               'Promo'),
    ('DATA_INIZIO',         'Data Inizio'),
    ('DATA_FINE',           'Data Fine'),
    ('COMPONENTE',          'Cod. Componente'),
    ('DESC_COMP',           'Descr. Componente'),
    ('PREZZO_COMP',         'Prezzo Comp.'),
    ('COEFF',               'Coeff.'),
    ('KIT',                 'Cod. Kit'),
    ('DESC_KIT',            'Descr. Kit'),
    ('PREZZO_VENDITA_KIT',  'Prezzo Vendita Kit'),
    ('PREZZO_KIT_CALC',     'Prezzo Kit Calc.'),
]


def _esegui_query():
    sql = """
        SELECT
            DTAAGGIO, PROMO, DATA_INIZIO, DATA_FINE,
            COMPONENTE, DESC_COMP, PREZZO_COMP, COEFF,
            KIT, DESC_KIT, PREZZO_VENDITA_KIT, PREZZO_KIT_CALC
        FROM V_QC110
        ORDER BY PROMO, KIT, COMPONENTE
    """
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql)
        cols = [c[0] for c in cursor.description]
        rows = [dict(zip(cols, row)) for row in cursor.fetchall()]
    return rows


def main(request):
    righe = _esegui_query()
    ctx = {
        'colonne': COLONNE,
        'righe': righe,
        'totale': len(righe),
    }
    return render(request, 'controllo_legami/main.html', ctx)


def export_excel(request):
    righe = _esegui_query()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Controllo Legami'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, (_, label) in enumerate(COLONNE, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, riga in enumerate(righe, 2):
        for col_idx, (campo, _) in enumerate(COLONNE, 1):
            ws.cell(row=row_idx, column=col_idx, value=riga.get(campo))

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="controllo_legami.xlsx"'
    wb.save(response)
    return response
