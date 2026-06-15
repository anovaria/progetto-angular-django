"""
Master Ordini - Views
Modulo per la consultazione e l'esportazione degli ordini fornitori.

Legge i dati dalla vista SQL dbo.V_OrdiniGenerale sul database 'goldreport'
tramite connessione pyodbc diretta (non ORM Django).

Funzionalità:
- index: pagina di ricerca con filtri (numero ordine, date, stato)
- export_excel: esportazione dei risultati filtrati in formato .xlsx
"""
from django.shortcuts import render
from django.http import HttpResponse
import pyodbc
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def get_connection():
    """Crea e restituisce una connessione pyodbc al database 'goldreport'.

    I parametri di connessione (driver, server, database, credenziali) vengono
    letti dalla configurazione DATABASES['goldreport'] di Django."""
    from django.conf import settings
    db = settings.DATABASES['goldreport']
    conn_str = (
        f"DRIVER={{{db['OPTIONS']['driver']}}};"
        f"SERVER={db['HOST']};"
        f"DATABASE={db['NAME']};"
        f"UID={db['USER']};"
        f"PWD={db['PASSWORD']};"
        f"{db['OPTIONS']['extra_params']}"
    )
    conn = pyodbc.connect(conn_str)
    return conn


def get_stati():
    """Recupera i valori distinti di TRAD (stato ordine) dalla vista V_OrdiniGenerale.
    Restituisce una lista di tuple (TRAD, STATO) ordinate per TRAD."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT TRAD, STATO FROM dbo.V_OrdiniGenerale WHERE TRAD IS NOT NULL ORDER BY TRAD")
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_num_ordini():
    """Recupera i numeri di ordine distinti (DCDCEXCDE) dalla vista V_OrdiniGenerale.
    Utilizzati per popolare il select2 nella pagina di ricerca."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT DCDCEXCDE FROM dbo.V_OrdiniGenerale WHERE DCDCEXCDE IS NOT NULL ORDER BY DCDCEXCDE")
    rows = [r[0] for r in cursor.fetchall()]
    conn.close()
    return rows


# Intestazioni della tabella dei risultati (e del file Excel esportato)
HEADERS = [
    'Dta Aggio', 'Data Ordine', 'Data Cons.', 'Cod. Ccom',
    'Descrizione Fornitore', 'Num. Ordine', 'Stato Ordine',
    'Codice art Forn.', 'Cod. Articolo', 'Descrizione Articolo',
    'S', 'Qta Inserita', 'Colli', 'Pz X UMB', 'Gestione',
    'Costo acq Singolo', 'Iva'
]


def index(request):
    """Vista principale del modulo Master Ordini.

    Se la richiesta contiene parametri GET, esegue la query filtrata su
    V_OrdiniGenerale e restituisce i risultati; altrimenti mostra il form vuoto.

    Filtri disponibili:
    - num_ordine: codice ordine (DCDCEXCDE)
    - data_ordine_da / data_ordine_a: intervallo date ordine
    - data_consegna_da / data_consegna_a: intervallo date consegna
    - stato: stato ordine (TRAD)
    """
    stati = get_stati()
    num_ordini = get_num_ordini()
    rows = None
    params_qs = ''

    if request.GET:
        # Legge e normalizza i parametri di filtro dalla query string
        num_ordine      = request.GET.get('num_ordine', '').strip()
        data_ordine_da  = request.GET.get('data_ordine_da', '').strip()
        data_ordine_a   = request.GET.get('data_ordine_a', '').strip()
        data_consegna_da = request.GET.get('data_consegna_da', '').strip()
        data_consegna_a  = request.GET.get('data_consegna_a', '').strip()
        stato           = request.GET.get('stato', '').strip()

        # Costruisce dinamicamente la clausola WHERE in base ai filtri valorizzati
        filters, params = [], []
        if num_ordine:
            filters.append("DCDCEXCDE = ?"); params.append(num_ordine)
        if data_ordine_da:
            filters.append("DATA_ORDINE >= ?"); params.append(data_ordine_da)
        if data_ordine_a:
            filters.append("DATA_ORDINE <= ?"); params.append(data_ordine_a)
        if data_consegna_da:
            filters.append("DATA_CONSEGNA >= ?"); params.append(data_consegna_da)
        if data_consegna_a:
            filters.append("DATA_CONSEGNA <= ?"); params.append(data_consegna_a)
        if stato:
            filters.append("TRAD = ?"); params.append(stato)

        where = ("WHERE " + " AND ".join(filters)) if filters else ""
        sql = f"""
            SELECT
                DTAAGGIO, DATA_ORDINE, DATA_CONSEGNA, CCOM, DESCRCCOM,
                DCDCEXCDE, STATO, CODARTFO, CODART, DESCRART,
                S, QTAINS, COLLIORD, PZXUL, GESTIONE, NETTONETTO, IVA
            FROM dbo.V_OrdiniGenerale
            {where}
            ORDER BY DCDCEXCDE, CODART
        """
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        conn.close()

        # Mantiene i parametri di filtro per il link "Esporta Excel"
        params_qs = request.GET.urlencode()

    context = {
        'stati': stati,
        'num_ordini': num_ordini,
        'today': date.today().isoformat(),
        'headers': HEADERS,
        'rows': rows,
        'params_qs': params_qs,
        'filters': request.GET,
    }
    return render(request, 'master_ordini/index.html', context)


def export_excel(request):
    """Esporta i risultati della ricerca in un file Excel (.xlsx).

    Utilizza gli stessi filtri della vista index, passati tramite query string.
    Il file generato ha:
    - Intestazioni con sfondo blu scuro e testo bianco
    - Colonna NETTONETTO formattata a 2 decimali
    - Prima colonna (DTAAGGIO) vuota nel file di output
    """
    # Ripete la stessa logica di filtro della vista index
    num_ordine = request.GET.get('num_ordine', '').strip()
    data_ordine_da = request.GET.get('data_ordine_da', '').strip()
    data_ordine_a = request.GET.get('data_ordine_a', '').strip()
    data_consegna_da = request.GET.get('data_consegna_da', '').strip()
    data_consegna_a = request.GET.get('data_consegna_a', '').strip()
    stato = request.GET.get('stato', '').strip()

    filters = []
    params = []

    if num_ordine:
        filters.append("DCDCEXCDE = ?")
        params.append(num_ordine)
    if data_ordine_da:
        filters.append("DATA_ORDINE >= ?")
        params.append(data_ordine_da)
    if data_ordine_a:
        filters.append("DATA_ORDINE <= ?")
        params.append(data_ordine_a)
    if data_consegna_da:
        filters.append("DATA_CONSEGNA >= ?")
        params.append(data_consegna_da)
    if data_consegna_a:
        filters.append("DATA_CONSEGNA <= ?")
        params.append(data_consegna_a)
    if stato:
        filters.append("TRAD = ?")
        params.append(stato)

    where = ("WHERE " + " AND ".join(filters)) if filters else ""

    sql = f"""
        SELECT
            DTAAGGIO, DATA_ORDINE, DATA_CONSEGNA, CCOM, DESCRCCOM,
            DCDCEXCDE, STATO, CODARTFO, CODART, DESCRART,
            S, QTAINS, COLLIORD, PZXUL, GESTIONE, NETTONETTO, IVA
        FROM dbo.V_OrdiniGenerale
        {where}
        ORDER BY DCDCEXCDE, CODART
    """

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(sql, params)
    rows = cursor.fetchall()
    conn.close()

    # Crea Excel con openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ModelloInvioOrdineStandard'

    headers = [
        '', 'Data Ordine', 'Data Cons.', 'Cod. Ccom',
        'Descrizione Fornitore', 'NumOrdine', 'Stato Ordine',
        'Codice art Forn.', 'Cod. Articolo', 'Descrizione Articolo',
        'S', 'Qta Inserita', 'Colli', 'Pz X UMB', 'Gestione',
        'Costo acq Singolo', 'Iva'
    ]

    # Stile header: testo bianco su sfondo blu #2F5496, centrato
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='2F5496')
    header_align = Alignment(horizontal='center', vertical='center')

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Scrittura dati: col 1 vuota (DTAAGGIO non esportata), col 16 (NETTONETTO) a 2 decimali
    NETTONETTO_COL = 16
    for row in rows:
        row_data = list(row)
        row_data[0] = None  # Prima colonna lasciata vuota nel file Excel
        if row_data[NETTONETTO_COL - 1] is not None:
            try:
                row_data[NETTONETTO_COL - 1] = round(float(row_data[NETTONETTO_COL - 1]), 2)
            except (ValueError, TypeError):
                pass
        ws.append(row_data)

    # Applica il formato numero a 2 decimali sulla colonna NETTONETTO
    for row_cells in ws.iter_rows(min_row=2, min_col=NETTONETTO_COL, max_col=NETTONETTO_COL):
        for cell in row_cells:
            cell.number_format = '#,##0.00'

    # Imposta la larghezza delle colonne secondo valori predefiniti
    col_widths = [12, 12, 12, 10, 35, 18, 20, 16, 14, 40, 4, 12, 8, 10, 12, 18, 6]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Risposta HTTP con Content-Disposition per il download diretto
    filename = f"ordini_fornitore.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
