"""
Masterdata - Views
Modulo per la consultazione e l'esportazione del catalogo articoli (masterdata).

Legge i dati dalla tabella dbo.t_masterdataall (e tabelle correlate) sul database
'goldreport' tramite l'ORM di Django (connections['goldreport']).

Funzionalità:
- masterdata_list: pagina di ricerca con filtri (reparto, stato, fornitore, testo, EAN, Cervinia)
  con paginazione lato server (PAGE_SIZE = 100 righe per pagina)
- masterdata_export: esportazione dell'intera selezione filtrata in formato .xlsx
"""
from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
from django.utils import timezone
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Numero di articoli per pagina nella vista lista
PAGE_SIZE = 100


def execute_masterdata_query(filters=None):
    """Esegue la query principale sul catalogo articoli con i filtri specificati.

    Effettua una LEFT JOIN con dbo.t_prezziPdv per i prezzi per punto vendita.
    Se il filtro 'cervinia' è attivo, aggiunge una INNER JOIN con t_assoCervinia
    per restituire solo gli articoli associati al PDV Cervinia.

    Argomenti:
        filters (dict): dizionario con chiavi opzionali:
            - reparto: codice reparto
            - stato: stato articolo
            - fornitore: codice CCOM
            - search: testo libero (cerca su descrizione, codice, EAN, fornitore)
            - solo_ean_princ: se valorizzato, filtra solo gli EAN principali
            - cervinia: se valorizzato, filtra solo articoli Cervinia

    Restituisce:
        list[dict]: lista di dizionari con un campo per ogni colonna.
    """
    # La INNER JOIN con t_assoCervinia viene aggiunta solo se il filtro è attivo
    cervinia_join = ""
    if filters and filters.get('cervinia'):
        cervinia_join = "INNER JOIN dbo.t_assoCervinia ac ON m.CODART = ac.CODART"

    query = f"""
    SELECT DISTINCT
        m.CODART,
        m.DESCRART,
        m.STATO,
        m.REP,
        m.DESCRREP,
        m.SREP,
        m.DESCRSREP,
        m.FAM,
        m.DESCRFAM,
        m.CCOM,
        m.DESCRCCOM,
        m.CODFORN,
        m.DESCFORN,
        m.EAN,
        m.PRACQ,
        m.IVA,
        p.PREZZO_AOSTA,
        p.PREZZO_CERVINIA
    FROM dbo.t_masterdataall m
    LEFT JOIN dbo.t_prezziPdv p ON m.CODART = p.CODART
    {cervinia_join}
    WHERE 1=1
    """

    params = []

    if filters:
        if filters.get('reparto'):
            query += " AND m.REP = %s"
            params.append(filters['reparto'])
        if filters.get('stato'):
            query += " AND m.STATO = %s"
            params.append(filters['stato'])
        if filters.get('fornitore'):
            query += " AND m.CCOM = %s"
            params.append(filters['fornitore'])
        if filters.get('search'):
            # Ricerca testuale su più campi con LIKE
            s = f"%{filters['search']}%"
            query += " AND (m.DESCRART LIKE %s OR m.CODART LIKE %s OR m.EAN LIKE %s OR m.DESCFORN LIKE %s)"
            params.extend([s, s, s, s])
        if filters.get('solo_ean_princ'):
            # Filtra solo gli EAN marcati come principali in t_t_Ean
            query += " AND m.EAN IN (SELECT EAN FROM dbo.t_t_Ean WHERE PRINC = 1)"

    query += " ORDER BY m.REP, m.DESCRART"

    with connections['goldreport'].cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]

    return results


def get_filter_options():
    """Recupera i valori distinti per i filtri della pagina lista.

    Legge in un'unica query i valori di REP, STATO e CCOM da t_masterdataall
    e li organizza in tre dizionari separati per evitare duplicati.

    Restituisce:
        dict con chiavi 'reparti', 'stati', 'fornitori': liste di tuple (codice, descrizione).
    """
    query = """
    SELECT DISTINCT REP, DESCRREP, STATO, CCOM, DESCRCCOM
    FROM dbo.t_masterdataall
    ORDER BY REP
    """
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()

    # Raccoglie valori unici per ogni tipo di filtro
    reparti, stati, fornitori = {}, {}, {}
    for rep, descrrep, stato, ccom, descrccom in rows:
        if rep and rep not in reparti:
            reparti[rep] = descrrep or f"Reparto {rep}"
        if stato and stato not in stati:
            stati[stato] = stato
        if ccom and ccom not in fornitori:
            fornitori[ccom] = descrccom or ccom

    return {
        'reparti': sorted(reparti.items()),
        'stati': sorted(stati.items()),
        # I fornitori sono ordinati per descrizione (non per codice)
        'fornitori': sorted(fornitori.items(), key=lambda x: x[1] or ''),
    }


def masterdata_list(request):
    """Vista principale del modulo Masterdata.

    Mostra il form di filtro e, se almeno un filtro è valorizzato,
    esegue la query e pagina i risultati con Paginator.
    Se nessun filtro è attivo, mostra lo stato "vuoto" senza query.
    """
    # Raccoglie i filtri dalla query string, escludendo quelli vuoti
    filters = {
        'reparto': request.GET.get('reparto'),
        'stato': request.GET.get('stato'),
        'fornitore': request.GET.get('fornitore'),
        'search': request.GET.get('search'),
        'cervinia': request.GET.get('cervinia'),
        'solo_ean_princ': request.GET.get('solo_ean_princ'),
    }
    filters = {k: v for k, v in filters.items() if v}

    filter_options = get_filter_options()
    page_obj = None
    totale = 0

    # Esegue la query solo se c'è almeno un filtro valorizzato
    if filters:
        articoli = execute_masterdata_query(filters)
        totale = len(articoli)
        # Paginazione: restituisce la pagina richiesta tramite il parametro 'page'
        paginator = Paginator(articoli, PAGE_SIZE)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'totale': totale,
        'filters': filters,
        'filter_options': filter_options,
    }
    return render(request, 'masterdata/list.html', context)


def masterdata_export(request):
    """Esporta il catalogo articoli filtrato in un file Excel (.xlsx).

    Utilizza gli stessi filtri della vista masterdata_list.
    Se nessun filtro è valorizzato, esporta l'intero catalogo.
    Il file include tutte le colonne della query (senza paginazione).
    """
    # Raccoglie e pulisce i filtri dalla query string
    filters = {
        'reparto': request.GET.get('reparto'),
        'stato': request.GET.get('stato'),
        'fornitore': request.GET.get('fornitore'),
        'search': request.GET.get('search'),
        'cervinia': request.GET.get('cervinia'),
        'solo_ean_princ': request.GET.get('solo_ean_princ'),
    }
    filters = {k: v for k, v in filters.items() if v}

    # Se nessun filtro, esporta tutto il catalogo
    results = execute_masterdata_query(filters if filters else None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Masterdata"

    # Stile intestazione: sfondo blu #2E5090, testo bianco grassetto
    header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ['CodArt', 'Descrizione', 'Stato', 'Rep', 'Reparto', 'SRep', 'Sottoreparto',
               'Famiglia', 'EAN', 'CodComm', 'Linea Comm', 'CodForn', 'Fornitore',
               'Prezzo Aosta', 'Prezzo Cervinia', 'Prezzo Acq', 'IVA']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Scrittura dati: una riga per articolo
    for row_num, item in enumerate(results, 2):
        ws.cell(row=row_num, column=1, value=item.get('CODART'))
        ws.cell(row=row_num, column=2, value=item.get('DESCRART'))
        ws.cell(row=row_num, column=3, value=item.get('STATO'))
        ws.cell(row=row_num, column=4, value=item.get('REP'))
        ws.cell(row=row_num, column=5, value=item.get('DESCRREP'))
        ws.cell(row=row_num, column=6, value=item.get('SREP'))
        ws.cell(row=row_num, column=7, value=item.get('DESCRSREP'))
        ws.cell(row=row_num, column=8, value=item.get('DESCRFAM'))
        ws.cell(row=row_num, column=9, value=item.get('EAN'))
        ws.cell(row=row_num, column=10, value=item.get('CCOM'))
        ws.cell(row=row_num, column=11, value=item.get('DESCRCCOM'))
        ws.cell(row=row_num, column=12, value=item.get('CODFORN'))
        ws.cell(row=row_num, column=13, value=item.get('DESCFORN'))
        ws.cell(row=row_num, column=14, value=item.get('PREZZO_AOSTA'))
        ws.cell(row=row_num, column=15, value=item.get('PREZZO_CERVINIA'))
        ws.cell(row=row_num, column=16, value=item.get('PRACQ'))
        ws.cell(row=row_num, column=17, value=item.get('IVA'))

    # Auto-dimensionamento colonne (max 50 caratteri di larghezza)
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Il nome del file include il timestamp per renderlo univoco
    filename = f'masterdata_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
