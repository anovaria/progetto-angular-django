from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

COLONNE = [
    ('DTAAGGIO',       'Data Agg.'),
    ('nr',             'Nr'),
    ('CODART',         'Cod. Art.'),
    ('DescrArticolo',  'Descrizione'),
]

LIMIT_DEFAULT = 500


def _esegui_query(ricerca=None, limit=None):
    where, params = [], []

    if ricerca:
        like = f"%{ricerca}%"
        where.append("(CODART LIKE %s OR DescrArticolo LIKE %s)")
        params.extend([like, like])

    where_clause = f"WHERE {' AND '.join(where)}" if where else ""
    top_clause   = f"TOP {limit}" if limit else ""

    sql = f"""
        SELECT {top_clause} DTAAGGIO, nr, CODART, DescrArticolo
        FROM V_promoDoppieDomani
        {where_clause}
        ORDER BY nr, CODART
    """
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql, params)
        cols = [c[0] for c in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]


def main(request):
    ricerca  = request.GET.get('ricerca', '').strip()
    filtrato = bool(ricerca)
    limit    = None if filtrato else LIMIT_DEFAULT
    righe    = _esegui_query(ricerca=ricerca or None, limit=limit)

    ctx = {
        'colonne':  COLONNE,
        'righe':    righe,
        'totale':   len(righe),
        'troncato': (not filtrato and len(righe) == LIMIT_DEFAULT),
        'ricerca':  ricerca,
    }
    return render(request, 'promo_doppie_domani/main.html', ctx)


def export_excel(request):
    ricerca = request.GET.get('ricerca', '').strip()
    righe   = _esegui_query(ricerca=ricerca or None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Promo Doppie Domani'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, (_, label) in enumerate(COLONNE, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, riga in enumerate(righe, 2):
        for col_idx, (campo, _) in enumerate(COLONNE, 1):
            ws.cell(row=row_idx, column=col_idx, value=riga.get(campo))

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="promo_doppie_domani.xlsx"'
    wb.save(response)
    return response
