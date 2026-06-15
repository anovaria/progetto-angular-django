"""
Views Django per la gestione Welfare Aziendale

Questo modulo gestisce tutte le operazioni legate ai buoni welfare aziendali:
- Dashboard con statistiche in tempo reale
- Ricerca voucher per front-office e cassa
- Registrazione consegne (cassa)
- Lista buoni da consegnare
- Contabilità e generazione report (Excel + PDF in ZIP)
- Import manuale email ePassi con area provvisoria e validazione
- API JSON per ricerca voucher e statistiche
- Storico consegne per punto.info
"""

from io import BytesIO
import zipfile
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.contrib import messages
from decimal import Decimal, InvalidOperation
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from .models import (
    TaglioBuono, RichiestaWelfare,
    DettaglioBuono, RichiestaProvvisoria
)


# ============================================================
# AUTENTICAZIONE (semplificata - ruoli gestiti da Angular)
# ============================================================

def get_current_user(request):
    """Recupera lo username dalla sessione portale."""
    session_user = request.session.get('user', {})
    return session_user.get('username', 'anonymous').lower()


def require_auth(view_func):
    """Inietta request.welfare_user dalla sessione portale."""
    def wrapper(request, *args, **kwargs):
        request.welfare_user = get_current_user(request)
        return view_func(request, *args, **kwargs)
    return wrapper

# ============================================================
# DASHBOARD PRINCIPALE
# ============================================================

@csrf_exempt
@require_auth
def dashboard(request):
    """
    Dashboard principale con statistiche e accesso rapido.

    Mostra un riepilogo dello stato corrente del sistema welfare:
    contatori per stato, valore economico del mese, ultime richieste
    in attesa e avvisi per richieste provvisorie da validare.
    """
    username = request.welfare_user

    # Calcola la data odierna e l'inizio del mese corrente per i filtri statistici
    oggi = timezone.now().date()
    inizio_mese = oggi.replace(day=1)

    # Raccoglie le statistiche aggregate da mostrare nelle card della dashboard
    stats = {
        # Buoni ricevuti dal sistema ePassi, pronti per la consegna fisica
        'totale_pronto': RichiestaWelfare.objects.filter(stato='PRONTO').count(),
        # Buoni in lavorazione intermedia (es. in preparazione)
        'totale_elaborato': RichiestaWelfare.objects.filter(stato='ELABORATO').count(),
        # Consegne effettuate nella giornata corrente
        'consegnati_oggi': RichiestaWelfare.objects.filter(
            stato='CONSEGNATO',
            data_consegna__date=oggi
        ).count(),
        # Consegne effettuate dall'inizio del mese corrente
        'consegnati_mese': RichiestaWelfare.objects.filter(
            stato='CONSEGNATO',
            data_consegna__date__gte=inizio_mese
        ).count(),
        # Valore economico totale dei buoni consegnati nel mese (somma totale_buono)
        'valore_consegnati_mese': RichiestaWelfare.objects.filter(
            stato='CONSEGNATO',
            data_consegna__date__gte=inizio_mese
        ).aggregate(tot=Sum('totale_buono'))['tot'] or 0,
        # Richieste importate via email ma non ancora validate dall'amministratore
        'provvisori_da_validare': RichiestaProvvisoria.objects.filter(
            processato=False
        ).count(),
    }

    # Recupera le ultime 10 richieste pronte per visualizzarle nella tabella rapida
    ultime_pronte = RichiestaWelfare.objects.filter(
        stato='PRONTO'
    ).order_by('-data_creazione')[:10]

    context = {
        'username': username,
        'stats': stats,
        'ultime_pronte': ultime_pronte,
        'oggi': oggi,
    }

    return render(request, 'welfare/dashboard.html', context)


# ============================================================
# RICERCA VOUCHER (Front-Office e Cassa)
# ============================================================

@csrf_exempt
@require_auth
def ricerca_voucher(request):
    """
    Ricerca una richiesta welfare per numero voucher.

    Permette la consultazione dello stato e del dettaglio di un voucher
    tramite il suo codice a 8 cifre (num_richiesta). Usato sia dal
    front-office per informazioni che dalla cassa prima della consegna.
    """
    username = request.welfare_user

    richiesta = None
    dettagli = []
    # Legge il parametro voucher dalla query string e rimuove spazi accidentali
    num_voucher = request.GET.get('voucher', '').strip()

    if num_voucher:
        try:
            # Ricerca esatta per numero richiesta (codice univoco ePassi)
            richiesta = RichiestaWelfare.objects.get(num_richiesta=num_voucher)
            # Carica i dettagli dei tagli buono con join sulla tabella TaglioBuono
            dettagli = richiesta.dettagli_buoni.select_related('taglio').all()
        except RichiestaWelfare.DoesNotExist:
            messages.warning(request, f'Voucher {num_voucher} non trovato')

    context = {
        'username': username,
        'num_voucher': num_voucher,
        'richiesta': richiesta,
        'dettagli': dettagli,
    }

    return render(request, 'welfare/ricerca_voucher.html', context)


# ============================================================
# GESTIONE CASSA
# ============================================================

@require_auth
def cassa_consegna(request):
    """
    Interfaccia cassa per registrare la consegna fisica dei buoni.

    Supporta due azioni via POST:
    - 'consegna': segna il voucher come consegnato al beneficiario
    - 'salva_tagli': registra il dettaglio dei tagli buono utilizzati
      (es. quanti buoni da 5€, quanti da 10€, ecc.)

    In GET mostra il form di ricerca e, se trovato, il dettaglio del voucher
    solo se è in stato PRONTO (non già consegnato o inevaso).
    """
    username = request.welfare_user

    richiesta = None
    dettagli = []
    # Recupera tutti i tagli buono configurati e attivi nel sistema
    tagli_disponibili = TaglioBuono.objects.filter(attivo=True)
    num_voucher = request.GET.get('voucher', '').strip()

    if num_voucher:
        try:
            # Cerca solo voucher in stato PRONTO: quelli già consegnati non possono essere rilavorati
            richiesta = RichiestaWelfare.objects.get(
                num_richiesta=num_voucher,
                stato='PRONTO'
            )
            dettagli = richiesta.dettagli_buoni.select_related('taglio').all()
        except RichiestaWelfare.DoesNotExist:
            messages.warning(request, f'Voucher {num_voucher} non trovato o non in stato PRONTO')

    if request.method == 'POST':
        richiesta_id = request.POST.get('richiesta_id')
        azione = request.POST.get('azione')

        if richiesta_id:
            richiesta = get_object_or_404(RichiestaWelfare, pk=richiesta_id)

            if azione == 'consegna':
                # Registra la consegna impostando data_consegna e utente_consegna
                richiesta.segna_consegnato('punto.info')
                messages.success(request, f'Voucher {richiesta.num_richiesta} consegnato!')
                # Redirect POST/GET per evitare doppio submit con F5
                return redirect('welfare:cassa_consegna')

            elif azione == 'salva_tagli':
                # Elimina i dettagli precedenti e li riscrive con i valori aggiornati
                DettaglioBuono.objects.filter(richiesta=richiesta).delete()

                # Itera su tutti i tagli disponibili e legge la quantità dal form
                for taglio in tagli_disponibili:
                    # Il nome del campo HTML segue la convenzione "qta_<valore_nominale>"
                    qta_key = f'qta_{taglio.valore_nominale}'
                    qta = request.POST.get(qta_key, '0')
                    try:
                        qta = int(qta)
                        # Crea il dettaglio solo se la quantità è positiva
                        if qta > 0:
                            DettaglioBuono.objects.create(
                                richiesta=richiesta,
                                taglio=taglio,
                                quantita=qta
                            )
                    except (ValueError, TypeError):
                        # Ignora valori non numerici nel campo quantità
                        pass

                messages.success(request, 'Dettaglio tagli salvato')
                # Torna alla stessa pagina con il voucher nella query string
                return redirect(f'?voucher={richiesta.num_richiesta}')

    context = {
        'username': username,
        'num_voucher': num_voucher,
        'richiesta': richiesta,
        'dettagli': dettagli,
        'tagli_disponibili': tagli_disponibili,
    }

    return render(request, 'welfare/cassa_consegna.html', context)


# ============================================================
# LISTA BUONI DA CONSEGNARE
# ============================================================

@csrf_exempt
@require_auth
def lista_da_consegnare(request):
    """
    Lista di tutti i buoni welfare in stato PRONTO, con filtri e ordinamento.

    Permette di filtrare per codice voucher e per azienda mittente.
    Supporta ordinamento per colonna (voucher, nominativo, data, qta, totale)
    con direzione ascendente/discendente tramite parametri GET.
    """
    username = request.welfare_user

    # Parte dalla queryset base: solo richieste in stato PRONTO
    richieste = RichiestaWelfare.objects.filter(stato='PRONTO')

    # Sezione filtri: ogni filtro è opzionale e si applica solo se valorizzato
    codice = request.GET.get('codice', '').strip()
    if codice:
        # Ricerca parziale case-insensitive sul codice voucher
        richieste = richieste.filter(num_richiesta__icontains=codice)

    azienda = request.GET.get('azienda', '').strip()
    if azienda:
        # Ricerca parziale sulla ragione sociale del mittente (azienda convenzionata)
        richieste = richieste.filter(nome_mittente__icontains=azienda)

    # Mappa i parametri GET ai nomi dei campi del modello per l'ordinamento
    SORT_FIELDS = {
        'voucher': 'num_richiesta',
        'nominativo': 'nominativo',
        'data': 'data_creazione',
        'qta': 'qta_buono',
        'totale': 'totale_buono',
    }
    sort = request.GET.get('sort', 'data')
    dir = request.GET.get('dir', 'desc')
    sort_field = SORT_FIELDS.get(sort, 'data_creazione')
    # Prefisso '-' per ordinamento discendente (convenzione Django ORM)
    if dir == 'desc':
        sort_field = f'-{sort_field}'
    richieste = richieste.order_by(sort_field)

    # Calcola il valore totale di tutti i buoni in attesa (aggregazione SQL)
    totale_valore = richieste.aggregate(tot=Sum('totale_buono'))['tot'] or 0

    context = {
        'username': username,
        'richieste': richieste,
        'filtro_codice': codice,
        'filtro_azienda': azienda,
        'totale_valore': totale_valore,
        'sort': sort,
        'dir': dir,
    }

    return render(request, 'welfare/lista_da_consegnare.html', context)


# ============================================================
# CONTABILITÀ E REPORT
# ============================================================

@csrf_exempt
@require_auth
def contabilita(request):
    """
    Interfaccia Contabilità con filtri avanzati e report.

    Mostra l'elenco delle richieste consegnate filtrate per anno/mese/giorno,
    con totali aggregati. Fornisce anche il link per scaricare il report
    in formato ZIP (Excel + PDF) tramite la view report_contabilita.
    """
    username = request.welfare_user

    oggi = timezone.now().date()
    # Usa i parametri GET per i filtri; default: mese e anno correnti
    anno = int(request.GET.get('anno', oggi.year))
    mese = int(request.GET.get('mese', oggi.month))
    giorno_param = request.GET.get('giorno', '')
    # Il giorno è opzionale: None significa "tutti i giorni del mese"
    giorno = int(giorno_param) if giorno_param else None

    # Query base: solo richieste CONSEGNATE nel mese/anno selezionato
    richieste = RichiestaWelfare.objects.filter(
        stato='CONSEGNATO',
        data_consegna__year=anno,
        data_consegna__month=mese
    )

    # Filtro giorno (opzionale): restringe ulteriormente al giorno specifico
    if giorno:
        richieste = richieste.filter(data_consegna__day=giorno)

    richieste = richieste.order_by('-data_consegna')

    # Calcola i totali aggregati per il riepilogo
    totali = richieste.aggregate(
        num_richieste=Count('id'),
        valore_totale=Sum('totale_buono')
    )

    # Formatta il valore totale in stile europeo (es. 1.234,56) per la visualizzazione
    _val = float(totali['valore_totale'] or 0)
    valore_fmt = '{:,.2f}'.format(_val).replace(',', 'X').replace('.', ',').replace('X', '.')

    # Recupera gli anni per cui esistono consegne registrate (per il menu a tendina)
    anni_disponibili = RichiestaWelfare.objects.filter(
        data_consegna__isnull=False
    ).dates('data_consegna', 'year', order='DESC')

    context = {
        'username': username,
        'richieste': richieste,
        'anno': anno,
        'mese': mese,
        'giorno': giorno,
        'totali': totali,
        'valore_fmt': valore_fmt,
        # Estrae solo l'intero anno dalla lista di date Django
        'anni_disponibili': [d.year for d in anni_disponibili],
        'mesi': range(1, 13),
        'giorni': range(1, 32),
    }

    return render(request, 'welfare/contabilita.html', context)


@csrf_exempt
@require_auth
def report_contabilita(request):
    """
    Genera e scarica un file ZIP contenente il report contabile in formato Excel e PDF.

    Accetta gli stessi filtri di contabilita() (anno, mese, giorno) più
    filtri sullo stato delle richieste (pronto, elaborato, consegnato, inevaso).
    Se nessuno stato è selezionato, viene usato 'CONSEGNATO' come default.
    Il file ZIP scaricato contiene: <NomeBase>.xlsx e <NomeBase>.pdf.
    """

    # Leggi i filtri dalla query string (stessi parametri della view contabilita)
    oggi = timezone.now().date()
    anno = int(request.GET.get('anno', oggi.year))
    mese = int(request.GET.get('mese', oggi.month))
    giorno_param = request.GET.get('giorno', '')
    giorno = int(giorno_param) if giorno_param else None

    # Legge i flag degli stati: il parametro '1' significa stato incluso nel report
    mostra_pronto = request.GET.get('pronto', '') == '1'
    mostra_elaborato = request.GET.get('elaborato', '') == '1'
    mostra_consegnato = request.GET.get('consegnato', '1') == '1'
    mostra_inevaso = request.GET.get('inevaso', '') == '1'

    # Costruisce la lista degli stati da includere nel filtro ORM
    stati_filtro = []
    if mostra_pronto:
        stati_filtro.append('PRONTO')
    if mostra_elaborato:
        stati_filtro.append('ELABORATO')
    if mostra_consegnato:
        stati_filtro.append('CONSEGNATO')
    if mostra_inevaso:
        stati_filtro.append('INEVASO')

    # Fallback: se nessuno stato è selezionato, includi almeno i CONSEGNATI
    if not stati_filtro:
        stati_filtro = ['CONSEGNATO']

    # Query con filtro multi-stato e periodo temporale
    richieste = RichiestaWelfare.objects.filter(
        stato__in=stati_filtro,
        data_consegna__year=anno,
        data_consegna__month=mese
    )

    if giorno:
        richieste = richieste.filter(data_consegna__day=giorno)

    # Ordina per data consegna crescente (utile per la lettura del report)
    richieste = richieste.order_by('data_consegna')

    # Costruisce il nome base del file: include giorno solo se filtro giorno attivo
    if giorno:
        nome_base = f"Welfare_{anno}_{mese:02d}_{giorno:02d}"
    else:
        nome_base = f"Welfare_{anno}_{mese:02d}"

    # Genera i due formati di report
    excel_buffer = genera_excel_report(richieste, mese, anno)
    pdf_buffer = genera_pdf_report(richieste, mese, anno)

    # Crea il file ZIP in memoria contenente entrambi i report
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.writestr(f"{nome_base}.xlsx", excel_buffer.getvalue())
        zip_file.writestr(f"{nome_base}.pdf", pdf_buffer.getvalue())

    zip_buffer.seek(0)

    # Invia il file ZIP come risposta HTTP con header per il download
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{nome_base}.zip"'
    return response


def genera_excel_report(richieste, mese, anno):
    """
    Genera il file Excel di riepilogo buoni welfare nel formato richiesto.

    Produce un foglio 'Riepilogo' con intestazione, header colorato e
    una riga per ogni richiesta. In fondo aggiunge una riga con i totali.
    Restituisce un oggetto BytesIO pronto per essere scritto su disco o
    incluso in un archivio ZIP.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo"

    # Definizione degli stili riutilizzati nel foglio
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Sfondo verde chiaro per la riga degli header di colonna
    header_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

    # Riga 1: Titolo del report con merge su tutte le colonne dati (A-F)
    ws['A1'] = f"Riepilogo Buoni Welfare - Mese: {mese} Anno: {anno}"
    ws['A1'].font = header_font
    ws.merge_cells('A1:F1')

    # Riga 3: Header delle colonne (riga 2 lasciata vuota come spaziatore)
    headers = ['TotaleBuono', 'Nominativo', 'Num_Richiesta', 'QtaBuono', 'ValoreBuono', 'Data_ConsegnaCliente']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = title_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Popolamento dati a partire dalla riga 4
    row = 4
    totale_generale = 0
    totale_qta = 0

    for r in richieste:
        ws.cell(row=row, column=1, value=float(r.totale_buono)).border = border
        ws.cell(row=row, column=2, value=r.nominativo).border = border
        ws.cell(row=row, column=3, value=r.num_richiesta).border = border
        ws.cell(row=row, column=4, value=r.qta_buono).border = border
        ws.cell(row=row, column=5, value=float(r.valore_buono)).border = border
        # Formatta la data nel formato italiano gg/mm/aaaa; stringa vuota se assente
        data_cons = r.data_consegna.strftime('%d/%m/%Y') if r.data_consegna else ''
        ws.cell(row=row, column=6, value=data_cons).border = border

        # Accumula i totali per la riga di sommario finale
        totale_generale += float(r.totale_buono)
        totale_qta += r.qta_buono
        row += 1

    # Riga totale: inserita due righe dopo l'ultimo dato (separatore visivo)
    row += 1
    ws.cell(row=row, column=1, value=totale_generale).font = Font(bold=True, size=12)
    ws.cell(row=row, column=4, value=totale_qta).font = Font(bold=True, size=12)

    # Imposta larghezze colonne ottimizzate per la leggibilità
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

    # Salva il workbook in un buffer in memoria invece che su file
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def genera_pdf_report(richieste, mese, anno):
    """
    Genera il file PDF di riepilogo buoni welfare nel formato richiesto.

    Le richieste vengono raggruppate per valore del buono (es. tutti i buoni
    da 5€ insieme, poi quelli da 10€, ecc.) in ordine decrescente di valore.
    Per ogni richiesta viene generato un 'box' con bordo contenente nominativo,
    numero richiesta, totale, dettaglio tagli e data consegna.
    Aggiunge il numero di pagina in calce a ogni pagina.
    Restituisce un oggetto BytesIO pronto per l'uso.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER

    buffer = BytesIO()

    # Configura il documento PDF con margini ridotti per massimizzare lo spazio utile
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=15*mm,
        bottomMargin=20*mm,
        leftMargin=15*mm,
        rightMargin=15*mm
    )

    elements = []

    # Stile per il titolo centrato in cima al documento
    title_style = ParagraphStyle(
        'Title',
        fontName='Helvetica-Bold',
        fontSize=12,
        alignment=TA_CENTER,
        spaceAfter=10*mm
    )
    elements.append(Paragraph(f"Riepilogo Buoni Wellfare Mese: {mese} Anno: {anno}", title_style))

    # Stile per la riga principale del box (nominativo + numero richiesta + totale)
    riga1_style = ParagraphStyle(
        'Riga1',
        fontName='Helvetica',
        fontSize=10,
        leading=12
    )

    # Stile per la riga secondaria del box (dettaglio tagli + data consegna)
    riga2_style = ParagraphStyle(
        'Riga2',
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        leftIndent=10*mm
    )

    # Stile per la riga totale di gruppo (mostrata dopo ogni gruppo per valore)
    totale_style = ParagraphStyle(
        'Totale',
        fontName='Helvetica-Bold',
        fontSize=10,
        spaceBefore=3*mm,
        spaceAfter=8*mm
    )

    # Raggruppa le richieste per valore nominale del buono (es. {5: [...], 10: [...]})
    richieste_list = list(richieste)
    valori_buono = {}
    for r in richieste_list:
        valore = int(r.valore_buono)
        if valore not in valori_buono:
            valori_buono[valore] = []
        valori_buono[valore].append(r)

    # Itera sui gruppi in ordine decrescente di valore (prima i tagli più grandi)
    for valore in sorted(valori_buono.keys(), reverse=True):
        gruppo = valori_buono[valore]
        totale_qta_gruppo = 0
        totale_valore_gruppo = 0

        for r in gruppo:
            data_cons = r.data_consegna.strftime('%d/%m/%Y') if r.data_consegna else '-'
            totale = int(r.totale_buono)
            qta = r.qta_buono

            # Riga 1 del box: NOMINATIVO + N° RICHIESTA + Totale in grassetto
            riga1 = f"<b>{r.nominativo}</b> &nbsp;&nbsp;&nbsp;&nbsp;&nbsp; N ° RICHIESTA <b>{r.num_richiesta}</b> &nbsp;&nbsp; Per un totale di <b>{totale} €</b>"

            # Riga 2 del box: dettaglio tagli e data consegna al cliente
            riga2 = f"DETTAGLIO: N°{qta} BUONI Da {valore} € --- Data Cons Cliente {data_cons}"

            # Ogni richiesta viene rappresentata come una tabella a 1 colonna con bordo
            box_content = [
                [Paragraph(riga1, riga1_style)],
                [Paragraph(riga2, riga2_style)],
            ]

            # La larghezza del box è quasi l'intera pagina (meno i margini)
            box = Table(box_content, colWidths=[170*mm])
            box.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
                ('TOPPADDING', (0, 0), (-1, -1), 3*mm),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3*mm),
                ('LEFTPADDING', (0, 0), (-1, -1), 3*mm),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3*mm),
            ]))

            elements.append(box)
            # Piccolo spazio verticale tra i box per leggibilità
            elements.append(Spacer(1, 2*mm))

            # Accumula i totali di gruppo per la riga di sommario
            totale_qta_gruppo += qta
            totale_valore_gruppo += totale

        # Riga di totale per il gruppo (es. "Totale buoni da 10€ X 5 = 50 €")
        elements.append(Paragraph(
            f"Totale buoni da {valore}€ X {totale_qta_gruppo} = {totale_valore_gruppo} €",
            totale_style
        ))

    # Funzione di callback per aggiungere il numero pagina in calce a ogni pagina
    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        canvas.drawCentredString(A4[0]/2, 10*mm, f"Pagina {doc.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)
    return buffer

# ============================================================
# IMPORT EMAIL (Admin)
# ============================================================

@require_auth
def import_email(request):
    """
    Interfaccia per l'import manuale di richieste welfare via email ePassi.

    Il flusso è: incolla HTML email -> parsing automatico -> salvataggio
    in area provvisoria (RichiestaProvvisoria) -> validazione manuale ->
    spostamento in archivio definitivo (RichiestaWelfare).

    Gestisce i duplicati sia nell'archivio definitivo che in quello provvisorio.
    """
    username = request.welfare_user
    risultato_parsing = None

    if request.method == 'POST':
        azione = request.POST.get('azione')

        if azione == 'parse_html':
            html_content = request.POST.get('html_content', '')
            # Esegue il parsing dell'HTML dell'email per estrarre i dati strutturati
            risultato_parsing = parse_email_epassi(html_content)

            if risultato_parsing.get('success'):
                num_richiesta = risultato_parsing['num_richiesta']

                # Controlla duplicati nell'archivio definitivo (già processata in passato)
                if RichiestaWelfare.objects.filter(num_richiesta=num_richiesta).exists():
                    messages.warning(request, f'Richiesta {num_richiesta} già presente nel sistema')
                    risultato_parsing['errore'] = 'Duplicato in archivio definitivo'

                # Controlla duplicati nell'area provvisoria (già importata, in attesa di validazione)
                elif RichiestaProvvisoria.objects.filter(num_richiesta=num_richiesta, processato=False).exists():
                    messages.warning(request, f'Richiesta {num_richiesta} già in attesa di validazione')
                    risultato_parsing['errore'] = 'Duplicato in area provvisoria'

                else:
                    # Salva nella tabella provvisoria: verrà validata dall'admin prima di diventare definitiva
                    RichiestaProvvisoria.objects.create(
                        data_creazione=timezone.now(),
                        num_richiesta=num_richiesta,
                        nominativo=risultato_parsing.get('nominativo', ''),
                        valore_buono=str(risultato_parsing['valore_buono']),
                        qta_buono=str(risultato_parsing['quantita']),
                        totale_buono=str(risultato_parsing['totale']),
                        mittente='import_manuale',
                        nome_mittente=risultato_parsing.get('punto_vendita', ''),
                    )
                    messages.success(request, f'Richiesta {num_richiesta} importata in area provvisoria')

    # Carica tutte le provvisorie non ancora processate per mostrarle nella lista di validazione
    provvisorie = RichiestaProvvisoria.objects.filter(
        processato=False
    ).order_by('-data_importazione')

    context = {
        'username': username,
        'provvisorie': provvisorie,
        'risultato_parsing': risultato_parsing,
    }

    return render(request, 'welfare/import_email.html', context)


def parse_email_epassi(html_content):
    """
    Parser regex per email di notifica ePassi.

    Estrae i campi strutturati dal corpo HTML dell'email:
    - num_richiesta: codice a 8 cifre della richiesta
    - punto_vendita: nome del punto vendita convenzionato
    - valore_buono: valore nominale del singolo buono (intero)
    - quantita: numero di buoni richiesti (intero)
    - totale: importo totale (float, con gestione separatore decimale IT/EN)
    - nominativo: nome e cognome del beneficiario

    Restituisce un dizionario con chiave 'success': True se i campi
    obbligatori (num_richiesta, valore_buono, quantita) sono presenti,
    altrimenti 'success': False con chiave 'errore' descrittiva.
    """
    risultato = {'success': False}

    try:
        # Pattern per codice richiesta: esattamente 8 cifre dopo "Codice richiesta"
        match_codice = re.search(r'Codice richiesta[:\s]*(\d{8})', html_content, re.IGNORECASE)
        if match_codice:
            risultato['num_richiesta'] = match_codice.group(1)

        # Pattern per punto vendita: tutto il testo tra "Punto vendita:" e "- Valore"
        match_pv = re.search(r'Punto vendita[:\s]*(.+?)\s*-\s*Valore', html_content, re.IGNORECASE)
        if match_pv:
            risultato['punto_vendita'] = match_pv.group(1).strip()

        # Pattern per valore del singolo buono (numero intero di euro)
        match_valore = re.search(r'Valore del buono[:\s]*(\d+)', html_content, re.IGNORECASE)
        if match_valore:
            risultato['valore_buono'] = int(match_valore.group(1))

        # Pattern per quantità buoni (accetta sia "Quantità" che "Quantita" senza accento)
        match_qta = re.search(r'Quantit[àa][:\s]*(\d+)', html_content, re.IGNORECASE)
        if match_qta:
            risultato['quantita'] = int(match_qta.group(1))

        # Pattern per prezzo totale: gestisce sia il simbolo € che formati numerici IT (virgola decimale)
        match_totale = re.search(r'Prezzo totale[:\s]*€?\s*([\d,.]+)', html_content, re.IGNORECASE)
        if match_totale:
            # Normalizza il numero: rimuove il separatore migliaia (.) e converte la virgola decimale
            totale_str = match_totale.group(1).replace('.', '').replace(',', '.')
            risultato['totale'] = float(totale_str)

        # Pattern per il nominativo del beneficiario: testo tra "l'utente" e "ha acquistato"
        match_nom = re.search(r"l'utente\s+([^<\n]+?)\s+ha acquistato", html_content, re.IGNORECASE)
        if match_nom:
            risultato['nominativo'] = match_nom.group(1).strip()

        # Verifica che i tre campi obbligatori siano stati estratti correttamente
        if all(k in risultato for k in ['num_richiesta', 'valore_buono', 'quantita']):
            risultato['success'] = True
        else:
            risultato['errore'] = 'Campi obbligatori mancanti'

    except Exception as e:
        risultato['errore'] = str(e)

    return risultato


@csrf_exempt
@require_http_methods(["POST"])
@require_auth
def valida_provvisoria(request, pk):
    """
    Valida una richiesta provvisoria e la trasferisce nell'archivio definitivo.

    Esegue un controllo finale anti-duplicato, poi crea la RichiestaWelfare
    in stato 'PRONTO' copiando tutti i dati dalla provvisoria.
    Marca la provvisoria come processata (processato=True) per escluderla
    dalla lista di validazione.
    In caso di errore (duplicato o conversione dati) aggiorna il campo
    'errore' della provvisoria per tracciabilità.
    Risponde con JSON per la gestione asincrona da parte del frontend.
    """
    provv = get_object_or_404(RichiestaProvvisoria, pk=pk)

    try:
        # Controllo anti-duplicato: la validazione potrebbe avvenire mentre un'altra
        # sessione sta già processando la stessa richiesta
        if RichiestaWelfare.objects.filter(num_richiesta=provv.num_richiesta).exists():
            provv.errore = 'Numero richiesta già esistente'
            provv.save()
            return JsonResponse({'error': 'Duplicato'}, status=400)

        # Crea la richiesta definitiva convertendo i campi stringa in tipi numerici
        # (i dati provvisori sono salvati come stringhe per flessibilità)
        richiesta = RichiestaWelfare.objects.create(
            data_creazione=provv.data_creazione,
            num_richiesta=provv.num_richiesta,
            nominativo=provv.nominativo,
            mittente=provv.mittente,
            nome_mittente=provv.nome_mittente,
            # Conversione stringa -> Decimal con gestione separatore decimale
            valore_buono=Decimal(provv.valore_buono.replace(',', '.')) if provv.valore_buono else 0,
            qta_buono=int(provv.qta_buono) if provv.qta_buono else 0,
            totale_buono=Decimal(provv.totale_buono.replace(',', '.')) if provv.totale_buono else 0,
            stato='PRONTO',
        )

        # Marca la provvisoria come processata per escluderla dalla lista
        provv.processato = True
        provv.save()

        return JsonResponse({
            'success': True,
            'richiesta_id': richiesta.id,
            'num_richiesta': richiesta.num_richiesta
        })

    except (ValueError, InvalidOperation) as e:
        # Errore di conversione: salva il messaggio nella provvisoria per diagnostica
        provv.errore = f'Errore conversione dati: {e}'
        provv.save()
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
@require_auth
def elimina_provvisoria(request, pk):
    """
    Elimina definitivamente una richiesta provvisoria dal database.

    Usata quando il parsing ha prodotto dati errati o la richiesta
    è un duplicato da scartare. L'eliminazione è irreversibile.
    Risponde con JSON per la gestione asincrona dal frontend.
    """
    provv = get_object_or_404(RichiestaProvvisoria, pk=pk)
    provv.delete()
    return JsonResponse({'success': True})


# gestione_utenti rimossa - i ruoli sono gestiti da Angular/AD


@csrf_exempt
def api_cerca_voucher(request):
    """
    API JSON per la ricerca voucher in modalità autocomplete.

    Accetta il parametro GET 'q' (query string) e restituisce fino a 10
    risultati che corrispondono per numero richiesta o nominativo.
    Richiede almeno 3 caratteri per evitare query troppo ampie.
    Non richiede autenticazione (usata da componenti Angular pubblici).
    """
    q = request.GET.get('q', '').strip()

    # Soglia minima per evitare ricerche troppo generiche e costose
    if len(q) < 3:
        return JsonResponse({'results': []})

    # Ricerca OR: trova corrispondenze nel codice voucher o nel nominativo beneficiario
    richieste = RichiestaWelfare.objects.filter(
        Q(num_richiesta__icontains=q) | Q(nominativo__icontains=q)
    )[:10]

    # Serializza i risultati in formato JSON compatibile con i componenti Angular
    results = [{
        'id': r.id,
        'num_richiesta': r.num_richiesta,
        'nominativo': r.nominativo,
        'stato': r.stato,
        'totale': float(r.totale_buono),
    } for r in richieste]

    return JsonResponse({'results': results})


@csrf_exempt
def api_stats(request):
    """
    API JSON per le statistiche della dashboard in tempo reale.

    Restituisce i contatori principali per l'aggiornamento polling
    del widget statistiche nella dashboard Angular.
    Non richiede autenticazione.
    """
    oggi = timezone.now().date()

    stats = {
        'pronto': RichiestaWelfare.objects.filter(stato='PRONTO').count(),
        'elaborato': RichiestaWelfare.objects.filter(stato='ELABORATO').count(),
        # Consegnati oggi: filtro preciso sulla data (non sul datetime) della consegna
        'consegnati_oggi': RichiestaWelfare.objects.filter(
            stato='CONSEGNATO',
            data_consegna__date=oggi
        ).count(),
    }

    return JsonResponse(stats)

@csrf_exempt
@require_auth
def storico_consegne(request):
    """
    Storico delle consegne effettuate dal punto.info.

    Mostra solo le richieste in stato CONSEGNATO, con filtri per
    anno/mese/giorno e ricerca per nominativo. Fornisce i totali
    aggregati del periodo selezionato.
    Simile a contabilita() ma orientato alla consultazione operativa
    del personale di front-office (non genera report scaricabili).
    """
    username = request.welfare_user

    oggi = timezone.now().date()
    anno = int(request.GET.get('anno', oggi.year))
    mese = int(request.GET.get('mese', oggi.month))
    giorno_param = request.GET.get('giorno', '')
    # Giorno opzionale: None = tutti i giorni del mese
    giorno = int(giorno_param) if giorno_param else None
    filtro_nominativo = request.GET.get('nominativo', '').strip()

    # Query base: solo CONSEGNATO nel periodo anno/mese selezionato
    richieste = RichiestaWelfare.objects.filter(
        stato='CONSEGNATO',
        data_consegna__year=anno,
        data_consegna__month=mese
    )

    # Filtro giorno (opzionale): restringe al singolo giorno
    if giorno:
        richieste = richieste.filter(data_consegna__day=giorno)

    # Filtro nominativo (opzionale): ricerca parziale case-insensitive
    if filtro_nominativo:
        richieste = richieste.filter(nominativo__icontains=filtro_nominativo)

    richieste = richieste.order_by('-data_consegna')

    # Calcola i totali aggregati per i widget riepilogo
    totali = richieste.aggregate(
        num_richieste=Count('id'),
        valore_totale=Sum('totale_buono')
    )

    # Formatta il valore in stile europeo (es. 1.234,56) per la visualizzazione
    _val = float(totali['valore_totale'] or 0)
    valore_fmt = '{:,.2f}'.format(_val).replace(',', 'X').replace('.', ',').replace('X', '.')

    # Anni disponibili per il menu a tendina del filtro periodo
    anni_disponibili = RichiestaWelfare.objects.filter(
        data_consegna__isnull=False
    ).dates('data_consegna', 'year', order='DESC')

    context = {
        'username': username,
        'richieste': richieste,
        'anno': anno,
        'mese': mese,
        'giorno': giorno,
        'filtro_nominativo': filtro_nominativo,
        'totali': totali,
        'valore_fmt': valore_fmt,
        'anni_disponibili': [d.year for d in anni_disponibili],
        'mesi': range(1, 13),
        'giorni': range(1, 32),
    }

    return render(request, 'welfare/storico_consegne.html', context)
