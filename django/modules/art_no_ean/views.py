from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from modules.bidone.views import carica_annotazioni

COLONNE = [
    ('DTAAGGIO',      'Data Agg.'),
    ('SETT',          'Sett.'),
    ('REP',           'Rep.'),
    ('SREP',          'SRep.'),
    ('FAM',           'Fam.'),
    ('CCOM',          'CCom'),
    ('DESCRCCOM',     'Fornitore'),
    ('TIPOART',       'Tipo Art.'),
    ('CODART',        'Cod. Art.'),
    ('DESCRART',      'Descrizione'),
    ('STATO',         'Stato'),
    ('TIPOEAN',       'Tipo EAN'),
    ('PRINC',         'Princ.'),
    ('EAN',           'EAN'),
]

APP_NAME = 'art_no_ean'


def _esegui_query():
    sql = """
        SELECT
            DTAAGGIO, SETT, REP, SREP, FAM,
            CCOM, DESCRCCOM, TIPOART, CODART, DESCRART,
            STATO, TIPOEAN, PRINC, EAN
        FROM t_ArtNoEan
        ORDER BY SETT, REP, SREP, CODART
    """
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql)
        cols = [c[0] for c in cursor.description]
        rows = [dict(zip(cols, row)) for row in cursor.fetchall()]
    return rows


def _build_key(r):
    return str(r.get('CODART', ''))


def _merge_annotazioni(righe, annotazioni):
    for r in righe:
        key = _build_key(r)
        ann = annotazioni.get(key, {})
        r['record_key']  = key
        r['ann_gestito'] = ann.get('gestito', False)
        r['ann_nota']    = ann.get('nota', '')
        r['ann_utente']  = ann.get('utente', '')
        r['ann_data']    = ann.get('aggiornato_il')


def main(request):
    righe = _esegui_query()
    _merge_annotazioni(righe, carica_annotazioni(APP_NAME))
    return render(request, 'art_no_ean/main.html', {
        'colonne':  COLONNE,
        'righe':    righe,
        'totale':   len(righe),
        'app_name': APP_NAME,
    })


def export_excel(request):
    righe = _esegui_query()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Articoli senza EAN'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, (_, label) in enumerate(COLONNE, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, riga in enumerate(righe, 2):
        for col_idx, (campo, _) in enumerate(COLONNE, 1):
            ws.cell(row=row_idx, column=col_idx, value=riga.get(campo))

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="art_no_ean.xlsx"'
    wb.save(response)
    return response
