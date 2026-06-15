from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import math

PAGE_SIZE = 100

COLUMNS = [
    ('dtaggio',           'Data Oggi'),
    ('SETT',              'Sett.'),
    ('REP',               'Rep'),
    ('DESCR_REPARTO',     'Reparto'),
    ('SREP',              'SRep'),
    ('DESCR_SOTTOREPARTO','Sottoreparto'),
    ('FAMI',              'Fam'),
    ('DESCR_FAMIGLIA',    'Famiglia'),
    ('SITO',              'Sito'),
    ('FOP',               'FOP'),
    ('CCOM',              'CCOM'),
    ('DESCRCCOM',         'Linea Comm.'),
    ('RIOCCOM',           'RIO CCOM'),
    ('CORSIA_GOLD',       'Corsia Gold'),
    ('CAMPATA_GOLD',      'Campata Gold'),
    ('CODART',            'Cod Art'),
    ('DESCRART',          'Descrizione'),
    ('GGSCAD',            'GG Scad'),
    ('STATO',             'Stato'),
    ('VL',                'VL'),
    ('PZXCRT',            'Pz x Crt'),
    ('COLLIXSTRATO',      'Colli x Strato'),
    ('STRATIXPALLET',     'Strati x Pallet'),
    ('COLLIPALLET',       'Colli Pallet'),
    ('PESOPEZZO',         'Peso Pezzo'),
    ('PESOCOLLO',         'Peso Collo'),
    ('PZXCART',           'Pz x Cart'),
    ('QTAMINACQ',         'Qta Min Acq'),
    ('ANNO',              'Anno'),
    ('VARACQ',            'Var Acq'),
    ('NETTONETTODash',    'Netto Netto'),
    ('prz_vend',          'Prezzo Vend'),
    ('IVA',               'IVA'),
]

COL_NAMES = [c[0] for c in COLUMNS]
COL_LIST  = ', '.join(f'[{c}]' for c in COL_NAMES)


def _build_where(filters):
    clauses, params = [], []
    if filters.get('sett'):
        clauses.append('SETT = %s')
        params.append(filters['sett'])
    if filters.get('rep'):
        clauses.append('REP = %s')
        params.append(filters['rep'])
    if filters.get('ccom'):
        clauses.append('CCOM = %s')
        params.append(filters['ccom'])
    if filters.get('anno'):
        clauses.append('ANNO = %s')
        params.append(filters['anno'])
    where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
    return where, params


def get_filter_options():
    query = """
        SELECT DISTINCT SETT, REP, DESCR_REPARTO, CCOM, DESCRCCOM, ANNO
        FROM dbo.t_masterdataCategory
        WHERE SETT IS NOT NULL OR REP IS NOT NULL
        ORDER BY REP
    """
    with connections['category'].cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()

    sett_set, rep_dict, ccom_dict, anno_set = set(), {}, {}, set()
    for sett, rep, descr_rep, ccom, descrccom, anno in rows:
        if sett is not None:
            sett_set.add(sett)
        if rep is not None:
            rep_dict[rep] = descr_rep or str(rep)
        if ccom is not None:
            ccom_dict[ccom] = descrccom or str(ccom)
        if anno is not None:
            anno_set.add(anno)

    return {
        'sett':    sorted(sett_set),
        'reparti': sorted(rep_dict.items()),
        'ccom':    sorted(ccom_dict.items(), key=lambda x: x[1]),
        'anni':    sorted(anno_set, reverse=True),
    }


def get_total(filters):
    where, params = _build_where(filters)
    with connections['category'].cursor() as cursor:
        cursor.execute(f"SELECT COUNT(*) FROM dbo.t_masterdataCategory {where}", params)
        return cursor.fetchone()[0]


def get_page_rows(filters, offset):
    where, params = _build_where(filters)
    query = f"""
        SELECT {COL_LIST}
        FROM dbo.t_masterdataCategory
        {where}
        ORDER BY REP, DESCRART
        OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
    """
    with connections['category'].cursor() as cursor:
        cursor.execute(query, params + [offset, PAGE_SIZE])
        columns = [col[0] for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def get_all_rows(filters):
    where, params = _build_where(filters)
    query = f"SELECT {COL_LIST} FROM dbo.t_masterdataCategory {where} ORDER BY REP, DESCRART"
    with connections['category'].cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def masterdatacategory_list(request):
    filters = {
        'sett': request.GET.get('sett') or '',
        'rep':  request.GET.get('rep')  or '',
        'ccom': request.GET.get('ccom') or '',
        'anno': request.GET.get('anno') or '',
    }
    active_filters = {k: v for k, v in filters.items() if v}

    filter_options = get_filter_options()
    page_number    = max(1, int(request.GET.get('page', 1)))
    totale         = get_total(active_filters)
    num_pages      = math.ceil(totale / PAGE_SIZE) if totale else 1
    page_number    = min(page_number, num_pages)
    offset         = (page_number - 1) * PAGE_SIZE
    rows           = get_page_rows(active_filters, offset)

    page_range = [p for p in range(1, num_pages + 1) if abs(p - page_number) <= 3]

    return render(request, 'masterdatacategory/list.html', {
        'rows':           rows,
        'totale':         totale,
        'columns':        COLUMNS,
        'filters':        filters,
        'filter_options': filter_options,
        'page_number':    page_number,
        'num_pages':      num_pages,
        'has_prev':       page_number > 1,
        'has_next':       page_number < num_pages,
        'prev_page':      page_number - 1,
        'next_page':      page_number + 1,
        'page_range':     page_range,
    })


def masterdatacategory_export(request):
    filters = {
        'sett': request.GET.get('sett') or '',
        'rep':  request.GET.get('rep')  or '',
        'ccom': request.GET.get('ccom') or '',
        'anno': request.GET.get('anno') or '',
    }
    active_filters = {k: v for k, v in filters.items() if v}
    rows = get_all_rows(active_filters)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MasterdataCategory"

    header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_num, (field, label) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_num, item in enumerate(rows, 2):
        for col_num, (field, _) in enumerate(COLUMNS, 1):
            ws.cell(row=row_num, column=col_num, value=item.get(field))

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'masterdatacategory_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
