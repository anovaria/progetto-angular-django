from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# v_GiacenzeNegativeNoSettore2: articoli con giacenza PDV negativa (escluso settore 2)
COLONNE = [
    ('DTAAGGIO',    'Data Agg.'),
    ('settore',     'Settore'),
    ('reparto',     'Reparto'),
    ('sottoRep',    'Sotto Rep.'),
    ('Codart',      'Cod. Art.'),
    ('DescrArt',    'Descrizione'),
    ('stato',       'Stato'),
    ('GiacPDV',     'Giac. PDV'),
]


def _esegui_query():
    # Le colonne "descrizione articolo" e "Giac PDV" hanno spazi nel nome:
    # le aliasiamo per evitare problemi nel template.
    sql = """
        SELECT
            DTAAGGIO,
            settore,
            reparto,
            sottoRep,
            Codart,
            [descrizione articolo] AS DescrArt,
            stato,
            [Giac PDV]             AS GiacPDV
        FROM v_GiacenzeNegativeNoSettore2
        ORDER BY settore, reparto, sottoRep, Codart
    """
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(sql)
        cols = [c[0] for c in cursor.description]
        rows = [dict(zip(cols, row)) for row in cursor.fetchall()]
    return rows


def main(request):
    righe = _esegui_query()
    ctx = {
        'colonne': COLONNE,
        'righe': righe,
        'totale': len(righe),
    }
    return render(request, 'giac_negative/main.html', ctx)


def export_excel(request):
    righe = _esegui_query()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Giacenze Negative'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, (_, label) in enumerate(COLONNE, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, riga in enumerate(righe, 2):
        for col_idx, (campo, _) in enumerate(COLONNE, 1):
            ws.cell(row=row_idx, column=col_idx, value=riga.get(campo))

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="giacenze_negative.xlsx"'
    wb.save(response)
    return response
