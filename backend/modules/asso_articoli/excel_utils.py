"""
Utility per export Excel con barcode embedded
Genera file Excel con formattazione e immagini barcode
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from .barcode_utils import generate_ean13_png, normalize_ean_for_barcode


def export_articoli_excel(articoli_data, filename='assortimenti', include_barcode=True):
    """
    Esporta lista articoli in Excel con formattazione
    
    Args:
        articoli_data: Lista di dict con dati articoli
        filename: Nome file (senza estensione)
        include_barcode: Se True, include colonna con barcode
    
    Returns:
        BytesIO: Buffer con file Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Assortimenti"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Codice Articolo',
        'Descrizione',
        'Settore',
        'Reparto',
        'Sotto-Reparto',
        'Famiglia',
        'CCOM',
        'Linea Prodotto',
        'CodArtFo',  # AGGIUNTO
        'Stato',
        'EAN',
    ]
    
    if include_barcode:
        headers.append('Barcode')
    
    headers.extend([
        'Giacenza PDV',
        'Giacenza Deposito',
        'Prezzo Acquisto',
        'IVA %',
        'Fornitore',
    ])
    
    # Scrivi headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin
    
    # Altezza header
    ws.row_dimensions[1].height = 25
    
    # Dati
    row_num = 2
    barcode_col = None
    
    for articolo in articoli_data:
        col_num = 1
        
        # Dati base
        ws.cell(row=row_num, column=col_num, value=articolo.get('codart', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('descrart', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('sett', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('rep', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('srep', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('fam', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('ccom', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('linea_prodotto', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('codartfo', '')).border = border_thin
        col_num += 1

        ws.cell(row=row_num, column=col_num, value=articolo.get('stato', '')).border = border_thin
        col_num += 1
        
        # EAN
        ean_value = articolo.get('ean', '')
        ws.cell(row=row_num, column=col_num, value=ean_value).border = border_thin
        col_num += 1
        
        # Barcode
        if include_barcode:
            barcode_col = col_num
            ean_value = articolo.get('ean', '')  # prendi EAN qui
            # Genera barcode PNG
            tipo_ean = articolo.get('tipoean')
            ean_normalized = normalize_ean_for_barcode(ean_value, tipo_ean)
            
        if ean_normalized:
            barcode_buffer = generate_ean13_png(ean_normalized)
            
            if barcode_buffer:
                img = XLImage(barcode_buffer)
                img.width = 150
                img.height = 70
                
                cell_address = f"{get_column_letter(col_num)}{row_num}"
                ws.add_image(img, cell_address)
                
                ws.row_dimensions[row_num].height = 55
        
            col_num += 1
        
        # Giacenze
        giacenza_pdv = articolo.get('giacenza_pdv', 0)
        cell = ws.cell(row=row_num, column=col_num, value=giacenza_pdv)
        cell.number_format = '#,##0.00'
        cell.border = border_thin
        col_num += 1
        
        giacenza_dep = articolo.get('giacenza_deposito', 0)
        cell = ws.cell(row=row_num, column=col_num, value=giacenza_dep)
        cell.number_format = '#,##0.00'
        cell.border = border_thin
        col_num += 1
        
        # Prezzo acquisto
        pracq = articolo.get('pracq', 0)
        cell = ws.cell(row=row_num, column=col_num, value=pracq)
        cell.number_format = '€#,##0.00'
        cell.border = border_thin
        col_num += 1
        
        # IVA
        iva = articolo.get('iva', 0)
        cell = ws.cell(row=row_num, column=col_num, value=iva)
        cell.number_format = '0.00"%"'
        cell.border = border_thin
        col_num += 1
        
        # Fornitore
        ws.cell(row=row_num, column=col_num, value=articolo.get('descforn', '')).border = border_thin
        
        row_num += 1
    
    # Auto-width per colonne (tranne barcode)
    for col_num in range(1, len(headers) + 1):
        if include_barcode and col_num == barcode_col:
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        else:
            ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Colonna descrizione più larga
    ws.column_dimensions['B'].width = 40
    
    # Freeze panes (header)
    ws.freeze_panes = 'A2'
    
    # Salva in buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def export_report_inventario(articoli_data, ccom=None, linea_prodotto=None):
    """
    Export specifico per report inventario (simile a r_StampaInv)
    
    Args:
        articoli_data: Lista articoli
        ccom: Filtro CCOM applicato
        linea_prodotto: Filtro linea prodotto
    
    Returns:
        BytesIO: Buffer Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Header report
    ws['A1'] = 'REPORT INVENTARIO ASSORTIMENTI'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    
    # Info filtri
    row = 2
    if ccom:
        ws[f'A{row}'] = f'CCOM: {ccom}'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    if linea_prodotto:
        ws[f'A{row}'] = f'Linea Prodotto: {linea_prodotto}'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    ws[f'A{row}'] = f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    row += 2
    
    # Headers tabella
    headers_row = row
    headers = ['Cod.Art', 'Descrizione', 'EAN', 'Barcode', 'Giac.PDV', 'Giac.Dep']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=headers_row, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Dati
    data_row = headers_row + 1
    for articolo in articoli_data:
        ws.cell(row=data_row, column=1, value=articolo.get('codart', ''))
        ws.cell(row=data_row, column=2, value=articolo.get('descrart', ''))
        ws.cell(row=data_row, column=3, value=articolo.get('ean', ''))
        
        # Barcode
        ean = articolo.get('ean', '')
        tipo_ean = articolo.get('tipoean')
        ean_normalized = normalize_ean_for_barcode(ean, tipo_ean)
        
        if ean_normalized:
            barcode_buffer = generate_ean13_png(ean_normalized)
            if barcode_buffer:
                img = XLImage(barcode_buffer)
                img.width = 100
                img.height = 50
                ws.add_image(img, f'D{data_row}')
                ws.row_dimensions[data_row].height = 40
        
        ws.cell(row=data_row, column=5, value=articolo.get('giacenza_pdv', 0))
        ws.cell(row=data_row, column=6, value=articolo.get('giacenza_deposito', 0))
        
        data_row += 1
    
    # Auto-width
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Salva
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
def export_report_reparti(articoli_data, tipo_reparto=''):
    """
    Export specifico per report reparti
    Ordinato per fornitore con riga intestazione ad ogni cambio
    """
    wb = Workbook()
    ws = wb.active
    ws.title = tipo_reparto or "Reparto"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    fornitore_font = Font(bold=True, size=12, color="1F4E79")
    fornitore_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers (senza Giacenza Deposito, con Note)
    headers = [
        'Codice Articolo',
        'Descrizione',
        'Famiglia',
        'Desc. Famiglia',
        'CodArtFo',
        'Stato',
        'EAN',
        'Barcode',
        'Giacenza PDV',
        'Note',
    ]
    
    # Scrivi headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin
    
    ws.row_dimensions[1].height = 25
    
    # Ordina per fornitore
    articoli_data_sorted = sorted(articoli_data, key=lambda x: (x.get('descforn') or '', x.get('fam') or '', x.get('codart') or ''))
    
    # Dati
    row_num = 2
    barcode_col = 8
    current_fornitore = None
    
    for articolo in articoli_data_sorted:
        fornitore = articolo.get('descforn') or 'Senza Fornitore'
        
        # Riga intestazione fornitore al cambio
        if fornitore != current_fornitore:
            current_fornitore = fornitore
            
            # Riga fornitore
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
            cell = ws.cell(row=row_num, column=1, value=f"▶ FORNITORE: {fornitore}")
            cell.font = fornitore_font
            cell.fill = fornitore_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_num].height = 22
            row_num += 1
        
        col_num = 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('codart', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('descrart', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('fam', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('descrfam', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('codartfo', '')).border = border_thin
        col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=articolo.get('stato', '')).border = border_thin
        col_num += 1
        
        # EAN
        ean_value = articolo.get('ean', '')
        ws.cell(row=row_num, column=col_num, value=ean_value).border = border_thin
        col_num += 1
        
        # Barcode
        tipo_ean = articolo.get('tipoean')
        ean_normalized = normalize_ean_for_barcode(ean_value, tipo_ean)
        
        if ean_normalized:
            barcode_buffer = generate_ean13_png(ean_normalized)
            
            if barcode_buffer:
                img = XLImage(barcode_buffer)
                img.width = 150
                img.height = 70
                
                cell_address = f"{get_column_letter(col_num)}{row_num}"
                ws.add_image(img, cell_address)
                
                ws.row_dimensions[row_num].height = 55
        
        col_num += 1
        
        # Giacenza PDV
        giacenza_pdv = articolo.get('giacenza_pdv', 0)
        cell = ws.cell(row=row_num, column=col_num, value=giacenza_pdv)
        cell.number_format = '#,##0.00'
        cell.border = border_thin
        col_num += 1
        
        # Note (vuoto con bordo per compilazione manuale)
        cell = ws.cell(row=row_num, column=col_num, value='')
        cell.border = border_thin
        
        row_num += 1
    
    # Larghezza colonne
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 25  # Note più larga per scrivere
    
    ws.freeze_panes = 'A2'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer