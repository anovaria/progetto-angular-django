"""
Views Django per la gestione Welfare Aziendale
"""

from io import BytesIO
import zipfile
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.contrib import messages
from decimal import Decimal, InvalidOperation
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from .models import (
    TaglioBuono, RichiestaWelfare, 
    DettaglioBuono, RichiestaProvvisoria
)


# ============================================================
# AUTENTICAZIONE (semplificata - ruoli gestiti da Angular)
# ============================================================

def get_current_user(request):
    """Recupera l'utente corrente dall'header LDAP."""
    username = request.META.get('HTTP_X_REMOTE_USER', '')
    if username:
        return username.split('\\')[-1].lower()
    return 'anonymous'


def require_auth(view_func):
    """Decorator - non blocca, passa sempre l'utente."""
    def wrapper(request, *args, **kwargs):
        request.welfare_user = get_current_user(request)
        return view_func(request, *args, **kwargs)
    return wrapper

# ============================================================
# DASHBOARD PRINCIPALE
# ============================================================

@csrf_exempt
@require_auth
def dashboard(request):
    """Dashboard principale con statistiche e accesso rapido."""
    username = request.welfare_user
    
    # Statistiche generali
    oggi = timezone.now().date()
    inizio_mese = oggi.replace(day=1)
    
    stats = {
        'totale_pronto': RichiestaWelfare.objects.filter(stato='PRONTO').count(),
        'totale_elaborato': RichiestaWelfare.objects.filter(stato='ELABORATO').count(),
        'consegnati_oggi': RichiestaWelfare.objects.filter(
            stato='CONSEGNATO',
            data_consegna__date=oggi
        ).count(),
        'consegnati_mese': RichiestaWelfare.objects.filter(
            stato='CONSEGNATO',
            data_consegna__date__gte=inizio_mese
        ).count(),
        'valore_consegnati_mese': RichiestaWelfare.objects.filter(
            stato='CONSEGNATO',
            data_consegna__date__gte=inizio_mese
        ).aggregate(tot=Sum('totale_buono'))['tot'] or 0,
        'provvisori_da_validare': RichiestaProvvisoria.objects.filter(
            processato=False
        ).count(),
    }
    
    # Ultime richieste pronte
    ultime_pronte = RichiestaWelfare.objects.filter(
        stato='PRONTO'
    ).order_by('-data_creazione')[:10]
    
    context = {
        'username': username,
        'stats': stats,
        'ultime_pronte': ultime_pronte,
        'oggi': oggi,
    }
    
    return render(request, 'welfare/dashboard.html', context)


# ============================================================
# RICERCA VOUCHER (Front-Office e Cassa)
# ============================================================

@csrf_exempt
@require_auth
def ricerca_voucher(request):
    """Ricerca richiesta per numero voucher."""
    username = request.welfare_user
    
    richiesta = None
    dettagli = []
    num_voucher = request.GET.get('voucher', '').strip()
    
    if num_voucher:
        try:
            richiesta = RichiestaWelfare.objects.get(num_richiesta=num_voucher)
            dettagli = richiesta.dettagli_buoni.select_related('taglio').all()
        except RichiestaWelfare.DoesNotExist:
            messages.warning(request, f'Voucher {num_voucher} non trovato')
    
    context = {
        'username': username,
        'num_voucher': num_voucher,
        'richiesta': richiesta,
        'dettagli': dettagli,
    }
    
    return render(request, 'welfare/ricerca_voucher.html', context)


# ============================================================
# GESTIONE CASSA
# ============================================================

@csrf_exempt
@require_auth
def cassa_consegna(request):
    """Interfaccia per registrare consegne e specificare tagli."""
    username = request.welfare_user
    
    richiesta = None
    dettagli = []
    tagli_disponibili = TaglioBuono.objects.filter(attivo=True)
    num_voucher = request.GET.get('voucher', '').strip()
    
    if num_voucher:
        try:
            richiesta = RichiestaWelfare.objects.get(
                num_richiesta=num_voucher,
                stato='PRONTO'
            )
            dettagli = richiesta.dettagli_buoni.select_related('taglio').all()
        except RichiestaWelfare.DoesNotExist:
            messages.warning(request, f'Voucher {num_voucher} non trovato o non in stato PRONTO')
    
    if request.method == 'POST':
        richiesta_id = request.POST.get('richiesta_id')
        azione = request.POST.get('azione')
        
        if richiesta_id:
            richiesta = get_object_or_404(RichiestaWelfare, pk=richiesta_id)
            
            if azione == 'consegna':
                richiesta.segna_consegnato('punto.info')
                messages.success(request, f'Voucher {richiesta.num_richiesta} consegnato!')
                return redirect('welfare:cassa_consegna')
            
            elif azione == 'salva_tagli':
                DettaglioBuono.objects.filter(richiesta=richiesta).delete()
                
                for taglio in tagli_disponibili:
                    qta_key = f'qta_{taglio.valore_nominale}'
                    qta = request.POST.get(qta_key, '0')
                    try:
                        qta = int(qta)
                        if qta > 0:
                            DettaglioBuono.objects.create(
                                richiesta=richiesta,
                                taglio=taglio,
                                quantita=qta
                            )
                    except (ValueError, TypeError):
                        pass
                
                messages.success(request, 'Dettaglio tagli salvato')
                return redirect(f'?voucher={richiesta.num_richiesta}')
    
    context = {
        'username': username,
        'num_voucher': num_voucher,
        'richiesta': richiesta,
        'dettagli': dettagli,
        'tagli_disponibili': tagli_disponibili,
    }
    
    return render(request, 'welfare/cassa_consegna.html', context)


# ============================================================
# LISTA BUONI DA CONSEGNARE
# ============================================================

@csrf_exempt
@require_auth
def lista_da_consegnare(request):
    """Lista di tutti i buoni in stato PRONTO."""
    username = request.welfare_user
    
    richieste = RichiestaWelfare.objects.filter(
        stato='PRONTO'
    ).order_by('data_creazione')
    
    # Filtri
    nominativo = request.GET.get('nominativo', '').strip()
    if nominativo:
        richieste = richieste.filter(nominativo__icontains=nominativo)
    
    azienda = request.GET.get('azienda', '').strip()
    if azienda:
        richieste = richieste.filter(nome_mittente__icontains=azienda)
    
    totale_valore = richieste.aggregate(tot=Sum('totale_buono'))['tot'] or 0
    
    context = {
        'username': username,
        'richieste': richieste,
        'filtro_nominativo': nominativo,
        'filtro_azienda': azienda,
        'totale_valore': totale_valore,
    }
    
    return render(request, 'welfare/lista_da_consegnare.html', context)


# ============================================================
# CONTABILITÀ E REPORT
# ============================================================

@csrf_exempt
@require_auth
def contabilita(request):
    """Interfaccia Contabilità con filtri avanzati e report."""
    username = request.welfare_user
    
    oggi = timezone.now().date()
    anno = int(request.GET.get('anno', oggi.year))
    mese = int(request.GET.get('mese', oggi.month))
    giorno_param = request.GET.get('giorno', '')
    giorno = int(giorno_param) if giorno_param else None
    
    mostra_pronto = request.GET.get('pronto', '') == '1'
    mostra_elaborato = request.GET.get('elaborato', '') == '1'
    mostra_consegnato = request.GET.get('consegnato', '1') == '1'
    mostra_inevaso = request.GET.get('inevaso', '') == '1'
    
    stati_filtro = []
    if mostra_pronto:
        stati_filtro.append('PRONTO')
    if mostra_elaborato:
        stati_filtro.append('ELABORATO')
    if mostra_consegnato:
        stati_filtro.append('CONSEGNATO')
    if mostra_inevaso:
        stati_filtro.append('INEVASO')
    
    if not stati_filtro:
        stati_filtro = ['CONSEGNATO']
    
    richieste = RichiestaWelfare.objects.filter(
        stato__in=stati_filtro,
        data_consegna__year=anno,
        data_consegna__month=mese
    )
    
    # Filtro giorno (opzionale)
    if giorno:
        richieste = richieste.filter(data_consegna__day=giorno)
    
    richieste = richieste.order_by('-data_consegna')
    
    totali = richieste.aggregate(
        num_richieste=Count('id'),
        valore_totale=Sum('totale_buono')
    )
    
    anni_disponibili = RichiestaWelfare.objects.filter(
        data_consegna__isnull=False
    ).dates('data_consegna', 'year', order='DESC')
    
    context = {
        'username': username,
        'richieste': richieste,
        'anno': anno,
        'mese': mese,
        'giorno': giorno,
        'mostra_pronto': mostra_pronto,
        'mostra_elaborato': mostra_elaborato,
        'mostra_consegnato': mostra_consegnato,
        'mostra_inevaso': mostra_inevaso,
        'totali': totali,
        'anni_disponibili': [d.year for d in anni_disponibili],
        'mesi': range(1, 13),
        'giorni': range(1, 32),
    }
    
    return render(request, 'welfare/contabilita.html', context)


@csrf_exempt
@require_auth
def report_contabilita(request):
    """Genera report Excel + PDF in ZIP con i filtri applicati."""
    
    # Leggi filtri
    oggi = timezone.now().date()
    anno = int(request.GET.get('anno', oggi.year))
    mese = int(request.GET.get('mese', oggi.month))
    giorno_param = request.GET.get('giorno', '')
    giorno = int(giorno_param) if giorno_param else None
    
    # Stati filtro
    mostra_pronto = request.GET.get('pronto', '') == '1'
    mostra_elaborato = request.GET.get('elaborato', '') == '1'
    mostra_consegnato = request.GET.get('consegnato', '1') == '1'
    mostra_inevaso = request.GET.get('inevaso', '') == '1'
    
    stati_filtro = []
    if mostra_pronto:
        stati_filtro.append('PRONTO')
    if mostra_elaborato:
        stati_filtro.append('ELABORATO')
    if mostra_consegnato:
        stati_filtro.append('CONSEGNATO')
    if mostra_inevaso:
        stati_filtro.append('INEVASO')
    
    if not stati_filtro:
        stati_filtro = ['CONSEGNATO']
    
    # Query
    richieste = RichiestaWelfare.objects.filter(
        stato__in=stati_filtro,
        data_consegna__year=anno,
        data_consegna__month=mese
    )
    
    if giorno:
        richieste = richieste.filter(data_consegna__day=giorno)
    
    richieste = richieste.order_by('data_consegna')
    
    # Nome file base
    if giorno:
        nome_base = f"Welfare_{anno}_{mese:02d}_{giorno:02d}"
    else:
        nome_base = f"Welfare_{anno}_{mese:02d}"
    
    # Genera Excel
    excel_buffer = genera_excel_report(richieste, mese, anno)
    
    # Genera PDF
    pdf_buffer = genera_pdf_report(richieste, mese, anno)
    
    # Crea ZIP
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.writestr(f"{nome_base}.xlsx", excel_buffer.getvalue())
        zip_file.writestr(f"{nome_base}.pdf", pdf_buffer.getvalue())
    
    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{nome_base}.zip"'
    return response


def genera_excel_report(richieste, mese, anno):
    """Genera il file Excel nel formato richiesto."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo"
    
    # Stili
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    
    # Intestazione
    ws['A1'] = f"Riepilogo Buoni Welfare - Mese: {mese} Anno: {anno}"
    ws['A1'].font = header_font
    ws.merge_cells('A1:F1')
    
    # Header colonne
    headers = ['TotaleBuono', 'Nominativo', 'Num_Richiesta', 'QtaBuono', 'ValoreBuono', 'Data_ConsegnaCliente']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = title_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Dati
    row = 4
    totale_generale = 0
    totale_qta = 0
    
    for r in richieste:
        ws.cell(row=row, column=1, value=float(r.totale_buono)).border = border
        ws.cell(row=row, column=2, value=r.nominativo).border = border
        ws.cell(row=row, column=3, value=r.num_richiesta).border = border
        ws.cell(row=row, column=4, value=r.qta_buono).border = border
        ws.cell(row=row, column=5, value=float(r.valore_buono)).border = border
        data_cons = r.data_consegna.strftime('%d/%m/%Y') if r.data_consegna else ''
        ws.cell(row=row, column=6, value=data_cons).border = border
        
        totale_generale += float(r.totale_buono)
        totale_qta += r.qta_buono
        row += 1
    
    # Riga totale
    row += 1
    ws.cell(row=row, column=1, value=totale_generale).font = Font(bold=True, size=12)
    ws.cell(row=row, column=4, value=totale_qta).font = Font(bold=True, size=12)
    
    # Larghezza colonne
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def genera_pdf_report(richieste, mese, anno):
    """Genera il file PDF nel formato richiesto."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER
    
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        topMargin=15*mm, 
        bottomMargin=20*mm,
        leftMargin=15*mm,
        rightMargin=15*mm
    )
    
    elements = []
    
    # Titolo
    title_style = ParagraphStyle(
        'Title',
        fontName='Helvetica-Bold',
        fontSize=12,
        alignment=TA_CENTER,
        spaceAfter=10*mm
    )
    elements.append(Paragraph(f"Riepilogo Buoni Wellfare Mese: {mese} Anno: {anno}", title_style))
    
    # Stili per le righe
    riga1_style = ParagraphStyle(
        'Riga1',
        fontName='Helvetica',
        fontSize=10,
        leading=12
    )
    
    riga2_style = ParagraphStyle(
        'Riga2',
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        leftIndent=10*mm
    )
    
    totale_style = ParagraphStyle(
        'Totale',
        fontName='Helvetica-Bold',
        fontSize=10,
        spaceBefore=3*mm,
        spaceAfter=8*mm
    )
    
    # Raggruppa per valore buono
    richieste_list = list(richieste)
    valori_buono = {}
    for r in richieste_list:
        valore = int(r.valore_buono)
        if valore not in valori_buono:
            valori_buono[valore] = []
        valori_buono[valore].append(r)
    
    # Per ogni gruppo di valore buono
    for valore in sorted(valori_buono.keys(), reverse=True):
        gruppo = valori_buono[valore]
        totale_qta_gruppo = 0
        totale_valore_gruppo = 0
        
        for r in gruppo:
            data_cons = r.data_consegna.strftime('%d/%m/%Y') if r.data_consegna else '-'
            totale = int(r.totale_buono)
            qta = r.qta_buono
            
            # Riga 1: NOMINATIVO      N° RICHIESTA xxxxxxxx  Per un totale di XXX €
            riga1 = f"<b>{r.nominativo}</b> &nbsp;&nbsp;&nbsp;&nbsp;&nbsp; N ° RICHIESTA <b>{r.num_richiesta}</b> &nbsp;&nbsp; Per un totale di <b>{totale} €</b>"
            
            # Riga 2: DETTAGLIO: N°X BUONI Da XX € --- Data Cons Cliente xx/xx/xxxx
            riga2 = f"DETTAGLIO: N°{qta} BUONI Da {valore} € --- Data Cons Cliente {data_cons}"
            
            # Box con bordo
            box_content = [
                [Paragraph(riga1, riga1_style)],
                [Paragraph(riga2, riga2_style)],
            ]
            
            box = Table(box_content, colWidths=[170*mm])
            box.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
                ('TOPPADDING', (0, 0), (-1, -1), 3*mm),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3*mm),
                ('LEFTPADDING', (0, 0), (-1, -1), 3*mm),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3*mm),
            ]))
            
            elements.append(box)
            elements.append(Spacer(1, 2*mm))
            
            totale_qta_gruppo += qta
            totale_valore_gruppo += totale
        
        # Totale per valore buono
        elements.append(Paragraph(
            f"Totale buoni da {valore}€ X {totale_qta_gruppo} = {totale_valore_gruppo} €",
            totale_style
        ))
    
    # Numero pagina
    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        canvas.drawCentredString(A4[0]/2, 10*mm, f"Pagina {doc.page}")
        canvas.restoreState()
    
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)
    return buffer
# ============================================================
# IMPORT EMAIL (Admin)
# ============================================================

@csrf_exempt
@require_auth
def import_email(request):
    """Interfaccia per import email."""
    username = request.welfare_user
    risultato_parsing = None
    
    if request.method == 'POST':
        azione = request.POST.get('azione')
        
        if azione == 'parse_html':
            html_content = request.POST.get('html_content', '')
            risultato_parsing = parse_email_eudaimon(html_content)
            
            if risultato_parsing.get('success'):
                num_richiesta = risultato_parsing['num_richiesta']
                
                # Controlla duplicati in richieste definitive
                if RichiestaWelfare.objects.filter(num_richiesta=num_richiesta).exists():
                    messages.warning(request, f'Richiesta {num_richiesta} già presente nel sistema')
                    risultato_parsing['errore'] = 'Duplicato in archivio definitivo'
                
                # Controlla duplicati in richieste provvisorie non processate
                elif RichiestaProvvisoria.objects.filter(num_richiesta=num_richiesta, processato=False).exists():
                    messages.warning(request, f'Richiesta {num_richiesta} già in attesa di validazione')
                    risultato_parsing['errore'] = 'Duplicato in area provvisoria'
                
                else:
                    RichiestaProvvisoria.objects.create(
                        data_creazione=timezone.now(),
                        num_richiesta=num_richiesta,
                        nominativo=risultato_parsing.get('nominativo', ''),
                        valore_buono=str(risultato_parsing['valore_buono']),
                        qta_buono=str(risultato_parsing['quantita']),
                        totale_buono=str(risultato_parsing['totale']),
                        mittente='import_manuale',
                        nome_mittente=risultato_parsing.get('punto_vendita', ''),
                    )
                    messages.success(request, f'Richiesta {num_richiesta} importata in area provvisoria')
    
    provvisorie = RichiestaProvvisoria.objects.filter(
        processato=False
    ).order_by('-data_importazione')
    
    context = {
        'username': username,
        'provvisorie': provvisorie,
        'risultato_parsing': risultato_parsing,
    }
    
    return render(request, 'welfare/import_email.html', context)


def parse_email_eudaimon(html_content):
    """
    Parser per email Eudaimon.
    Estrae: codice richiesta, punto vendita, valore, quantita, totale.
    """
    risultato = {'success': False}
    
    try:
        # Pattern per codice richiesta (8 cifre)
        match_codice = re.search(r'Codice richiesta[:\s]*(\d{8})', html_content, re.IGNORECASE)
        if match_codice:
            risultato['num_richiesta'] = match_codice.group(1)
        
        # Pattern per punto vendita (prende tutto fino a "- Valore")
        match_pv = re.search(r'Punto vendita[:\s]*(.+?)\s*-\s*Valore', html_content, re.IGNORECASE)
        if match_pv:
            risultato['punto_vendita'] = match_pv.group(1).strip()
        
        # Pattern per valore buono
        match_valore = re.search(r'Valore del buono[:\s]*(\d+)', html_content, re.IGNORECASE)
        if match_valore:
            risultato['valore_buono'] = int(match_valore.group(1))
        
        # Pattern per quantità
        match_qta = re.search(r'Quantit[àa][:\s]*(\d+)', html_content, re.IGNORECASE)
        if match_qta:
            risultato['quantita'] = int(match_qta.group(1))
        
        # Pattern per prezzo totale
        match_totale = re.search(r'Prezzo totale[:\s]*€?\s*([\d,.]+)', html_content, re.IGNORECASE)
        if match_totale:
            totale_str = match_totale.group(1).replace('.', '').replace(',', '.')
            risultato['totale'] = float(totale_str)
        
        # Pattern per nominativo (dall'utente)
        match_nom = re.search(r"l'utente\s+([^<\n]+?)\s+ha acquistato", html_content, re.IGNORECASE)
        if match_nom:
            risultato['nominativo'] = match_nom.group(1).strip()
        
        # Verifica campi obbligatori
        if all(k in risultato for k in ['num_richiesta', 'valore_buono', 'quantita']):
            risultato['success'] = True
        else:
            risultato['errore'] = 'Campi obbligatori mancanti'
    
    except Exception as e:
        risultato['errore'] = str(e)
    
    return risultato


@csrf_exempt
@require_http_methods(["POST"])
@require_auth
def valida_provvisoria(request, pk):
    """Valida una richiesta provvisoria e la sposta in definitiva."""
    provv = get_object_or_404(RichiestaProvvisoria, pk=pk)
    
    try:
        if RichiestaWelfare.objects.filter(num_richiesta=provv.num_richiesta).exists():
            provv.errore = 'Numero richiesta già esistente'
            provv.save()
            return JsonResponse({'error': 'Duplicato'}, status=400)
        
        richiesta = RichiestaWelfare.objects.create(
            data_creazione=provv.data_creazione,
            num_richiesta=provv.num_richiesta,
            nominativo=provv.nominativo,
            mittente=provv.mittente,
            nome_mittente=provv.nome_mittente,
            valore_buono=Decimal(provv.valore_buono.replace(',', '.')) if provv.valore_buono else 0,
            qta_buono=int(provv.qta_buono) if provv.qta_buono else 0,
            totale_buono=Decimal(provv.totale_buono.replace(',', '.')) if provv.totale_buono else 0,
            stato='PRONTO',
        )
        
        provv.processato = True
        provv.save()
        
        return JsonResponse({
            'success': True,
            'richiesta_id': richiesta.id,
            'num_richiesta': richiesta.num_richiesta
        })
    
    except (ValueError, InvalidOperation) as e:
        provv.errore = f'Errore conversione dati: {e}'
        provv.save()
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
@require_auth
def elimina_provvisoria(request, pk):
    """Elimina una richiesta provvisoria."""
    provv = get_object_or_404(RichiestaProvvisoria, pk=pk)
    provv.delete()
    return JsonResponse({'success': True})


# gestione_utenti rimossa - i ruoli sono gestiti da Angular/AD


@csrf_exempt
def api_cerca_voucher(request):
    """API per ricerca voucher."""
    q = request.GET.get('q', '').strip()
    
    if len(q) < 3:
        return JsonResponse({'results': []})
    
    richieste = RichiestaWelfare.objects.filter(
        Q(num_richiesta__icontains=q) | Q(nominativo__icontains=q)
    )[:10]
    
    results = [{
        'id': r.id,
        'num_richiesta': r.num_richiesta,
        'nominativo': r.nominativo,
        'stato': r.stato,
        'totale': float(r.totale_buono),
    } for r in richieste]
    
    return JsonResponse({'results': results})


@csrf_exempt
def api_stats(request):
    """API per statistiche dashboard."""
    oggi = timezone.now().date()
    
    stats = {
        'pronto': RichiestaWelfare.objects.filter(stato='PRONTO').count(),
        'elaborato': RichiestaWelfare.objects.filter(stato='ELABORATO').count(),
        'consegnati_oggi': RichiestaWelfare.objects.filter(
            stato='CONSEGNATO',
            data_consegna__date=oggi
        ).count(),
    }
    
    return JsonResponse(stats)

@csrf_exempt
@require_auth
def storico_consegne(request):
    """Storico consegne per punto.info - solo CONSEGNATO."""
    username = request.welfare_user
    
    oggi = timezone.now().date()
    anno = int(request.GET.get('anno', oggi.year))
    mese = int(request.GET.get('mese', oggi.month))
    giorno_param = request.GET.get('giorno', '')
    giorno = int(giorno_param) if giorno_param else None
    filtro_nominativo = request.GET.get('nominativo', '').strip()
    
    richieste = RichiestaWelfare.objects.filter(
        stato='CONSEGNATO',
        data_consegna__year=anno,
        data_consegna__month=mese
    )
    
    # Filtro giorno (opzionale)
    if giorno:
        richieste = richieste.filter(data_consegna__day=giorno)
    
    # Filtro nominativo (opzionale)
    if filtro_nominativo:
        richieste = richieste.filter(nominativo__icontains=filtro_nominativo)
    
    richieste = richieste.order_by('-data_consegna')
    
    totali = richieste.aggregate(
        num_richieste=Count('id'),
        valore_totale=Sum('totale_buono')
    )
    
    anni_disponibili = RichiestaWelfare.objects.filter(
        data_consegna__isnull=False
    ).dates('data_consegna', 'year', order='DESC')
    
    context = {
        'username': username,
        'richieste': richieste,
        'anno': anno,
        'mese': mese,
        'giorno': giorno,
        'filtro_nominativo': filtro_nominativo,
        'totali': totali,
        'anni_disponibili': [d.year for d in anni_disponibili],
        'mesi': range(1, 13),
        'giorni': range(1, 32),
    }
    
    return render(request, 'welfare/storico_consegne.html', context)