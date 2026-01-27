# modules/plu/views.py
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.db import connections
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def execute_plu_query(filters=None):
    """
    Esegue la query PLU corretta (stessa logica dell'Excel dei colleghi)
    Con JOIN a t_Reparti per nome reparto
    Solo lettura, nessuna modifica al DB
    """
    base_query = """
    WITH codvel AS (
        SELECT DISTINCT 
            dbo.t_AnagArticoli.REP, 
            dbo.t_t_Ean.CODART, 
            dbo.t_AnagArticoli.DESCRART, 
            dbo.t_t_Ean.EAN
        FROM dbo.t_t_Ean 
        INNER JOIN dbo.t_AnagArticoli ON dbo.t_t_Ean.CODART = dbo.t_AnagArticoli.CODART
        WHERE dbo.t_t_Ean.TIPO = 5 AND dbo.t_AnagArticoli.EANPRINC = 0
    ), 
    codean AS (
        SELECT DISTINCT 
            t_t_Ean_1.CODART, 
            t_t_Ean_1.EAN
        FROM dbo.t_t_Ean AS t_t_Ean_1 
        INNER JOIN dbo.t_AnagArticoli AS t_AnagArticoli_1 ON t_t_Ean_1.CODART = t_AnagArticoli_1.CODART
        WHERE t_t_Ean_1.TIPO NOT IN (5) AND t_t_Ean_1.PRINC = 1 AND t_AnagArticoli_1.EANPRINC = 1
    )
    SELECT DISTINCT 
        m.CCOM as ccom, 
        m.DESCRCCOM as descrccom, 
        a.REP as rep,
        r.RepDescrizione as reparto_nome,
        a.CODART AS codArticolo, 
        a.DESCRART AS descrizione, 
        SUBSTRING(a.EAN, 1, 2) AS bancobilancia, 
        SUBSTRING(a.EAN, 3, 7) AS plu, 
        b.EAN as ean
    FROM codvel AS a 
    INNER JOIN codean AS b ON a.CODART = b.CODART 
    LEFT OUTER JOIN dbo.t_masterData AS m ON a.CODART = m.CODART
    LEFT OUTER JOIN dbo.t_Reparti AS r ON a.REP = r.NrReparto
    WHERE 1=1
    """
    
    params = []
    
    if filters:
        if filters.get('reparto'):
            base_query += " AND a.REP = %s"
            params.append(filters['reparto'])
        
        if filters.get('banco'):
            base_query += " AND SUBSTRING(a.EAN, 1, 2) = %s"
            params.append(filters['banco'])
        
        if filters.get('fornitore'):
            base_query += " AND m.CCOM = %s"
            params.append(filters['fornitore'])
        
        if filters.get('search'):
            search_term = f"%{filters['search']}%"
            base_query += """ AND (
                a.DESCRART LIKE %s 
                OR a.CODART LIKE %s 
                OR SUBSTRING(a.EAN, 3, 7) LIKE %s
                OR m.DESCRCCOM LIKE %s
            )"""
            params.extend([search_term, search_term, search_term, search_term])
    
    base_query += " ORDER BY a.REP, SUBSTRING(a.EAN, 3, 7)"
    
    with connections['goldreport'].cursor() as cursor:
        cursor.execute(base_query, params)
        columns = [col[0] for col in cursor.description]
        results = []
        
        for row in cursor.fetchall():
            row_dict = dict(zip(columns, row))
            row_dict['ean_formatted'] = str(row_dict.get('ean', '')).zfill(13) if row_dict.get('ean') else ''
            
            # Se manca reparto_nome, usa il codice
            if not row_dict.get('reparto_nome'):
                row_dict['reparto_nome'] = f"Reparto {row_dict.get('rep', '')}"
            
            try:
                row_dict['plu_int'] = int(row_dict.get('plu', 0)) if row_dict.get('plu') else 0
            except (ValueError, TypeError):
                row_dict['plu_int'] = 0
            
            results.append(row_dict)
        
        return results


class RepartoPlUViewSet(viewsets.ViewSet):
    """
    ViewSet per gestione PLU articoli - usa query raw
    """
    permission_classes = [AllowAny]
    
    def list(self, request):
        """
        GET /api/plu/
        Lista articoli con filtri
        """
        filters = {
            'reparto': request.query_params.get('reparto'),
            'banco': request.query_params.get('banco'),
            'fornitore': request.query_params.get('fornitore'),
            'search': request.query_params.get('search'),
        }
        filters = {k: v for k, v in filters.items() if v}
        
        results = execute_plu_query(filters if filters else None)
        return Response(results)
    
    def retrieve(self, request, pk=None):
        """
        GET /api/plu/{codArticolo}/
        Dettaglio singolo articolo
        """
        results = execute_plu_query()
        
        for item in results:
            if str(item.get('codArticolo')) == str(pk):
                return Response(item)
        
        return Response({'detail': 'Articolo non trovato'}, status=404)
    
    @action(detail=False, methods=['get'])
    def reparti(self, request):
        """
        GET /api/plu/reparti/
        Lista reparti con conteggio e nome
        """
        results = execute_plu_query()
        
        reparti_count = {}
        for item in results:
            rep = item.get('rep')
            if rep not in reparti_count:
                reparti_count[rep] = {
                    'reparto': rep,
                    'reparto_nome': item.get('reparto_nome'),
                    'totale_articoli': 0
                }
            reparti_count[rep]['totale_articoli'] += 1
        
        return Response(sorted(reparti_count.values(), key=lambda x: str(x['reparto'] or '')))
    
    @action(detail=False, methods=['get'])
    def banchi(self, request):
        """
        GET /api/plu/banchi/
        Lista banchi con conteggio
        """
        results = execute_plu_query()
        
        banchi_count = {}
        for item in results:
            banco = item.get('bancobilancia')
            if banco and banco not in banchi_count:
                banchi_count[banco] = {
                    'banco': banco,
                    'totale_articoli': 0
                }
            if banco:
                banchi_count[banco]['totale_articoli'] += 1
        
        return Response(sorted(banchi_count.values(), key=lambda x: str(x['banco'] or '')))
    
    @action(detail=False, methods=['get'])
    def fornitori(self, request):
        """
        GET /api/plu/fornitori/
        Lista fornitori
        """
        results = execute_plu_query()
        
        fornitori = {}
        for item in results:
            ccom = item.get('ccom')
            descrccom = item.get('descrccom')
            if ccom and descrccom and ccom not in fornitori:
                fornitori[ccom] = {
                    'codice': ccom,
                    'descrizione': descrccom
                }
        
        return Response(sorted(fornitori.values(), key=lambda x: x['descrizione'] or ''))
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        GET /api/plu/stats/
        Statistiche generali
        """
        results = execute_plu_query()
        
        reparti = set()
        banchi = set()
        fornitori = set()
        per_reparto = {}
        
        for item in results:
            rep = item.get('rep')
            banco = item.get('bancobilancia')
            ccom = item.get('ccom')
            
            if rep:
                reparti.add(rep)
                if rep not in per_reparto:
                    per_reparto[rep] = {
                        'reparto': rep,
                        'reparto_nome': item.get('reparto_nome'),
                        'totale_articoli': 0,
                        'totale_plu': set(),
                        'banchi': set()
                    }
                per_reparto[rep]['totale_articoli'] += 1
                if item.get('plu'):
                    per_reparto[rep]['totale_plu'].add(item.get('plu'))
                if banco:
                    per_reparto[rep]['banchi'].add(banco)
            
            if banco:
                banchi.add(banco)
            if ccom:
                fornitori.add(ccom)
        
        stats_reparti = []
        for rep_data in per_reparto.values():
            stats_reparti.append({
                'reparto': rep_data['reparto'],
                'reparto_nome': rep_data['reparto_nome'],
                'totale_articoli': rep_data['totale_articoli'],
                'totale_plu': len(rep_data['totale_plu']),
                'banchi': list(rep_data['banchi'])
            })
        
        return Response({
            'totale_articoli': len(results),
            'totale_reparti': len(reparti),
            'totale_banchi': len(banchi),
            'totale_fornitori': len(fornitori),
            'per_reparto': sorted(stats_reparti, key=lambda x: str(x['reparto'] or ''))
        })
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        GET /api/plu/export_excel/
        Export Excel
        """
        filters = {
            'reparto': request.query_params.get('reparto'),
            'banco': request.query_params.get('banco'),
            'fornitore': request.query_params.get('fornitore'),
        }
        filters = {k: v for k, v in filters.items() if v}
        
        results = execute_plu_query(filters if filters else None)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PLU Articoli"
        
        header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        headers = [
            'PLU', 'Cod.Articolo', 'Descrizione', 'EAN', 'Barcode',
            'Reparto', 'Banco', 'Cod.Forn', 'Fornitore'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row_num, item in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=item.get('plu'))
            ws.cell(row=row_num, column=2, value=item.get('codArticolo'))
            ws.cell(row=row_num, column=3, value=item.get('descrizione'))
            ws.cell(row=row_num, column=4, value=item.get('ean_formatted'))
            # Barcode - testo per font barcode
            barcode_cell = ws.cell(row=row_num, column=5, value=item.get('ean_formatted'))
            barcode_cell.font = Font(name='Libre Barcode EAN13 Text', size=48)
            ws.cell(row=row_num, column=6, value=item.get('reparto_nome'))
            ws.cell(row=row_num, column=7, value=item.get('bancobilancia'))
            ws.cell(row=row_num, column=8, value=item.get('ccom'))
            ws.cell(row=row_num, column=9, value=item.get('descrccom'))
        
        # Larghezza colonne
        column_widths = [10, 15, 45, 18, 30, 20, 8, 12, 35]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Altezza righe per barcode
        for row in range(2, len(results) + 2):
            ws.row_dimensions[row].height = 35
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'plu_articoli_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response